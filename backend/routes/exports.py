"""
Export module for generating PDF and Excel reports.
Supports: Journal, Ledger, P&L, Balance Sheet exports.
"""
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from typing import Optional
from database import db
from auth import get_current_user
from datetime import datetime, timezone
from routes.bookkeeping import get_indian_fy_dates
import logging
import io

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/exports")


def format_inr(amount: float) -> str:
    """Format amount in Indian number system."""
    is_negative = amount < 0
    abs_amount = abs(amount)
    whole = int(abs_amount)
    decimal = round((abs_amount - whole) * 100)
    
    s = str(whole)
    result = ""
    
    if len(s) <= 3:
        result = s
    else:
        result = s[-3:]
        s = s[:-3]
        while s:
            result = s[-2:] + "," + result
            s = s[:-2]
    
    formatted = f"{result}.{decimal:02d}"
    return f"({formatted})" if is_negative else formatted


async def get_journal_data(user_id: str, start_date: str, end_date: str):
    """Fetch journal entries for export."""
    query = {
        "user_id": user_id,
        "date": {"$gte": start_date, "$lte": end_date},
    }
    entries = await db.journal_entries.find(
        query, {"_id": 0}
    ).sort("date", 1).to_list(5000)
    return entries


async def get_pnl_data(user_id: str, start_date: str, end_date: str):
    """Fetch P&L data for export."""
    pipeline = [
        {"$match": {
            "user_id": user_id,
            "date": {"$gte": start_date, "$lte": end_date},
        }},
        {"$unwind": "$entries"},
        {"$match": {"entries.account_type": "Nominal"}},
        {"$group": {
            "_id": {
                "account_name": "$entries.account_name",
                "account_group": "$entries.account_group",
            },
            "total_debit": {"$sum": "$entries.debit"},
            "total_credit": {"$sum": "$entries.credit"},
        }},
    ]
    results = await db.journal_entries.aggregate(pipeline).to_list(500)
    
    income_items = {}
    expense_items = {}
    
    for r in results:
        name = r["_id"]["account_name"].replace(" A/c", "")
        group = r["_id"]["account_group"]
        amount = round(abs(r["total_credit"] - r["total_debit"]), 2)
        
        if group == "Income":
            income_items[name] = income_items.get(name, 0) + amount
        elif group == "Expense":
            expense_items[name] = expense_items.get(name, 0) + amount
    
    return income_items, expense_items


async def get_balance_sheet_data(user_id: str, as_of_date: str):
    """Fetch balance sheet data for export."""
    pipeline = [
        {"$match": {"user_id": user_id, "date": {"$lte": as_of_date}}},
        {"$unwind": "$entries"},
        {"$group": {
            "_id": {
                "account_name": "$entries.account_name",
                "account_type": "$entries.account_type",
                "account_group": "$entries.account_group",
            },
            "total_debit": {"$sum": "$entries.debit"},
            "total_credit": {"$sum": "$entries.credit"},
        }},
    ]
    return await db.journal_entries.aggregate(pipeline).to_list(500)


# ══════════════════════════════════════════════════════════════
# PDF EXPORTS
# ══════════════════════════════════════════════════════════════

@router.get("/journal/pdf")
async def export_journal_pdf(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user=Depends(get_current_user),
):
    """Export Journal Entries as PDF."""
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    if not start_date or not end_date:
        fy_start, fy_end = get_indian_fy_dates()
        start_date = start_date or fy_start.strftime("%Y-%m-%d")
        end_date = end_date or fy_end.strftime("%Y-%m-%d")
    
    entries = await get_journal_data(user["id"], start_date, end_date)
    
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=15*mm, rightMargin=15*mm, topMargin=20*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    elements = []
    
    # Title
    title_style = ParagraphStyle('JournalTitle', parent=styles['Title'], fontSize=18, spaceAfter=6, textColor=rl_colors.HexColor('#1E3A5F'))
    elements.append(Paragraph("JOURNAL ENTRIES", title_style))
    
    sub_style = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=10, textColor=rl_colors.grey)
    elements.append(Paragraph(f"Period: {start_date} to {end_date}", sub_style))
    elements.append(Paragraph(f"Generated: {datetime.now(timezone.utc).strftime('%d %b %Y, %H:%M UTC')}", sub_style))
    elements.append(Spacer(1, 15))
    
    # Build table data
    header = ["Entry #", "Date", "Type", "Account Name", "Debit (₹)", "Credit (₹)", "Narration"]
    data = [header]
    
    total_debit = 0
    total_credit = 0
    
    for entry in entries:
        entry_num = entry.get("entry_number", "")
        date = entry.get("date", "")
        ref_type = entry.get("reference_type", "")
        narration = entry.get("narration", "")[:40]
        
        for i, line in enumerate(entry.get("entries", [])):
            dr = line.get("debit", 0)
            cr = line.get("credit", 0)
            total_debit += dr
            total_credit += cr
            
            data.append([
                str(entry_num) if i == 0 else "",
                date if i == 0 else "",
                ref_type.capitalize() if i == 0 else "",
                line.get("account_name", ""),
                format_inr(dr) if dr > 0 else "-",
                format_inr(cr) if cr > 0 else "-",
                narration if i == 0 else "",
            ])
    
    # Totals row
    data.append(["", "", "", "TOTAL", format_inr(total_debit), format_inr(total_credit), ""])
    
    col_widths = [50, 65, 70, 150, 80, 80, 180]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (4, 0), (5, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.HexColor('#E5E7EB')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [rl_colors.white, rl_colors.HexColor('#F9FAFB')]),
        ('BACKGROUND', (0, -1), (-1, -1), rl_colors.HexColor('#E8F5E9')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)
    
    # Footer
    elements.append(Spacer(1, 20))
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=rl_colors.grey)
    elements.append(Paragraph(f"Total Entries: {len(entries)} | Generated by Visor Finance", footer_style))
    
    doc.build(elements)
    buf.seek(0)
    
    filename = f"Journal_{start_date}_to_{end_date}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/pnl/pdf")
async def export_pnl_pdf(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user=Depends(get_current_user),
):
    """Export Profit & Loss Statement as PDF."""
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    if not start_date or not end_date:
        fy_start, fy_end = get_indian_fy_dates()
        start_date = start_date or fy_start.strftime("%Y-%m-%d")
        end_date = end_date or fy_end.strftime("%Y-%m-%d")
    
    income_items, expense_items = await get_pnl_data(user["id"], start_date, end_date)
    
    total_income = sum(income_items.values())
    total_expenses = sum(expense_items.values())
    surplus = total_income - total_expenses
    
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=20*mm, rightMargin=20*mm, topMargin=25*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()
    elements = []
    
    # Title
    title_style = ParagraphStyle('PnLTitle', parent=styles['Title'], fontSize=18, spaceAfter=6, textColor=rl_colors.HexColor('#1E3A5F'), alignment=1)
    elements.append(Paragraph("INCOME & EXPENDITURE STATEMENT", title_style))
    
    sub_style = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=10, textColor=rl_colors.grey, alignment=1)
    elements.append(Paragraph(f"For the period {start_date} to {end_date}", sub_style))
    elements.append(Spacer(1, 20))
    
    # Income Section
    income_header = ParagraphStyle('SectionHeader', parent=styles['Heading2'], fontSize=12, textColor=rl_colors.HexColor('#059669'), spaceBefore=10)
    elements.append(Paragraph("I. INCOME", income_header))
    
    income_data = [["Particulars", "Amount (₹)"]]
    for cat, amt in sorted(income_items.items(), key=lambda x: -x[1]):
        income_data.append([cat, format_inr(amt)])
    income_data.append(["Total Income (A)", format_inr(total_income)])
    
    income_table = Table(income_data, colWidths=[120*mm, 50*mm])
    income_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#D1FAE5')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.HexColor('#E5E7EB')),
        ('BACKGROUND', (0, -1), (-1, -1), rl_colors.HexColor('#A7F3D0')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(income_table)
    elements.append(Spacer(1, 15))
    
    # Expense Section
    expense_header = ParagraphStyle('SectionHeader', parent=styles['Heading2'], fontSize=12, textColor=rl_colors.HexColor('#DC2626'), spaceBefore=10)
    elements.append(Paragraph("II. EXPENDITURE", expense_header))
    
    expense_data = [["Particulars", "Amount (₹)"]]
    for cat, amt in sorted(expense_items.items(), key=lambda x: -x[1]):
        expense_data.append([cat, format_inr(amt)])
    expense_data.append(["Total Expenditure (B)", format_inr(total_expenses)])
    
    expense_table = Table(expense_data, colWidths=[120*mm, 50*mm])
    expense_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#FEE2E2')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.HexColor('#E5E7EB')),
        ('BACKGROUND', (0, -1), (-1, -1), rl_colors.HexColor('#FECACA')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(expense_table)
    elements.append(Spacer(1, 20))
    
    # Surplus/Deficit Box
    surplus_label = "SURPLUS (A - B)" if surplus >= 0 else "DEFICIT (B - A)"
    surplus_color = rl_colors.HexColor('#059669') if surplus >= 0 else rl_colors.HexColor('#DC2626')
    surplus_bg = rl_colors.HexColor('#D1FAE5') if surplus >= 0 else rl_colors.HexColor('#FEE2E2')
    
    surplus_data = [[surplus_label, format_inr(abs(surplus))]]
    surplus_table = Table(surplus_data, colWidths=[120*mm, 50*mm])
    surplus_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), surplus_bg),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('TEXTCOLOR', (0, 0), (-1, -1), surplus_color),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('BOX', (0, 0), (-1, -1), 2, surplus_color),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(surplus_table)
    
    # Footer
    elements.append(Spacer(1, 30))
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=rl_colors.grey, alignment=1)
    elements.append(Paragraph(f"Generated by Visor Finance on {datetime.now(timezone.utc).strftime('%d %b %Y, %H:%M UTC')}", footer_style))
    
    doc.build(elements)
    buf.seek(0)
    
    filename = f"PnL_Statement_{start_date}_to_{end_date}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/balance-sheet/pdf")
async def export_balance_sheet_pdf(
    as_of_date: Optional[str] = None,
    user=Depends(get_current_user),
):
    """Export Balance Sheet as PDF."""
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    if not as_of_date:
        as_of_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    results = await get_balance_sheet_data(user["id"], as_of_date)
    
    # Categorize accounts
    assets = []
    liabilities = []
    income_total = 0
    expense_total = 0
    
    for r in results:
        name = r["_id"]["account_name"]
        acc_type = r["_id"]["account_type"]
        acc_group = r["_id"]["account_group"]
        balance = round(r["total_debit"] - r["total_credit"], 2)
        
        if acc_type == "Nominal":
            if acc_group == "Income":
                income_total += abs(balance)
            elif acc_group == "Expense":
                expense_total += abs(balance)
        elif acc_type in ("Real", "Personal"):
            if balance > 0:
                assets.append({"name": name, "amount": balance})
            elif balance < 0:
                liabilities.append({"name": name, "amount": abs(balance)})
    
    total_assets = sum(a["amount"] for a in assets)
    total_liabilities = sum(liab["amount"] for liab in liabilities)
    net_worth = income_total - expense_total + total_assets - total_liabilities
    
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=20*mm, rightMargin=20*mm, topMargin=25*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()
    elements = []
    
    # Title
    title_style = ParagraphStyle('BSTitle', parent=styles['Title'], fontSize=18, spaceAfter=6, textColor=rl_colors.HexColor('#1E3A5F'), alignment=1)
    elements.append(Paragraph("STATEMENT OF FINANCIAL POSITION", title_style))
    
    sub_style = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=10, textColor=rl_colors.grey, alignment=1)
    elements.append(Paragraph(f"As at {as_of_date}", sub_style))
    elements.append(Spacer(1, 20))
    
    # Assets Section
    section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=12, textColor=rl_colors.HexColor('#1E40AF'), spaceBefore=10)
    elements.append(Paragraph("I. ASSETS", section_style))
    
    asset_data = [["Particulars", "Amount (₹)"]]
    for a in sorted(assets, key=lambda x: -x["amount"]):
        asset_data.append([a["name"], format_inr(a["amount"])])
    asset_data.append(["Total Assets", format_inr(total_assets)])
    
    asset_table = Table(asset_data, colWidths=[120*mm, 50*mm])
    asset_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#DBEAFE')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.HexColor('#E5E7EB')),
        ('BACKGROUND', (0, -1), (-1, -1), rl_colors.HexColor('#93C5FD')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(asset_table)
    elements.append(Spacer(1, 15))
    
    # Liabilities Section
    elements.append(Paragraph("II. LIABILITIES", section_style))
    
    liability_data = [["Particulars", "Amount (₹)"]]
    for liab in sorted(liabilities, key=lambda x: -x["amount"]):
        liability_data.append([liab["name"], format_inr(liab["amount"])])
    liability_data.append(["Total Liabilities", format_inr(total_liabilities)])
    
    liability_table = Table(liability_data, colWidths=[120*mm, 50*mm])
    liability_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#FEE2E2')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.HexColor('#E5E7EB')),
        ('BACKGROUND', (0, -1), (-1, -1), rl_colors.HexColor('#FECACA')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(liability_table)
    elements.append(Spacer(1, 15))
    
    # Net Worth Section
    elements.append(Paragraph("III. NET WORTH", section_style))
    
    nw_data = [
        ["Accumulated Surplus/(Deficit)", format_inr(income_total - expense_total)],
        ["Net Worth", format_inr(net_worth)],
    ]
    nw_table = Table(nw_data, colWidths=[120*mm, 50*mm])
    nw_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.HexColor('#E5E7EB')),
        ('BACKGROUND', (0, -1), (-1, -1), rl_colors.HexColor('#D1FAE5')),
        ('TEXTCOLOR', (0, -1), (-1, -1), rl_colors.HexColor('#059669')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(nw_table)
    
    # Balance Check
    elements.append(Spacer(1, 20))
    is_balanced = abs(total_assets - (total_liabilities + net_worth)) < 0.01
    balance_text = "Balance Sheet is BALANCED" if is_balanced else "Balance Sheet has DISCREPANCY"
    balance_color = rl_colors.HexColor('#059669') if is_balanced else rl_colors.HexColor('#DC2626')
    balance_style = ParagraphStyle('Balance', parent=styles['Normal'], fontSize=10, textColor=balance_color, alignment=1)
    elements.append(Paragraph(f"[{balance_text}]", balance_style))
    
    # Footer
    elements.append(Spacer(1, 30))
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=rl_colors.grey, alignment=1)
    elements.append(Paragraph(f"Generated by Visor Finance on {datetime.now(timezone.utc).strftime('%d %b %Y, %H:%M UTC')}", footer_style))
    
    doc.build(elements)
    buf.seek(0)
    
    filename = f"Balance_Sheet_{as_of_date}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/ledger/pdf")
async def export_full_ledger_pdf(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user=Depends(get_current_user),
):
    """Export complete General Ledger as PDF."""
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    if not start_date or not end_date:
        fy_start, fy_end = get_indian_fy_dates()
        start_date = start_date or fy_start.strftime("%Y-%m-%d")
        end_date = end_date or fy_end.strftime("%Y-%m-%d")
    
    # Get all accounts
    pipeline = [
        {"$match": {"user_id": user["id"], "date": {"$gte": start_date, "$lte": end_date}}},
        {"$unwind": "$entries"},
        {"$group": {
            "_id": "$entries.account_name",
            "account_type": {"$first": "$entries.account_type"},
            "account_group": {"$first": "$entries.account_group"},
        }},
        {"$sort": {"_id": 1}},
    ]
    accounts = await db.journal_entries.aggregate(pipeline).to_list(500)
    
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=15*mm, rightMargin=15*mm, topMargin=20*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    elements = []
    
    # Title Page
    title_style = ParagraphStyle('LedgerTitle', parent=styles['Title'], fontSize=20, spaceAfter=6, textColor=rl_colors.HexColor('#1E3A5F'), alignment=1)
    elements.append(Paragraph("GENERAL LEDGER", title_style))
    
    sub_style = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=11, textColor=rl_colors.grey, alignment=1)
    elements.append(Paragraph(f"Period: {start_date} to {end_date}", sub_style))
    elements.append(Paragraph(f"Total Accounts: {len(accounts)}", sub_style))
    elements.append(Spacer(1, 20))
    
    # Table of Contents
    toc_style = ParagraphStyle('TOC', parent=styles['Heading2'], fontSize=12, textColor=rl_colors.HexColor('#374151'))
    elements.append(Paragraph("Index of Accounts", toc_style))
    
    toc_data = [["#", "Account Name", "Type", "Group"]]
    for i, acc in enumerate(accounts, 1):
        toc_data.append([str(i), acc["_id"], acc.get("account_type", ""), acc.get("account_group", "")])
    
    toc_table = Table(toc_data, colWidths=[30, 200, 80, 80])
    toc_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#E5E7EB')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.HexColor('#D1D5DB')),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(toc_table)
    elements.append(PageBreak())
    
    # Individual Ledgers
    for acc in accounts:
        account_name = acc["_id"]
        
        # Get ledger entries
        query = {
            "user_id": user["id"],
            "entries.account_name": account_name,
            "date": {"$gte": start_date, "$lte": end_date},
        }
        journal_docs = await db.journal_entries.find(query, {"_id": 0}).sort("date", 1).to_list(2000)
        
        ledger_entries = []
        running_balance = 0.0
        
        for jdoc in journal_docs:
            for entry in jdoc.get("entries", []):
                if entry["account_name"] == account_name:
                    running_balance += entry["debit"] - entry["credit"]
                    contra = [e["account_name"] for e in jdoc.get("entries", []) if e["account_name"] != account_name]
                    ledger_entries.append({
                        "date": jdoc["date"],
                        "entry_num": jdoc.get("entry_number", ""),
                        "contra": ", ".join(contra)[:30],
                        "narration": jdoc.get("narration", "")[:40],
                        "debit": entry["debit"],
                        "credit": entry["credit"],
                        "balance": round(running_balance, 2),
                    })
        
        total_debit = sum(e["debit"] for e in ledger_entries)
        total_credit = sum(e["credit"] for e in ledger_entries)
        
        # Account Header
        acc_title = ParagraphStyle('AccTitle', parent=styles['Heading2'], fontSize=14, textColor=rl_colors.HexColor('#1E40AF'), spaceBefore=10)
        elements.append(Paragraph(f"Ledger: {account_name}", acc_title))
        
        acc_sub = ParagraphStyle('AccSub', parent=styles['Normal'], fontSize=9, textColor=rl_colors.grey)
        elements.append(Paragraph(f"Type: {acc.get('account_type', '')} | Group: {acc.get('account_group', '')}", acc_sub))
        elements.append(Spacer(1, 8))
        
        # Ledger Table
        header = ["Date", "Entry #", "Contra Account", "Narration", "Debit (₹)", "Credit (₹)", "Balance (₹)"]
        data = [header]
        
        for e in ledger_entries:
            data.append([
                e["date"],
                str(e["entry_num"]),
                e["contra"],
                e["narration"],
                format_inr(e["debit"]) if e["debit"] > 0 else "-",
                format_inr(e["credit"]) if e["credit"] > 0 else "-",
                format_inr(e["balance"]),
            ])
        
        # Totals row
        data.append(["", "", "", "TOTAL", format_inr(total_debit), format_inr(total_credit), format_inr(running_balance)])
        
        col_widths = [55, 45, 90, 150, 70, 70, 75]
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#4338CA')),
            ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ALIGN', (4, 0), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.HexColor('#E5E7EB')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [rl_colors.white, rl_colors.HexColor('#F9FAFB')]),
            ('BACKGROUND', (0, -1), (-1, -1), rl_colors.HexColor('#E8F5E9')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        elements.append(table)
        elements.append(PageBreak())
    
    # Footer on last page
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=rl_colors.grey, alignment=1)
    elements.append(Paragraph(f"End of General Ledger | Generated by Visor Finance on {datetime.now(timezone.utc).strftime('%d %b %Y, %H:%M UTC')}", footer_style))
    
    doc.build(elements)
    buf.seek(0)
    
    filename = f"General_Ledger_{start_date}_to_{end_date}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ══════════════════════════════════════════════════════════════
# EXCEL EXPORTS
# ══════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════
# TAX SUMMARY EXPORT (PDF)
# ══════════════════════════════════════════════════════════════

@router.get("/tax-summary/pdf")
async def export_tax_summary_pdf(
    fy: str = "2025-26",
    user=Depends(get_current_user),
):
    """Export comprehensive Tax Summary Report as PDF — includes all deductions, 
    regime comparison, capital gains, TDS status, and uploaded documents summary.
    Designed to be shared with CAs for ITR filing."""
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from routes.tax import income_tax_calculator, get_tax_summary

    # ── Fetch all tax data ────────────────────────────────────
    tax_calc = await income_tax_calculator(user, fy)
    tax_summary = await get_tax_summary(user)  # noqa: F841

    salary_profile = await db.salary_profiles.find_one({"user_id": user["id"]}, {"_id": 0})
    income_profile = await db.tax_income_profiles.find_one({"user_id": user["id"]}, {"_id": 0})
    tax_docs = await db.tax_documents.find({"user_id": user["id"]}, {"_id": 0}).to_list(20)
    auto_deds = await db.auto_tax_deductions.find(
        {"user_id": user["id"], "fy": fy}, {"_id": 0}
    ).to_list(500)
    user_deds = await db.user_tax_deductions.find(  # noqa: F841
        {"user_id": user["id"]}, {"_id": 0}
    ).to_list(100)

    # Freelancer / Business profiles
    freelancer = await db.freelancer_profiles.find_one({"user_id": user["id"], "fy": fy}, {"_id": 0})
    business = await db.business_profiles.find_one({"user_id": user["id"], "fy": fy}, {"_id": 0})

    # ── Build PDF ────────────────────────────────────────────
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=18*mm, rightMargin=18*mm,
        topMargin=20*mm, bottomMargin=18*mm,
    )
    styles = getSampleStyleSheet()
    elements = []

    # ── Styles ───────────────────────────────────────────────
    NAVY = rl_colors.HexColor('#1E3A5F')
    GREEN = rl_colors.HexColor('#059669')
    LIGHT_BLUE = rl_colors.HexColor('#DBEAFE')
    LIGHT_GREEN = rl_colors.HexColor('#D1FAE5')
    LIGHT_RED = rl_colors.HexColor('#FEE2E2')
    LIGHT_AMBER = rl_colors.HexColor('#FEF3C7')
    GREY_BG = rl_colors.HexColor('#F3F4F6')

    title_s = ParagraphStyle('TaxTitle', parent=styles['Title'], fontSize=20, textColor=NAVY, alignment=1, spaceAfter=4)
    sub_s = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=10, textColor=rl_colors.grey, alignment=1)
    section_s = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=13, textColor=NAVY, spaceBefore=14, spaceAfter=6)
    note_s = ParagraphStyle('Note', parent=styles['Normal'], fontSize=8, textColor=rl_colors.grey)
    body_s = ParagraphStyle('Body', parent=styles['Normal'], fontSize=9)

    def make_table(data, col_widths, header_color=NAVY):
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), header_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (-1, 0), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.HexColor('#E5E7EB')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [rl_colors.white, GREY_BG]),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        return t

    # ══ PAGE 1: HEADER ═══════════════════════════════════════
    elements.append(Paragraph("TAX SUMMARY REPORT", title_s))
    fy_parts = fy.split("-")
    ay = f"{int(fy_parts[0]) + 1}-{int(fy_parts[1]) + 1:02d}"
    elements.append(Paragraph(f"Financial Year: {fy} | Assessment Year: {ay}", sub_s))
    elements.append(Paragraph(f"Generated: {datetime.now(timezone.utc).strftime('%d %b %Y, %H:%M UTC')}", sub_s))
    elements.append(Spacer(1, 8))
    elements.append(HRFlowable(width="100%", thickness=1, color=NAVY))
    elements.append(Spacer(1, 8))

    # ── Taxpayer Profile ─────────────────────────────────────
    if salary_profile or income_profile:
        elements.append(Paragraph("TAXPAYER PROFILE", section_s))
        profile_data = [["Field", "Details"]]
        if salary_profile:
            profile_data.append(["Employer", salary_profile.get("employer_name", "N/A")])
            profile_data.append(["City", f"{salary_profile.get('residence_city', 'N/A')} ({salary_profile.get('city_type', 'N/A')})"])
            gross_m = salary_profile.get("gross_monthly", 0)
            profile_data.append(["Gross Monthly Salary", format_inr(gross_m)])
            profile_data.append(["Annual Gross", f"{format_inr(salary_profile.get('gross_annual', gross_m * 12))}"])
            profile_data.append(["Monthly TDS", f"{format_inr(salary_profile.get('tds_monthly', 0))}"])
        if income_profile:
            types = income_profile.get("income_types", [])
            profile_data.append(["Income Types", ", ".join(t.title() for t in types)])
        if freelancer:
            profile_data.append(["Freelancer (44ADA)", f"Gross: {format_inr(freelancer.get('gross_receipts', 0))}"])
        if business:
            profile_data.append(["Business (44AD)", f"Turnover: {format_inr(business.get('gross_turnover', 0))}"])

        elements.append(make_table(profile_data, [80*mm, 90*mm]))
        elements.append(Spacer(1, 8))

    # ── Income Summary ───────────────────────────────────────
    elements.append(Paragraph("I. INCOME SUMMARY", section_s))
    income = tax_calc["income"]
    income_data = [
        ["Particulars", "Amount (Rs.)"],
        ["Salary / Employment Income", format_inr(income["salary"])],
        ["Other Income", format_inr(income["other"])],
        ["Gross Total Income", format_inr(income["gross_total"])],
    ]
    cg = tax_calc.get("capital_gains", {})
    if cg.get("stcg", 0) > 0 or cg.get("ltcg", 0) > 0:
        income_data.append(["Short-Term Capital Gains (STCG)", format_inr(cg.get("stcg", 0))])
        income_data.append(["Long-Term Capital Gains (LTCG)", format_inr(cg.get("ltcg", 0))])
        income_data.append([f"LTCG Exemption (u/s 112A)", f"({format_inr(cg.get('ltcg_exemption', 125000))})"])
        income_data.append(["Taxable Capital Gains", format_inr(cg.get("stcg", 0) + cg.get("ltcg_taxable", 0))])

    t = make_table(income_data, [110*mm, 60*mm])
    # Highlight gross total row
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 3), (-1, 3), LIGHT_GREEN),
        ('FONTNAME', (0, 3), (-1, 3), 'Helvetica-Bold'),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 8))

    # ── Deductions Breakdown ─────────────────────────────────
    elements.append(Paragraph("II. DEDUCTIONS (Chapter VI-A) — Old Regime", section_s))
    ded_data = [["Section", "Claimed (Rs.)", "Limit (Rs.)", "Eligible (Rs.)"]]
    for d in tax_calc.get("deductions", []):
        section = d.get("section", "")
        amount = d.get("amount", 0) or 0
        limit = d.get("limit", 0) or 0
        capped = d.get("capped_amount", min(amount, limit) if limit > 0 else amount)
        ded_data.append([
            d.get("label", f"Section {section}"),
            format_inr(amount),
            format_inr(limit) if limit > 0 else "No limit",
            format_inr(capped),
        ])

    # Auto-detected deductions not in tax_calc deductions
    auto_sections = {}
    for d in auto_deds:
        sec = d.get("section", "")
        auto_sections[sec] = auto_sections.get(sec, 0) + d.get("amount", 0)

    # Standard deduction
    old_std = tax_calc["old_regime"]["standard_deduction"]
    if old_std > 0:
        ded_data.append(["Standard Deduction (u/s 16(ia))", format_inr(old_std), format_inr(50000), format_inr(old_std)])

    total_old_ded = tax_calc["old_regime"]["total_deductions"]
    ded_data.append(["TOTAL DEDUCTIONS (Old Regime)", "", "", format_inr(total_old_ded)])

    t = make_table(ded_data, [60*mm, 35*mm, 35*mm, 40*mm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, -1), (-1, -1), LIGHT_BLUE),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 8))

    # ── Old vs New Regime Comparison ──────────────────────────
    elements.append(Paragraph("III. TAX COMPUTATION — Old vs New Regime", section_s))
    old = tax_calc["old_regime"]
    new = tax_calc["new_regime"]
    comp = tax_calc["comparison"]

    regime_data = [
        ["Particulars", "Old Regime (Rs.)", "New Regime (Rs.)"],
        ["Gross Total Income", format_inr(income["gross_total"]), format_inr(income["gross_total"])],
        ["Standard Deduction", format_inr(old["standard_deduction"]), format_inr(new["standard_deduction"])],
        ["Chapter VI-A Deductions", format_inr(old["chapter_via_deductions"]), "N/A"],
        ["NPS Deduction (80CCD)", "Incl. in above", format_inr(new.get("nps_deduction", 0))],
        ["Total Deductions", format_inr(old["total_deductions"]), format_inr(new["total_deductions"])],
        ["Taxable Income", format_inr(old["taxable_income"]), format_inr(new["taxable_income"])],
        ["Tax on Income", format_inr(old["tax_on_income"]), format_inr(new["tax_on_income"])],
        ["Rebate u/s 87A", f"({format_inr(old['rebate_87a'])})", f"({format_inr(new['rebate_87a'])})"],
        ["Tax after Rebate", format_inr(old["tax_after_rebate"]), format_inr(new["tax_after_rebate"])],
        ["Surcharge", format_inr(old["surcharge"]), format_inr(new["surcharge"])],
        ["Health & Education Cess (4%)", format_inr(old["cess"]), format_inr(new["cess"])],
        ["Tax on Regular Income", format_inr(old["total_tax_on_income"]), format_inr(new["total_tax_on_income"])],
        ["Capital Gains Tax", format_inr(old["capital_gains_tax"]), format_inr(new["capital_gains_tax"])],
        ["TOTAL TAX LIABILITY", format_inr(old["total_tax"]), format_inr(new["total_tax"])],
    ]

    t = Table(regime_data, colWidths=[75*mm, 47*mm, 47*mm], repeatRows=1)
    # Determine which column to highlight as "better"
    old_better = comp["better_regime"] == "old"
    winner_col = 1 if old_better else 2

    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), NAVY),
        ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.HexColor('#E5E7EB')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [rl_colors.white, GREY_BG]),
        ('BACKGROUND', (0, -1), (-1, -1), LIGHT_GREEN if comp["better_regime"] != "equal" else LIGHT_AMBER),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        # Highlight winner column header
        ('BACKGROUND', (winner_col, -1), (winner_col, -1), rl_colors.HexColor('#A7F3D0')),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 6))

    # Recommendation box
    regime_label = "OLD REGIME" if old_better else "NEW REGIME"
    if comp["better_regime"] != "equal":
        savings = comp["savings"]
        rec_text = f"RECOMMENDATION: {regime_label} saves Rs. {format_inr(savings)} more. Effective rate: {comp['old_effective_rate'] if old_better else comp['new_effective_rate']}%"
        rec_style = ParagraphStyle('Rec', parent=styles['Normal'], fontSize=10, textColor=GREEN, alignment=1)
        elements.append(Paragraph(f"<b>{rec_text}</b>", rec_style))
    elements.append(Spacer(1, 8))

    # ── Slab-wise Breakdown ──────────────────────────────────
    elements.append(Paragraph("IV. SLAB-WISE TAX BREAKDOWN", section_s))
    for regime_name, regime_key in [("Old Regime", "old_regime"), ("New Regime", "new_regime")]:
        slab_data = [["Income Range", "Rate", "Taxable Income (Rs.)", "Tax (Rs.)"]]
        for slab in tax_calc[regime_key].get("slab_breakdown", []):
            slab_data.append([slab["range"], slab["rate"], format_inr(slab["income"]), format_inr(slab["tax"])])
        if len(slab_data) > 1:
            elements.append(Paragraph(f"{regime_name}:", body_s))
            elements.append(make_table(slab_data, [55*mm, 20*mm, 47*mm, 47*mm]))
            elements.append(Spacer(1, 6))

    # ── TDS Summary ──────────────────────────────────────────
    if salary_profile and salary_profile.get("tds_monthly", 0) > 0:
        elements.append(Paragraph("V. TDS & TAX PAYMENT STATUS", section_s))
        monthly_tds = salary_profile.get("tds_monthly", 0)
        annual_tds = monthly_tds * 12
        better_tax = old["total_tax"] if old_better else new["total_tax"]
        due = max(0, better_tax - annual_tds)
        refund = max(0, annual_tds - better_tax)

        tds_data = [
            ["Particulars", "Amount (Rs.)"],
            ["TDS Deducted by Employer (Annual)", format_inr(annual_tds)],
            [f"Tax Liability ({regime_label if comp['better_regime'] != 'equal' else 'New Regime'})", format_inr(better_tax)],
        ]
        if refund > 0:
            tds_data.append(["EXPECTED REFUND", format_inr(refund)])
        elif due > 0:
            tds_data.append(["TAX DUE (Self-Assessment)", format_inr(due)])
        else:
            tds_data.append(["Status", "TDS matches liability"])

        t = make_table(tds_data, [110*mm, 60*mm])
        if refund > 0:
            t.setStyle(TableStyle([('BACKGROUND', (0, -1), (-1, -1), LIGHT_GREEN), ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold')]))
        elif due > 0:
            t.setStyle(TableStyle([('BACKGROUND', (0, -1), (-1, -1), LIGHT_RED), ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold')]))
        elements.append(t)
        elements.append(Spacer(1, 8))

    # ── Uploaded Documents ───────────────────────────────────
    if tax_docs:
        elements.append(Paragraph("VI. UPLOADED TAX DOCUMENTS", section_s))
        doc_data = [["Document Type", "Filename", "Uploaded On"]]
        for td in tax_docs:
            dtype = td.get("document_type", "Unknown").replace("_", " ").title()
            fname = td.get("filename", "N/A")[:40]
            uploaded = td.get("created_at", "")[:10]
            doc_data.append([dtype, fname, uploaded])
        elements.append(make_table(doc_data, [55*mm, 80*mm, 35*mm]))
        elements.append(Spacer(1, 8))

    # ── Notes & Disclaimer ───────────────────────────────────
    elements.append(Spacer(1, 12))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=rl_colors.grey))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph("NOTES:", ParagraphStyle('NoteH', parent=styles['Normal'], fontSize=9, textColor=NAVY)))
    for note in tax_calc.get("notes", []):
        elements.append(Paragraph(f"  - {note}", note_s))

    elements.append(Spacer(1, 10))
    disclaimer_s = ParagraphStyle('Disclaimer', parent=styles['Normal'], fontSize=7, textColor=rl_colors.HexColor('#9CA3AF'), alignment=1)
    elements.append(Paragraph(
        "DISCLAIMER: This is a computer-generated estimate based on available data. "
        "It is not a substitute for professional tax advice. Please consult a Chartered Accountant "
        "before filing your Income Tax Return. Generated by Visor Finance.",
        disclaimer_s,
    ))

    # Footer
    elements.append(Spacer(1, 6))
    footer_s = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=rl_colors.grey, alignment=1)
    elements.append(Paragraph(f"Visor Finance | Tax Summary Report | {datetime.now(timezone.utc).strftime('%d %b %Y')}", footer_s))

    doc.build(elements)
    buf.seek(0)

    filename = f"Visor_Tax_Summary_FY{fy}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ══════════════════════════════════════════════════════════════
# EXCEL EXPORTS (Original)
# ══════════════════════════════════════════════════════════════

@router.get("/journal/excel")
async def export_journal_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user=Depends(get_current_user),
):
    """Export Journal Entries as Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    if not start_date or not end_date:
        fy_start, fy_end = get_indian_fy_dates()
        start_date = start_date or fy_start.strftime("%Y-%m-%d")
        end_date = end_date or fy_end.strftime("%Y-%m-%d")
    
    entries = await get_journal_data(user["id"], start_date, end_date)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Journal Entries"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    total_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:G1')
    ws['A1'] = "JOURNAL ENTRIES"
    ws['A1'].font = Font(bold=True, size=16, color="1E3A5F")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:G2')
    ws['A2'] = f"Period: {start_date} to {end_date}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ["Entry #", "Date", "Type", "Account Name", "Debit (₹)", "Credit (₹)", "Narration"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Data
    row = 5
    total_debit = 0
    total_credit = 0
    
    for entry in entries:
        entry_num = entry.get("entry_number", "")
        date = entry.get("date", "")
        ref_type = entry.get("reference_type", "")
        narration = entry.get("narration", "")
        
        for i, line in enumerate(entry.get("entries", [])):
            dr = line.get("debit", 0)
            cr = line.get("credit", 0)
            total_debit += dr
            total_credit += cr
            
            ws.cell(row=row, column=1, value=entry_num if i == 0 else "").border = thin_border
            ws.cell(row=row, column=2, value=date if i == 0 else "").border = thin_border
            ws.cell(row=row, column=3, value=ref_type.capitalize() if i == 0 else "").border = thin_border
            ws.cell(row=row, column=4, value=line.get("account_name", "")).border = thin_border
            
            debit_cell = ws.cell(row=row, column=5, value=dr if dr > 0 else None)
            debit_cell.number_format = '₹#,##0.00'
            debit_cell.border = thin_border
            
            credit_cell = ws.cell(row=row, column=6, value=cr if cr > 0 else None)
            credit_cell.number_format = '₹#,##0.00'
            credit_cell.border = thin_border
            
            ws.cell(row=row, column=7, value=narration if i == 0 else "").border = thin_border
            
            row += 1
    
    # Totals row
    ws.cell(row=row, column=4, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=4).fill = total_fill
    ws.cell(row=row, column=4).border = thin_border
    
    total_dr_cell = ws.cell(row=row, column=5, value=total_debit)
    total_dr_cell.font = Font(bold=True)
    total_dr_cell.fill = total_fill
    total_dr_cell.number_format = '₹#,##0.00'
    total_dr_cell.border = thin_border
    
    total_cr_cell = ws.cell(row=row, column=6, value=total_credit)
    total_cr_cell.font = Font(bold=True)
    total_cr_cell.fill = total_fill
    total_cr_cell.number_format = '₹#,##0.00'
    total_cr_cell.border = thin_border
    
    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 40
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    filename = f"Journal_{start_date}_to_{end_date}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/pnl/excel")
async def export_pnl_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user=Depends(get_current_user),
):
    """Export P&L Statement as Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    if not start_date or not end_date:
        fy_start, fy_end = get_indian_fy_dates()
        start_date = start_date or fy_start.strftime("%Y-%m-%d")
        end_date = end_date or fy_end.strftime("%Y-%m-%d")
    
    income_items, expense_items = await get_pnl_data(user["id"], start_date, end_date)
    
    total_income = sum(income_items.values())
    total_expenses = sum(expense_items.values())
    surplus = total_income - total_expenses
    
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L Statement"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    income_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    expense_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    subtotal_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:B1')
    ws['A1'] = "INCOME & EXPENDITURE STATEMENT"
    ws['A1'].font = Font(bold=True, size=16, color="1E3A5F")
    
    ws.merge_cells('A2:B2')
    ws['A2'] = f"Period: {start_date} to {end_date}"
    
    # Income Section
    row = 4
    ws.cell(row=row, column=1, value="I. INCOME").font = Font(bold=True, size=12, color="059669")
    row += 1
    
    ws.cell(row=row, column=1, value="Particulars").font = header_font
    ws.cell(row=row, column=1).fill = income_fill
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2, value="Amount (₹)").font = header_font
    ws.cell(row=row, column=2).fill = income_fill
    ws.cell(row=row, column=2).border = thin_border
    row += 1
    
    for cat, amt in sorted(income_items.items(), key=lambda x: -x[1]):
        ws.cell(row=row, column=1, value=cat).border = thin_border
        amt_cell = ws.cell(row=row, column=2, value=amt)
        amt_cell.number_format = '₹#,##0.00'
        amt_cell.border = thin_border
        row += 1
    
    ws.cell(row=row, column=1, value="Total Income (A)").font = Font(bold=True)
    ws.cell(row=row, column=1).fill = subtotal_fill
    ws.cell(row=row, column=1).border = thin_border
    total_inc_cell = ws.cell(row=row, column=2, value=total_income)
    total_inc_cell.font = Font(bold=True)
    total_inc_cell.fill = subtotal_fill
    total_inc_cell.number_format = '₹#,##0.00'
    total_inc_cell.border = thin_border
    row += 2
    
    # Expense Section
    ws.cell(row=row, column=1, value="II. EXPENDITURE").font = Font(bold=True, size=12, color="DC2626")
    row += 1
    
    ws.cell(row=row, column=1, value="Particulars").font = header_font
    ws.cell(row=row, column=1).fill = expense_fill
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2, value="Amount (₹)").font = header_font
    ws.cell(row=row, column=2).fill = expense_fill
    ws.cell(row=row, column=2).border = thin_border
    row += 1
    
    for cat, amt in sorted(expense_items.items(), key=lambda x: -x[1]):
        ws.cell(row=row, column=1, value=cat).border = thin_border
        amt_cell = ws.cell(row=row, column=2, value=amt)
        amt_cell.number_format = '₹#,##0.00'
        amt_cell.border = thin_border
        row += 1
    
    ws.cell(row=row, column=1, value="Total Expenditure (B)").font = Font(bold=True)
    ws.cell(row=row, column=1).fill = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")
    ws.cell(row=row, column=1).border = thin_border
    total_exp_cell = ws.cell(row=row, column=2, value=total_expenses)
    total_exp_cell.font = Font(bold=True)
    total_exp_cell.fill = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")
    total_exp_cell.number_format = '₹#,##0.00'
    total_exp_cell.border = thin_border
    row += 2
    
    # Surplus/Deficit
    surplus_label = "SURPLUS (A - B)" if surplus >= 0 else "DEFICIT (B - A)"
    ws.cell(row=row, column=1, value=surplus_label).font = Font(bold=True, size=12)
    surplus_cell = ws.cell(row=row, column=2, value=abs(surplus))
    surplus_cell.font = Font(bold=True, size=12, color="059669" if surplus >= 0 else "DC2626")
    surplus_cell.number_format = '₹#,##0.00'
    
    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    filename = f"PnL_Statement_{start_date}_to_{end_date}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/balance-sheet/excel")
async def export_balance_sheet_excel(
    as_of_date: Optional[str] = None,
    user=Depends(get_current_user),
):
    """Export Balance Sheet as Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    if not as_of_date:
        as_of_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    results = await get_balance_sheet_data(user["id"], as_of_date)
    
    assets = []
    liabilities = []
    income_total = 0
    expense_total = 0
    
    for r in results:
        name = r["_id"]["account_name"]
        acc_type = r["_id"]["account_type"]
        acc_group = r["_id"]["account_group"]
        balance = round(r["total_debit"] - r["total_credit"], 2)
        
        if acc_type == "Nominal":
            if acc_group == "Income":
                income_total += abs(balance)
            elif acc_group == "Expense":
                expense_total += abs(balance)
        elif acc_type in ("Real", "Personal"):
            if balance > 0:
                assets.append({"name": name, "amount": balance})
            elif balance < 0:
                liabilities.append({"name": name, "amount": abs(balance)})
    
    total_assets = sum(a["amount"] for a in assets)
    total_liabilities = sum(liab["amount"] for liab in liabilities)
    net_worth = income_total - expense_total + total_assets - total_liabilities
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    asset_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    liability_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    subtotal_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:B1')
    ws['A1'] = "STATEMENT OF FINANCIAL POSITION"
    ws['A1'].font = Font(bold=True, size=16, color="1E3A5F")
    
    ws.merge_cells('A2:B2')
    ws['A2'] = f"As at {as_of_date}"
    
    # Assets Section
    row = 4
    ws.cell(row=row, column=1, value="I. ASSETS").font = Font(bold=True, size=12, color="1E40AF")
    row += 1
    
    ws.cell(row=row, column=1, value="Particulars").font = header_font
    ws.cell(row=row, column=1).fill = asset_fill
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2, value="Amount (₹)").font = header_font
    ws.cell(row=row, column=2).fill = asset_fill
    ws.cell(row=row, column=2).border = thin_border
    row += 1
    
    for a in sorted(assets, key=lambda x: -x["amount"]):
        ws.cell(row=row, column=1, value=a["name"]).border = thin_border
        amt_cell = ws.cell(row=row, column=2, value=a["amount"])
        amt_cell.number_format = '₹#,##0.00'
        amt_cell.border = thin_border
        row += 1
    
    ws.cell(row=row, column=1, value="Total Assets").font = Font(bold=True)
    ws.cell(row=row, column=1).fill = subtotal_fill
    ws.cell(row=row, column=1).border = thin_border
    total_asset_cell = ws.cell(row=row, column=2, value=total_assets)
    total_asset_cell.font = Font(bold=True)
    total_asset_cell.fill = subtotal_fill
    total_asset_cell.number_format = '₹#,##0.00'
    total_asset_cell.border = thin_border
    row += 2
    
    # Liabilities Section
    ws.cell(row=row, column=1, value="II. LIABILITIES").font = Font(bold=True, size=12, color="DC2626")
    row += 1
    
    ws.cell(row=row, column=1, value="Particulars").font = header_font
    ws.cell(row=row, column=1).fill = liability_fill
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2, value="Amount (₹)").font = header_font
    ws.cell(row=row, column=2).fill = liability_fill
    ws.cell(row=row, column=2).border = thin_border
    row += 1
    
    for liab in sorted(liabilities, key=lambda x: -x["amount"]):
        ws.cell(row=row, column=1, value=liab["name"]).border = thin_border
        amt_cell = ws.cell(row=row, column=2, value=liab["amount"])
        amt_cell.number_format = '₹#,##0.00'
        amt_cell.border = thin_border
        row += 1
    
    ws.cell(row=row, column=1, value="Total Liabilities").font = Font(bold=True)
    ws.cell(row=row, column=1).fill = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")
    ws.cell(row=row, column=1).border = thin_border
    total_liab_cell = ws.cell(row=row, column=2, value=total_liabilities)
    total_liab_cell.font = Font(bold=True)
    total_liab_cell.fill = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")
    total_liab_cell.number_format = '₹#,##0.00'
    total_liab_cell.border = thin_border
    row += 2
    
    # Net Worth Section
    ws.cell(row=row, column=1, value="III. NET WORTH").font = Font(bold=True, size=12, color="059669")
    row += 1
    
    ws.cell(row=row, column=1, value="Accumulated Surplus/(Deficit)").border = thin_border
    surplus_cell = ws.cell(row=row, column=2, value=income_total - expense_total)
    surplus_cell.number_format = '₹#,##0.00'
    surplus_cell.border = thin_border
    row += 1
    
    ws.cell(row=row, column=1, value="Net Worth").font = Font(bold=True, color="059669")
    ws.cell(row=row, column=1).fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    ws.cell(row=row, column=1).border = thin_border
    nw_cell = ws.cell(row=row, column=2, value=net_worth)
    nw_cell.font = Font(bold=True, color="059669")
    nw_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    nw_cell.number_format = '₹#,##0.00'
    nw_cell.border = thin_border
    
    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    filename = f"Balance_Sheet_{as_of_date}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/ledger/excel")
async def export_full_ledger_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user=Depends(get_current_user),
):
    """Export complete General Ledger as Excel with multiple sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    if not start_date or not end_date:
        fy_start, fy_end = get_indian_fy_dates()
        start_date = start_date or fy_start.strftime("%Y-%m-%d")
        end_date = end_date or fy_end.strftime("%Y-%m-%d")
    
    # Get all accounts
    pipeline = [
        {"$match": {"user_id": user["id"], "date": {"$gte": start_date, "$lte": end_date}}},
        {"$unwind": "$entries"},
        {"$group": {
            "_id": "$entries.account_name",
            "account_type": {"$first": "$entries.account_type"},
            "account_group": {"$first": "$entries.account_group"},
        }},
        {"$sort": {"_id": 1}},
    ]
    accounts = await db.journal_entries.aggregate(pipeline).to_list(500)
    
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")
    total_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Index Sheet
    ws_index = wb.active
    ws_index.title = "Index"
    ws_index['A1'] = "GENERAL LEDGER"
    ws_index['A1'].font = Font(bold=True, size=16, color="1E3A5F")
    ws_index['A2'] = f"Period: {start_date} to {end_date}"
    
    ws_index['A4'] = "#"
    ws_index['B4'] = "Account Name"
    ws_index['C4'] = "Type"
    ws_index['D4'] = "Group"
    for cell in ['A4', 'B4', 'C4', 'D4']:
        ws_index[cell].font = header_font
        ws_index[cell].fill = header_fill
    
    for i, acc in enumerate(accounts, 1):
        ws_index[f'A{4+i}'] = i
        ws_index[f'B{4+i}'] = acc["_id"]
        ws_index[f'C{4+i}'] = acc.get("account_type", "")
        ws_index[f'D{4+i}'] = acc.get("account_group", "")
    
    ws_index.column_dimensions['A'].width = 5
    ws_index.column_dimensions['B'].width = 40
    ws_index.column_dimensions['C'].width = 15
    ws_index.column_dimensions['D'].width = 15
    
    # Individual Ledger Sheets
    for acc in accounts[:30]:  # Limit to 30 sheets to avoid Excel limits
        account_name = acc["_id"]
        safe_name = account_name.replace("/", "_").replace(":", "_").replace("*", "_")[:28]
        
        ws = wb.create_sheet(title=safe_name)
        
        # Get ledger entries
        query = {
            "user_id": user["id"],
            "entries.account_name": account_name,
            "date": {"$gte": start_date, "$lte": end_date},
        }
        journal_docs = await db.journal_entries.find(query, {"_id": 0}).sort("date", 1).to_list(2000)
        
        ledger_entries = []
        running_balance = 0.0
        
        for jdoc in journal_docs:
            for entry in jdoc.get("entries", []):
                if entry["account_name"] == account_name:
                    running_balance += entry["debit"] - entry["credit"]
                    contra = [e["account_name"] for e in jdoc.get("entries", []) if e["account_name"] != account_name]
                    ledger_entries.append({
                        "date": jdoc["date"],
                        "entry_num": jdoc.get("entry_number", ""),
                        "contra": ", ".join(contra),
                        "narration": jdoc.get("narration", ""),
                        "debit": entry["debit"],
                        "credit": entry["credit"],
                        "balance": round(running_balance, 2),
                    })
        
        total_debit = sum(e["debit"] for e in ledger_entries)
        total_credit = sum(e["credit"] for e in ledger_entries)
        
        # Account Header
        ws['A1'] = f"Ledger: {account_name}"
        ws['A1'].font = Font(bold=True, size=14, color="1E40AF")
        ws['A2'] = f"Type: {acc.get('account_type', '')} | Group: {acc.get('account_group', '')}"
        
        # Table headers
        headers = ["Date", "Entry #", "Contra Account", "Narration", "Debit (₹)", "Credit (₹)", "Balance (₹)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        # Data rows
        row = 5
        for e in ledger_entries:
            ws.cell(row=row, column=1, value=e["date"]).border = thin_border
            ws.cell(row=row, column=2, value=e["entry_num"]).border = thin_border
            ws.cell(row=row, column=3, value=e["contra"][:40]).border = thin_border
            ws.cell(row=row, column=4, value=e["narration"][:50]).border = thin_border
            
            dr_cell = ws.cell(row=row, column=5, value=e["debit"] if e["debit"] > 0 else None)
            dr_cell.number_format = '₹#,##0.00'
            dr_cell.border = thin_border
            
            cr_cell = ws.cell(row=row, column=6, value=e["credit"] if e["credit"] > 0 else None)
            cr_cell.number_format = '₹#,##0.00'
            cr_cell.border = thin_border
            
            bal_cell = ws.cell(row=row, column=7, value=e["balance"])
            bal_cell.number_format = '₹#,##0.00'
            bal_cell.border = thin_border
            
            row += 1
        
        # Totals row
        ws.cell(row=row, column=4, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=4).fill = total_fill
        ws.cell(row=row, column=4).border = thin_border
        
        ws.cell(row=row, column=5, value=total_debit).font = Font(bold=True)
        ws.cell(row=row, column=5).fill = total_fill
        ws.cell(row=row, column=5).number_format = '₹#,##0.00'
        ws.cell(row=row, column=5).border = thin_border
        
        ws.cell(row=row, column=6, value=total_credit).font = Font(bold=True)
        ws.cell(row=row, column=6).fill = total_fill
        ws.cell(row=row, column=6).number_format = '₹#,##0.00'
        ws.cell(row=row, column=6).border = thin_border
        
        ws.cell(row=row, column=7, value=running_balance).font = Font(bold=True)
        ws.cell(row=row, column=7).fill = total_fill
        ws.cell(row=row, column=7).number_format = '₹#,##0.00'
        ws.cell(row=row, column=7).border = thin_border
        
        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    filename = f"General_Ledger_{start_date}_to_{end_date}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
