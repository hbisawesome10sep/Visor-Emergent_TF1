from fastapi import FastAPI, APIRouter, HTTPException, Depends, Header
from fastapi.responses import RedirectResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
from encryption import generate_user_dek, encrypt_field, decrypt_field, encrypt_sensitive_fields, decrypt_sensitive_fields

# Sensitive fields that need encryption
USER_SENSITIVE_FIELDS = ["pan", "aadhaar"]
LOAN_SENSITIVE_FIELDS = ["account_number"]

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_SECRET = os.environ['JWT_SECRET']
EMERGENT_LLM_KEY = os.environ['EMERGENT_LLM_KEY']
ALPHA_VANTAGE_KEY = os.environ.get('ALPHA_VANTAGE_KEY', '')

app = FastAPI()
api_router = APIRouter(prefix="/api")

# ── Logging ──
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ══════════════════════════════════════
#  MODELS
# ══════════════════════════════════════

class UserCreate(BaseModel):
    email: str
    password: str
    full_name: str
    dob: str
    pan: str
    aadhaar: str

class UserLogin(BaseModel):
    email: str
    password: str

class UserResponse(BaseModel):
    id: str
    email: str
    full_name: str
    dob: str
    pan: str
    aadhaar_last4: str
    created_at: str

class TransactionCreate(BaseModel):
    type: str
    amount: float
    category: str
    description: str = ""
    date: str
    is_recurring: bool = False
    recurring_frequency: Optional[str] = None
    is_split: bool = False
    split_count: int = 1
    notes: Optional[str] = None
    buy_sell: Optional[str] = None  # "buy" or "sell" for investment transactions
    units: Optional[float] = None   # Number of units bought/sold
    price_per_unit: Optional[float] = None  # Price per unit

class TransactionResponse(BaseModel):
    id: str
    user_id: str
    type: str
    amount: float
    category: str
    description: str
    date: str
    is_recurring: bool = False
    recurring_frequency: Optional[str] = None
    is_split: bool = False
    split_count: int = 1
    notes: Optional[str] = None
    buy_sell: Optional[str] = None
    units: Optional[float] = None
    price_per_unit: Optional[float] = None
    created_at: str

class GoalCreate(BaseModel):
    title: str
    target_amount: float
    current_amount: float = 0
    deadline: str
    category: str

class GoalUpdate(BaseModel):
    title: Optional[str] = None
    target_amount: Optional[float] = None
    current_amount: Optional[float] = None
    deadline: Optional[str] = None
    category: Optional[str] = None

class GoalResponse(BaseModel):
    id: str
    user_id: str
    title: str
    target_amount: float
    current_amount: float
    deadline: str
    category: str
    created_at: str

class AIMessageCreate(BaseModel):
    message: str

class ChatMessage(BaseModel):
    id: str
    user_id: str
    role: str
    content: str
    created_at: str

# ══════════════════════════════════════
#  BOOKKEEPING MODELS
# ══════════════════════════════════════

class FixedAssetCreate(BaseModel):
    name: str
    category: str  # Property, Vehicle, Electronics, Furniture, Other
    purchase_date: str
    purchase_value: float
    current_value: float
    depreciation_rate: float = 10.0  # Annual depreciation %
    notes: Optional[str] = None

class FixedAssetUpdate(BaseModel):
    name: Optional[str] = None
    category: Optional[str] = None
    purchase_date: Optional[str] = None
    purchase_value: Optional[float] = None
    current_value: Optional[float] = None
    depreciation_rate: Optional[float] = None
    notes: Optional[str] = None

class FixedAssetResponse(BaseModel):
    id: str
    user_id: str
    name: str
    category: str
    purchase_date: str
    purchase_value: float
    current_value: float
    depreciation_rate: float
    accumulated_depreciation: float
    notes: Optional[str]
    created_at: str

class AccountCreate(BaseModel):
    name: str
    account_type: str  # Assets, Liabilities, Income, Expenses
    account_group: str  # Sub-group like "Cash & Bank", "Investments", etc.
    opening_balance: float = 0
    notes: Optional[str] = None

class AccountResponse(BaseModel):
    id: str
    user_id: str
    name: str
    account_type: str
    account_group: str
    opening_balance: float
    current_balance: float
    notes: Optional[str]
    created_at: str

# ══════════════════════════════════════
#  LOAN/LIABILITY MODELS
# ══════════════════════════════════════

class LoanCreate(BaseModel):
    name: str
    loan_type: str  # Home Loan, Car Loan, Personal Loan, Education Loan, Credit Card, Other
    principal_amount: float
    interest_rate: float  # Annual interest rate %
    tenure_months: int
    start_date: str
    emi_amount: Optional[float] = None  # If not provided, will be calculated
    lender: Optional[str] = None
    account_number: Optional[str] = None
    notes: Optional[str] = None

class LoanUpdate(BaseModel):
    name: Optional[str] = None
    loan_type: Optional[str] = None
    principal_amount: Optional[float] = None
    interest_rate: Optional[float] = None
    tenure_months: Optional[int] = None
    start_date: Optional[str] = None
    emi_amount: Optional[float] = None
    lender: Optional[str] = None
    account_number: Optional[str] = None
    notes: Optional[str] = None

class LoanResponse(BaseModel):
    id: str
    user_id: str
    name: str
    loan_type: str
    principal_amount: float
    interest_rate: float
    tenure_months: int
    start_date: str
    emi_amount: float
    lender: Optional[str]
    account_number: Optional[str]
    notes: Optional[str]
    outstanding_principal: float
    total_interest_paid: float
    total_principal_paid: float
    remaining_emis: int
    created_at: str

class EMIScheduleItem(BaseModel):
    month: int
    date: str
    opening_balance: float
    emi: float
    principal: float
    interest: float
    closing_balance: float
    status: str  # paid, upcoming, overdue

# ══════════════════════════════════════
#  AUTH HELPERS
# ══════════════════════════════════════

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_token(user_id: str, email: str) -> str:
    payload = {
        "user_id": user_id,
        "email": email,
        "exp": datetime.now(timezone.utc).timestamp() + 86400 * 7
    }
    return jwt.encode(payload, JWT_SECRET, algorithm="HS256")

async def get_current_user(authorization: str = Header(None)):
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Not authenticated")
    token = authorization.split(" ")[1]
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
        user = await db.users.find_one({"id": payload["user_id"]}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# ══════════════════════════════════════
#  AUTH ENDPOINTS
# ══════════════════════════════════════

@api_router.post("/auth/register")
async def register(user_data: UserCreate):
    existing = await db.users.find_one({"email": user_data.email.lower()}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    if len(user_data.pan) != 10:
        raise HTTPException(status_code=400, detail="PAN must be 10 characters")
    if len(user_data.aadhaar.replace(" ", "").replace("-", "")) != 12:
        raise HTTPException(status_code=400, detail="Aadhaar must be 12 digits")

    user_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    aadhaar_clean = user_data.aadhaar.replace(" ", "").replace("-", "")
    
    # Generate user-specific encryption key
    user_dek = generate_user_dek()
    
    user_doc = {
        "id": user_id,
        "email": user_data.email.lower(),
        "password": hash_password(user_data.password),
        "full_name": user_data.full_name,
        "dob": user_data.dob,
        "pan": encrypt_field(user_data.pan.upper(), user_dek),
        "aadhaar": encrypt_field(aadhaar_clean, user_dek),
        "encryption_key": user_dek,
        "created_at": now,
    }
    await db.users.insert_one(user_doc)
    
    token = create_token(user_id, user_data.email.lower())
    return {
        "token": token,
        "user": {
            "id": user_id,
            "email": user_data.email.lower(),
            "full_name": user_data.full_name,
            "dob": user_data.dob,
            "pan": user_data.pan.upper(),
            "aadhaar_last4": aadhaar_clean[-4:],
            "created_at": now,
        }
    }

@api_router.post("/auth/login")
async def login(credentials: UserLogin):
    user = await db.users.find_one({"email": credentials.email.lower()}, {"_id": 0})
    if not user or not verify_password(credentials.password, user["password"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    token = create_token(user["id"], user["email"])
    
    # Decrypt sensitive fields if encrypted
    dek = user.get("encryption_key", "")
    pan = user.get("pan", "")
    aadhaar = user.get("aadhaar", "")
    if dek and pan.startswith("ENC:"):
        pan = decrypt_field(pan, dek)
    if dek and aadhaar.startswith("ENC:"):
        aadhaar = decrypt_field(aadhaar, dek)
    
    return {
        "token": token,
        "user": {
            "id": user["id"],
            "email": user["email"],
            "full_name": user["full_name"],
            "dob": user.get("dob", ""),
            "pan": pan,
            "aadhaar_last4": aadhaar[-4:] if len(aadhaar) >= 4 else "",
            "created_at": user.get("created_at", ""),
        }
    }

@api_router.get("/auth/profile")
async def get_profile(user=Depends(get_current_user)):
    dek = user.get("encryption_key", "")
    pan = user.get("pan", "")
    aadhaar = user.get("aadhaar", "")
    if dek and pan.startswith("ENC:"):
        pan = decrypt_field(pan, dek)
    if dek and aadhaar.startswith("ENC:"):
        aadhaar = decrypt_field(aadhaar, dek)
    return {
        "id": user["id"],
        "email": user["email"],
        "full_name": user["full_name"],
        "dob": user.get("dob", ""),
        "pan": pan,
        "aadhaar_last4": aadhaar[-4:] if len(aadhaar) >= 4 else "",
        "created_at": user.get("created_at", ""),
        "is_encrypted": bool(dek),
    }

# ══════════════════════════════════════
#  TRANSACTION ENDPOINTS
# ══════════════════════════════════════

@api_router.get("/transactions", response_model=List[TransactionResponse])
async def get_transactions(
    type: Optional[str] = None,
    category: Optional[str] = None,
    search: Optional[str] = None,
    user=Depends(get_current_user)
):
    query = {"user_id": user["id"]}
    if type:
        query["type"] = type
    if category:
        query["category"] = category
    if search:
        query["$or"] = [
            {"description": {"$regex": search, "$options": "i"}},
            {"category": {"$regex": search, "$options": "i"}},
            {"notes": {"$regex": search, "$options": "i"}},
        ]
    
    txns = await db.transactions.find(query, {"_id": 0}).sort("date", -1).to_list(500)
    return txns

@api_router.post("/transactions", response_model=TransactionResponse)
async def create_transaction(txn: TransactionCreate, user=Depends(get_current_user)):
    txn_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    txn_doc = {
        "id": txn_id,
        "user_id": user["id"],
        "type": txn.type,
        "amount": txn.amount,
        "category": txn.category,
        "description": txn.description,
        "date": txn.date,
        "is_recurring": txn.is_recurring,
        "recurring_frequency": txn.recurring_frequency,
        "is_split": txn.is_split,
        "split_count": txn.split_count,
        "notes": txn.notes,
        "buy_sell": txn.buy_sell,
        "units": txn.units,
        "price_per_unit": txn.price_per_unit,
        "created_at": now,
    }
    await db.transactions.insert_one(txn_doc)
    return TransactionResponse(**txn_doc)

@api_router.delete("/transactions/{txn_id}")
async def delete_transaction(txn_id: str, user=Depends(get_current_user)):
    result = await db.transactions.delete_one({"id": txn_id, "user_id": user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Transaction not found")
    return {"message": "Transaction deleted"}

@api_router.put("/transactions/{txn_id}", response_model=TransactionResponse)
async def update_transaction(txn_id: str, txn: TransactionCreate, user=Depends(get_current_user)):
    existing = await db.transactions.find_one({"id": txn_id, "user_id": user["id"]}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Transaction not found")
    update_data = {
        "type": txn.type,
        "amount": txn.amount,
        "category": txn.category,
        "description": txn.description,
        "date": txn.date,
        "is_recurring": txn.is_recurring,
        "recurring_frequency": txn.recurring_frequency,
        "is_split": txn.is_split,
        "split_count": txn.split_count,
        "notes": txn.notes,
        "buy_sell": txn.buy_sell,
        "units": txn.units,
        "price_per_unit": txn.price_per_unit,
    }
    await db.transactions.update_one({"id": txn_id, "user_id": user["id"]}, {"$set": update_data})
    updated = await db.transactions.find_one({"id": txn_id}, {"_id": 0})
    return TransactionResponse(**updated)

# ══════════════════════════════════════
#  GOAL ENDPOINTS
# ══════════════════════════════════════

@api_router.get("/goals", response_model=List[GoalResponse])
async def get_goals(user=Depends(get_current_user)):
    goals = await db.goals.find({"user_id": user["id"]}, {"_id": 0}).to_list(100)
    return goals

@api_router.post("/goals", response_model=GoalResponse)
async def create_goal(goal: GoalCreate, user=Depends(get_current_user)):
    goal_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    goal_doc = {
        "id": goal_id,
        "user_id": user["id"],
        "title": goal.title,
        "target_amount": goal.target_amount,
        "current_amount": goal.current_amount,
        "deadline": goal.deadline,
        "category": goal.category,
        "created_at": now,
    }
    await db.goals.insert_one(goal_doc)
    return GoalResponse(**goal_doc)

@api_router.put("/goals/{goal_id}", response_model=GoalResponse)
async def update_goal(goal_id: str, goal_update: GoalUpdate, user=Depends(get_current_user)):
    existing = await db.goals.find_one({"id": goal_id, "user_id": user["id"]}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Goal not found")
    
    update_data = {k: v for k, v in goal_update.dict().items() if v is not None}
    if update_data:
        await db.goals.update_one({"id": goal_id}, {"$set": update_data})
    
    updated = await db.goals.find_one({"id": goal_id}, {"_id": 0})
    return GoalResponse(**updated)

@api_router.delete("/goals/{goal_id}")
async def delete_goal(goal_id: str, user=Depends(get_current_user)):
    result = await db.goals.delete_one({"id": goal_id, "user_id": user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Goal not found")
    return {"message": "Goal deleted"}

# ══════════════════════════════════════
#  DASHBOARD & HEALTH SCORE
# ══════════════════════════════════════

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user=Depends(get_current_user),
):
    user_id = user["id"]
    
    # Build query filter
    query = {"user_id": user_id}
    if start_date and end_date:
        query["date"] = {"$gte": start_date, "$lte": end_date}
    
    txns = await db.transactions.find(query, {"_id": 0}).to_list(1000)
    goals = await db.goals.find({"user_id": user_id}, {"_id": 0}).to_list(100)
    
    total_income = sum(t["amount"] for t in txns if t["type"] == "income")
    total_expenses = sum(t["amount"] for t in txns if t["type"] == "expense")
    total_investments = sum(t["amount"] for t in txns if t["type"] == "investment")
    net_balance = total_income - total_expenses - total_investments
    
    # Category breakdown for expenses
    category_breakdown = {}
    for t in txns:
        if t["type"] == "expense":
            cat = t["category"]
            category_breakdown[cat] = category_breakdown.get(cat, 0) + t["amount"]
    
    # Recent transactions (last 5)
    recent = sorted(txns, key=lambda x: x.get("date", ""), reverse=True)[:5]
    
    # Monthly summary (current month approximation)
    now = datetime.now(timezone.utc)
    current_month = now.strftime("%Y-%m")
    monthly_income = sum(t["amount"] for t in txns if t["type"] == "income" and t.get("date", "").startswith(current_month))
    monthly_expenses = sum(t["amount"] for t in txns if t["type"] == "expense" and t.get("date", "").startswith(current_month))
    monthly_investments = sum(t["amount"] for t in txns if t["type"] == "investment" and t.get("date", "").startswith(current_month))
    
    # Goal progress
    total_goal_target = sum(g["target_amount"] for g in goals) if goals else 0
    total_goal_current = sum(g["current_amount"] for g in goals) if goals else 0
    goal_progress = (total_goal_current / total_goal_target * 100) if total_goal_target > 0 else 0
    
    # Savings calculation
    savings = total_income - total_expenses - total_investments
    savings_rate = (savings / total_income * 100) if total_income > 0 else 0
    expense_ratio = (total_expenses / total_income * 100) if total_income > 0 else 0
    investment_ratio = (total_investments / total_income * 100) if total_income > 0 else 0

    # Monthly savings
    monthly_savings = monthly_income - monthly_expenses - monthly_investments

    # Budget tracking per category (as % of income)
    budget_items = []
    for cat, amount in sorted(category_breakdown.items(), key=lambda x: -x[1]):
        pct = (amount / total_income * 100) if total_income > 0 else 0
        budget_items.append({"category": cat, "amount": amount, "percentage": round(pct, 1)})

    # Investment breakdown
    invest_breakdown = {}
    for t in txns:
        if t["type"] == "investment":
            cat = t["category"]
            invest_breakdown[cat] = invest_breakdown.get(cat, 0) + t["amount"]

    # Get user's account creation date for date range limits
    user_created_at = user.get("created_at", now.isoformat())

    # ── Compute Health Score based on filtered date range ──
    hs_income = sum(t["amount"] for t in txns if t["type"] == "income") or 1
    hs_expenses = sum(t["amount"] for t in txns if t["type"] == "expense")
    hs_investments = sum(t["amount"] for t in txns if t["type"] == "investment")
    
    hs_savings_rate = max(0, (hs_income - hs_expenses) / hs_income * 100)
    hs_investment_rate = (hs_investments / hs_income * 100)
    hs_expense_ratio = (hs_expenses / hs_income * 100)
    
    hs_goal_target = sum(g["target_amount"] for g in goals) if goals else 1
    hs_goal_current = sum(g["current_amount"] for g in goals) if goals else 0
    hs_goal_score = min(100, (hs_goal_current / hs_goal_target * 100))
    
    hs_savings_score = min(100, hs_savings_rate * 2.5)
    hs_invest_score = min(100, hs_investment_rate * 5)
    hs_expense_score = max(0, 100 - hs_expense_ratio)
    
    hs_overall = (hs_savings_score * 0.3 + hs_invest_score * 0.2 + hs_expense_score * 0.3 + hs_goal_score * 0.2)
    hs_overall = min(100, max(0, hs_overall))
    
    if hs_overall >= 80:
        hs_grade = "Excellent"
    elif hs_overall >= 65:
        hs_grade = "Good"
    elif hs_overall >= 45:
        hs_grade = "Fair"
    elif hs_overall >= 25:
        hs_grade = "Needs Work"
    else:
        hs_grade = "Critical"

    # ── Trend Analysis Data (for the selected date range) ──
    # Group transactions by week/period for trend visualization
    trend_data = []
    trend_insights = []
    
    if txns:
        # Sort by date
        sorted_txns = sorted(txns, key=lambda x: x.get("date", ""))
        
        # Group by week
        weekly_data = {}
        for t in sorted_txns:
            date_str = t.get("date", "")
            if not date_str:
                continue
            try:
                dt = datetime.strptime(date_str, "%Y-%m-%d")
                week_start = dt - timedelta(days=dt.weekday())
                week_key = week_start.strftime("%b %d")
            except:
                continue
            
            if week_key not in weekly_data:
                weekly_data[week_key] = {"income": 0, "expenses": 0, "investments": 0}
            
            if t["type"] == "income":
                weekly_data[week_key]["income"] += t["amount"]
            elif t["type"] == "expense":
                weekly_data[week_key]["expenses"] += t["amount"]
            elif t["type"] == "investment":
                weekly_data[week_key]["investments"] += t["amount"]
        
        # Convert to list for frontend
        for week_label, data in weekly_data.items():
            trend_data.append({
                "label": week_label,
                "income": round(data["income"], 2),
                "expenses": round(data["expenses"], 2),
                "investments": round(data["investments"], 2),
            })
        
        # Generate insights based on trend
        if len(weekly_data) >= 2:
            weeks = list(weekly_data.keys())
            last_week = weekly_data.get(weeks[-1], {})
            prev_week = weekly_data.get(weeks[-2], {})
            
            # Expense trend
            exp_change = last_week.get("expenses", 0) - prev_week.get("expenses", 0)
            if exp_change > 0:
                exp_pct = (exp_change / max(prev_week.get("expenses", 1), 1)) * 100
                trend_insights.append({
                    "type": "warning",
                    "icon": "trending-up",
                    "title": "Expenses Increasing",
                    "message": f"Your spending increased by ₹{exp_change:,.0f} ({exp_pct:.1f}%) from last week",
                })
            elif exp_change < 0:
                exp_pct = abs(exp_change / max(prev_week.get("expenses", 1), 1)) * 100
                trend_insights.append({
                    "type": "success",
                    "icon": "trending-down",
                    "title": "Expenses Decreasing",
                    "message": f"Great! You saved ₹{abs(exp_change):,.0f} ({exp_pct:.1f}%) compared to last week",
                })
            
            # Income trend
            inc_change = last_week.get("income", 0) - prev_week.get("income", 0)
            if inc_change > 0:
                inc_pct = (inc_change / max(prev_week.get("income", 1), 1)) * 100
                trend_insights.append({
                    "type": "success",
                    "icon": "cash-plus",
                    "title": "Income Growing",
                    "message": f"Income increased by ₹{inc_change:,.0f} ({inc_pct:.1f}%)",
                })
            
            # Investment trend
            inv_change = last_week.get("investments", 0) - prev_week.get("investments", 0)
            if inv_change > 0:
                trend_insights.append({
                    "type": "success",
                    "icon": "chart-line",
                    "title": "Investing More",
                    "message": f"You invested ₹{last_week.get('investments', 0):,.0f} this week",
                })
        
        # Top spending category
        if category_breakdown:
            top_cat = max(category_breakdown.items(), key=lambda x: x[1])
            top_pct = (top_cat[1] / total_expenses * 100) if total_expenses > 0 else 0
            trend_insights.append({
                "type": "info",
                "icon": "tag",
                "title": f"Top Spending: {top_cat[0]}",
                "message": f"₹{top_cat[1]:,.0f} ({top_pct:.1f}% of total expenses)",
            })
        
        # Savings insight
        if savings > 0:
            trend_insights.append({
                "type": "success",
                "icon": "piggy-bank",
                "title": "Savings Summary",
                "message": f"You saved ₹{savings:,.0f} ({savings_rate:.1f}% of income) in this period",
            })
        elif savings < 0:
            trend_insights.append({
                "type": "warning",
                "icon": "alert",
                "title": "Spending Exceeds Income",
                "message": f"You're ₹{abs(savings):,.0f} over budget. Review your expenses.",
            })

    return {
        "total_income": total_income,
        "total_expenses": total_expenses,
        "total_investments": total_investments,
        "net_balance": net_balance,
        "savings": savings,
        "savings_rate": round(savings_rate, 1),
        "expense_ratio": round(expense_ratio, 1),
        "investment_ratio": round(investment_ratio, 1),
        "category_breakdown": category_breakdown,
        "budget_items": budget_items,
        "invest_breakdown": invest_breakdown,
        "recent_transactions": recent,
        "monthly_income": monthly_income,
        "monthly_expenses": monthly_expenses,
        "monthly_investments": monthly_investments,
        "monthly_savings": monthly_savings,
        "goal_count": len(goals),
        "goal_progress": round(goal_progress, 1),
        "transaction_count": len(txns),
        "user_created_at": user_created_at,
        "date_range": {
            "start": start_date,
            "end": end_date,
        } if start_date and end_date else None,
        "trend_data": trend_data,
        "trend_insights": trend_insights,
        "health_score": {
            "overall": round(hs_overall, 1),
            "grade": hs_grade,
            "breakdown": {
                "savings": round(hs_savings_score, 1),
                "investments": round(hs_invest_score, 1),
                "spending": round(hs_expense_score, 1),
                "goals": round(hs_goal_score, 1),
            },
            "metrics": {
                "savings_rate": round(hs_savings_rate, 1),
                "investment_rate": round(hs_investment_rate, 1),
                "expense_ratio": round(hs_expense_ratio, 1),
                "goal_progress": round(hs_goal_score, 1),
            },
        },
    }

@api_router.get("/health-score")
async def get_health_score(user=Depends(get_current_user)):
    user_id = user["id"]
    txns = await db.transactions.find({"user_id": user_id}, {"_id": 0}).to_list(1000)
    goals = await db.goals.find({"user_id": user_id}, {"_id": 0}).to_list(100)
    
    total_income = sum(t["amount"] for t in txns if t["type"] == "income") or 1
    total_expenses = sum(t["amount"] for t in txns if t["type"] == "expense")
    total_investments = sum(t["amount"] for t in txns if t["type"] == "investment")
    
    savings_rate = max(0, (total_income - total_expenses) / total_income * 100)
    investment_rate = (total_investments / total_income * 100)
    expense_ratio = (total_expenses / total_income * 100)
    
    # Goal progress
    total_goal_target = sum(g["target_amount"] for g in goals) if goals else 1
    total_goal_current = sum(g["current_amount"] for g in goals) if goals else 0
    goal_score = min(100, (total_goal_current / total_goal_target * 100))
    
    # Calculate overall score (weighted)
    savings_score = min(100, savings_rate * 2.5)  # 40% savings = 100
    invest_score = min(100, investment_rate * 5)   # 20% investment = 100
    expense_score = max(0, 100 - expense_ratio)     # Lower expense ratio = higher score
    
    overall = (savings_score * 0.3 + invest_score * 0.2 + expense_score * 0.3 + goal_score * 0.2)
    overall = min(100, max(0, overall))
    
    if overall >= 80:
        grade = "Excellent"
    elif overall >= 65:
        grade = "Good"
    elif overall >= 45:
        grade = "Fair"
    elif overall >= 25:
        grade = "Needs Work"
    else:
        grade = "Critical"
    
    return {
        "overall_score": round(overall, 1),
        "grade": grade,
        "savings_rate": round(savings_rate, 1),
        "investment_rate": round(investment_rate, 1),
        "expense_ratio": round(expense_ratio, 1),
        "goal_progress": round(goal_score, 1),
        "breakdown": {
            "savings": round(savings_score, 1),
            "investments": round(invest_score, 1),
            "spending": round(expense_score, 1),
            "goals": round(goal_score, 1),
        }
    }

# ══════════════════════════════════════
#  AI CHAT ENDPOINTS
# ══════════════════════════════════════

@api_router.post("/ai/chat")
async def ai_chat(msg: AIMessageCreate, user=Depends(get_current_user)):
    from emergentintegrations.llm.chat import LlmChat, UserMessage
    
    user_id = user["id"]
    now = datetime.now(timezone.utc).isoformat()
    
    # Save user message
    user_msg_id = str(uuid.uuid4())
    await db.chat_history.insert_one({
        "id": user_msg_id,
        "user_id": user_id,
        "role": "user",
        "content": msg.message,
        "created_at": now,
    })
    
    # ── Gather ALL user data for full context ──
    txns = await db.transactions.find({"user_id": user_id}, {"_id": 0}).to_list(500)
    goals = await db.goals.find({"user_id": user_id}, {"_id": 0}).to_list(50)
    risk_doc = await db.risk_profiles.find_one({"user_id": user_id}, {"_id": 0})
    holdings = await db.holdings.find({"user_id": user_id}, {"_id": 0}).to_list(100)
    sips = await db.recurring_transactions.find({"user_id": user_id}, {"_id": 0}).to_list(50)
    budgets = await db.budgets.find({"user_id": user_id}, {"_id": 0}).to_list(50)
    loans = await db.loans.find({"user_id": user_id}, {"_id": 0}).to_list(20)
    
    total_income = sum(t["amount"] for t in txns if t["type"] == "income")
    total_expenses = sum(t["amount"] for t in txns if t["type"] == "expense")
    total_investments_txn = sum(t["amount"] for t in txns if t["type"] == "investment")
    
    # ── Holdings / Portfolio ──
    total_holdings_invested = sum(h.get("invested_value", 0) or (h.get("buy_price", 0) * h.get("quantity", 0)) for h in holdings)
    total_holdings_current = sum(h.get("current_value", 0) or (h.get("buy_price", 0) * h.get("quantity", 0)) for h in holdings)
    holdings_gain_loss = total_holdings_current - total_holdings_invested
    holdings_gain_pct = (holdings_gain_loss / total_holdings_invested * 100) if total_holdings_invested > 0 else 0
    
    holdings_summary = []
    for h in holdings[:15]:
        name = h.get("name", "Unknown")
        qty = h.get("quantity", 0)
        invested = h.get("invested_value", 0) or (h.get("buy_price", 0) * qty)
        current = h.get("current_value", 0) or invested
        gain = current - invested
        gain_pct = (gain / invested * 100) if invested > 0 else 0
        holdings_summary.append(f"{name}: Qty={qty}, Invested=₹{invested:,.0f}, Current=₹{current:,.0f} ({gain_pct:+.1f}%)")
    
    # ── Expense Category Breakdown ──
    category_breakdown = {}
    for t in txns:
        if t["type"] == "expense":
            cat = t["category"]
            category_breakdown[cat] = category_breakdown.get(cat, 0) + t["amount"]
    
    # ── Monthly Trends (last 6 months) ──
    monthly_trends = {}
    for t in txns:
        dt_str = t.get("date", "")
        if dt_str:
            month_key = dt_str[:7]
            if month_key not in monthly_trends:
                monthly_trends[month_key] = {"income": 0, "expenses": 0, "investments": 0}
            if t["type"] == "income":
                monthly_trends[month_key]["income"] += t["amount"]
            elif t["type"] == "expense":
                monthly_trends[month_key]["expenses"] += t["amount"]
            elif t["type"] == "investment":
                monthly_trends[month_key]["investments"] += t["amount"]
    monthly_trend_str = "\n".join(f"  {m}: Income=₹{d['income']:,.0f}, Expenses=₹{d['expenses']:,.0f}, Invest=₹{d['investments']:,.0f}" for m, d in sorted(monthly_trends.items())[-6:])
    
    # ── Goals ──
    goal_summary = []
    for g in goals:
        pct = (g['current_amount'] / g['target_amount'] * 100) if g['target_amount'] > 0 else 0
        goal_summary.append(f"{g['title']}: ₹{g['current_amount']:,.0f}/₹{g['target_amount']:,.0f} ({pct:.0f}%) - Deadline: {g.get('deadline', 'N/A')}")
    
    # ── SIPs ──
    sip_summary = []
    for s in sips:
        sip_summary.append(f"{s.get('description', 'SIP')}: ₹{s.get('amount', 0):,.0f}/{s.get('frequency', 'monthly')} in {s.get('category', 'N/A')}")
    
    # ── Budgets ──
    budget_summary = []
    for b in budgets:
        spent = b.get("spent", 0)
        limit_amt = b.get("limit", 0) or b.get("amount", 0)
        usage_pct = (spent / limit_amt * 100) if limit_amt > 0 else 0
        budget_summary.append(f"{b.get('category', 'N/A')}: ₹{spent:,.0f}/₹{limit_amt:,.0f} ({usage_pct:.0f}% used)")
    
    # ── Loans ──
    loan_summary = []
    for l in loans:
        loan_summary.append(f"{l.get('name', 'Loan')}: ₹{l.get('principal', 0):,.0f} @ {l.get('interest_rate', 0)}% for {l.get('tenure_years', 0)}yrs, EMI=₹{l.get('emi', 0):,.0f}")
    
    # ── Health Score ──
    hs_income = total_income or 1
    hs_savings_rate = max(0, (hs_income - total_expenses) / hs_income * 100)
    hs_investment_rate = (total_investments_txn / hs_income * 100)
    hs_overall = min(100, hs_savings_rate * 0.35 + min(hs_investment_rate, 30) * 0.25 + max(0, 100 - (total_expenses / hs_income * 100)) * 0.25 + min((len(goals) > 0) * 50 + sum(1 for g in goals if g['current_amount'] >= g['target_amount'] * 0.5) * 25, 100) * 0.15)
    
    # ── Buy/Sell investment transactions ──
    buy_sells = [t for t in txns if t.get("buy_sell")]
    buy_sell_summary = []
    for t in buy_sells[:10]:
        buy_sell_summary.append(f"{t.get('buy_sell','').upper()} {t.get('description','')}: {t.get('units',0)} units @ ₹{t.get('price_per_unit',0):,.0f} = ₹{t['amount']:,.0f} on {t.get('date','')}")
    
    # ── Capital Gains ──
    cap_gains_context = ""
    sell_txns_for_cg = [t for t in txns if t.get("buy_sell") == "sell"]
    if sell_txns_for_cg:
        total_cg = 0
        for st in sell_txns_for_cg:
            desc = st.get("description", "")
            buy_match = next((b for b in txns if b.get("buy_sell") == "buy" and b.get("description", "").lower() == desc.lower()), None)
            if buy_match:
                cost = buy_match.get("amount", 0)
                gain = st["amount"] - cost
                total_cg += gain
        cap_gains_context = f"\n- Estimated Capital Gains from Sell Transactions: ₹{total_cg:,.0f}"

    risk_context = ""
    if risk_doc:
        risk_context = f"\n- Risk Profile: {risk_doc.get('profile', 'Not assessed')} (Score: {risk_doc.get('score', 0):.1f}/5)"

    context = f"""User Financial Profile (COMPLETE APP DATA):

INCOME & EXPENSES:
- Total Income: ₹{total_income:,.2f}
- Total Expenses: ₹{total_expenses:,.2f}
- Investment Transactions: ₹{total_investments_txn:,.2f}
- Net Balance: ₹{total_income - total_expenses - total_investments_txn:,.2f}
- Savings Rate: {((total_income - total_expenses) / max(total_income, 1) * 100):.1f}%
- Top Expense Categories: {', '.join(f'{k}: ₹{v:,.0f}' for k, v in sorted(category_breakdown.items(), key=lambda x: -x[1])[:5])}

MONTHLY TRENDS (Last 6 Months):
{monthly_trend_str if monthly_trend_str else '  No monthly data available'}

FINANCIAL HEALTH SCORE: {hs_overall:.1f}/100{risk_context}

INVESTMENT PORTFOLIO ({len(holdings)} holdings):
- Total Invested: ₹{total_holdings_invested:,.2f}
- Current Value: ₹{total_holdings_current:,.2f}
- Total Gain/Loss: ₹{holdings_gain_loss:,.2f} ({holdings_gain_pct:+.1f}%)
- Holdings:
  {chr(10).join('  ' + h for h in holdings_summary) if holdings_summary else '  None'}

BUY/SELL TRANSACTIONS:
  {chr(10).join('  ' + b for b in buy_sell_summary) if buy_sell_summary else '  None'}{cap_gains_context}

SIPS & RECURRING INVESTMENTS ({len(sips)}):
  {chr(10).join('  ' + s for s in sip_summary) if sip_summary else '  None'}

FINANCIAL GOALS ({len(goals)}):
  {chr(10).join('  ' + g for g in goal_summary) if goal_summary else '  None set'}

BUDGETS:
  {chr(10).join('  ' + b for b in budget_summary) if budget_summary else '  No budgets set'}

LOANS/EMIs:
  {chr(10).join('  ' + l for l in loan_summary) if loan_summary else '  No loans'}

RECENT TRANSACTIONS (Last 10):
  {chr(10).join('  ' + f"{t['date']} | {t['type'].upper()} | {t['category']} | ₹{t['amount']:,.0f} | {t.get('description','')}" for t in sorted(txns, key=lambda x: x.get('date',''), reverse=True)[:10]) if txns else '  None'}
"""
    
    system_msg = """You are Visor AI, an expert Indian personal finance advisor with FULL access to the user's financial data in this app. You can see their transactions, investments, holdings, goals, budgets, loans, SIPs, health score, and portfolio.

Key guidelines:
- Always use ₹ (Indian Rupee) for currency
- Reference Indian tax slabs, Section 80C, 80D deductions where relevant
- Suggest Indian investment instruments (PPF, NPS, ELSS, SIP, FD, Gold ETFs)
- Consider Indian inflation rates (~5-6%) in calculations
- Be concise, actionable, and encouraging
- Format numbers in Indian system (lakhs, crores)
- When the user asks about ANY aspect of their finances, reference their ACTUAL data from the context above
- Provide personalized advice based on their real financial situation
- Keep responses under 200 words unless detailed analysis is needed
- If the user asks about something not in their data, let them know and suggest they add it"""
    
    try:
        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=f"visor-{user_id}-{datetime.now(timezone.utc).strftime('%Y%m%d')}",
            system_message=system_msg,
        )
        chat.with_model("openai", "gpt-5.2")
        
        user_message = UserMessage(text=f"{context}\n\nUser Question: {msg.message}")
        response_text = await chat.send_message(user_message)
        
    except Exception as e:
        logger.error(f"AI chat error: {e}")
        response_text = "I'm having trouble connecting right now. Please try again in a moment. In the meantime, here's a tip: Review your monthly expenses and identify subscriptions you no longer use—small savings add up!"
    
    # Save AI response
    ai_msg_id = str(uuid.uuid4())
    ai_now = datetime.now(timezone.utc).isoformat()
    await db.chat_history.insert_one({
        "id": ai_msg_id,
        "user_id": user_id,
        "role": "assistant",
        "content": response_text,
        "created_at": ai_now,
    })
    
    return {
        "id": ai_msg_id,
        "role": "assistant",
        "content": response_text,
        "created_at": ai_now,
    }

@api_router.get("/ai/history")
async def get_chat_history(user=Depends(get_current_user)):
    messages = await db.chat_history.find(
        {"user_id": user["id"]}, {"_id": 0}
    ).sort("created_at", 1).to_list(100)
    return messages

@api_router.delete("/ai/history")
async def clear_chat_history(user=Depends(get_current_user)):
    await db.chat_history.delete_many({"user_id": user["id"]})
    return {"message": "Chat history cleared"}

# ══════════════════════════════════════
#  RISK PROFILE
# ══════════════════════════════════════

class RiskProfileCreate(BaseModel):
    answers: list  # list of {question_id, value, category}
    score: float
    profile: str  # Conservative, Moderate, Aggressive
    breakdown: dict  # category scores

@api_router.post("/risk-profile")
async def save_risk_profile(data: RiskProfileCreate, user=Depends(get_current_user)):
    doc = {
        "user_id": user["id"],
        "answers": data.answers,
        "score": data.score,
        "profile": data.profile,
        "breakdown": data.breakdown,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.risk_profiles.delete_many({"user_id": user["id"]})
    await db.risk_profiles.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.get("/risk-profile")
async def get_risk_profile(user=Depends(get_current_user)):
    doc = await db.risk_profiles.find_one({"user_id": user["id"]}, {"_id": 0})
    if not doc:
        return None
    return doc

# ══════════════════════════════════════
#  TAX PLANNING
# ══════════════════════════════════════

# Tax section mapping: category -> tax section
TAX_SECTION_MAP = {
    "PPF": "80C", "ELSS": "80C", "NPS": "80C", "EPF": "80C",
    "Life Insurance": "80C", "FD": "80C", "Fixed Deposit": "80C", "NSC": "80C",
    "ULIP": "80C", "Sukanya Samriddhi": "80C", "Tax Saver FD": "80C",
    "Health Insurance": "80D",
    "NPS Additional": "80CCD1B",
    "Education Loan": "80E",
}
TAX_LIMITS = {
    "80C": 150000,
    "80D": 25000,      # 50000 for senior citizens
    "80CCD1B": 50000,
    "80E": 0,          # No limit
    "80TTA": 10000,
}
TAX_SECTION_LABELS = {
    "80C": "Section 80C",
    "80D": "Section 80D",
    "80CCD1B": "Section 80CCD(1B)",
    "80E": "Section 80E",
    "80TTA": "Section 80TTA",
}
TAX_SECTION_ICONS = {
    "80C": "shield-lock-outline",
    "80D": "hospital-box-outline",
    "80CCD1B": "cash-plus",
    "80E": "school-outline",
    "80TTA": "bank-outline",
}

@api_router.get("/tax-summary")
async def get_tax_summary(user=Depends(get_current_user)):
    # Fetch investments
    txns = await db.transactions.find(
        {"user_id": user["id"], "type": "investment"}, {"_id": 0}
    ).to_list(1000)
    holdings_list = []
    async for doc in db.holdings.find({"user_id": user["id"]}):
        doc["id"] = str(doc.pop("_id"))
        holdings_list.append(doc)

    sections: dict = {}
    for sec_id, limit in TAX_LIMITS.items():
        sections[sec_id] = {
            "section": sec_id,
            "label": TAX_SECTION_LABELS.get(sec_id, sec_id),
            "icon": TAX_SECTION_ICONS.get(sec_id, "file-document-outline"),
            "limit": limit,
            "used": 0,
            "items": [],
        }

    # Map transactions to sections
    for t in txns:
        cat = t.get("category", "")
        sec = TAX_SECTION_MAP.get(cat)
        if sec and sec in sections:
            amt = t.get("amount", 0)
            sections[sec]["used"] += amt
            sections[sec]["items"].append({"name": cat, "amount": amt, "source": "transaction"})

    # Map holdings to sections
    for h in holdings_list:
        cat = h.get("category", "")
        sec = TAX_SECTION_MAP.get(cat)
        if sec and sec in sections:
            amt = h.get("quantity", 0) * h.get("buy_price", 0)
            sections[sec]["used"] += amt
            sections[sec]["items"].append({"name": h.get("name", cat), "amount": amt, "source": "holding"})

    # Cap used at limit (except unlimited sections like 80E)
    for sec_id in sections:
        sections[sec_id]["used"] = round(sections[sec_id]["used"], 2)
        limit = sections[sec_id]["limit"]
        pct = (sections[sec_id]["used"] / limit * 100) if limit > 0 else 0
        sections[sec_id]["percentage"] = round(min(pct, 100), 1)
        remaining = max(limit - sections[sec_id]["used"], 0) if limit > 0 else 0
        sections[sec_id]["remaining"] = round(remaining, 2)

    # Total tax saved estimate (30% slab)
    total_deductions = sum(
        min(s["used"], s["limit"]) if s["limit"] > 0 else s["used"]
        for s in sections.values()
    )
    tax_saved_30 = round(total_deductions * 0.30, 2)
    tax_saved_20 = round(total_deductions * 0.20, 2)

    # Only include sections that have usage or are important
    active = [s for s in sections.values() if s["used"] > 0 or s["section"] in ("80C", "80D")]

    return {
        "sections": active,
        "total_deductions": round(total_deductions, 2),
        "tax_saved_30_slab": tax_saved_30,
        "tax_saved_20_slab": tax_saved_20,
        "fy": "2025-26",
    }


@api_router.get("/capital-gains")
async def get_capital_gains(user=Depends(get_current_user)):
    """Calculate capital gains tax on sold investments."""
    # Get all investment transactions with buy_sell field
    txns = await db.transactions.find(
        {"user_id": user["id"], "type": "investment"}, {"_id": 0}
    ).to_list(1000)
    
    # Group by category/description for FIFO matching
    buys = {}  # {category: [{date, amount, units, price_per_unit}]}
    sells = []  # [{date, amount, units, price_per_unit, category, description}]
    
    for t in txns:
        key = t.get("description", t.get("category", "Unknown"))
        buy_sell = t.get("buy_sell", "buy")  # Default to buy for legacy data
        
        if buy_sell == "sell":
            sells.append({
                "date": t.get("date", ""),
                "amount": t.get("amount", 0),
                "units": t.get("units", 0),
                "price_per_unit": t.get("price_per_unit", 0),
                "category": t.get("category", ""),
                "description": t.get("description", ""),
                "key": key,
            })
        else:
            if key not in buys:
                buys[key] = []
            buys[key].append({
                "date": t.get("date", ""),
                "amount": t.get("amount", 0),
                "units": t.get("units", t.get("amount", 0)),  # Fallback to amount if units not set
                "price_per_unit": t.get("price_per_unit", 1),
            })
    
    # Calculate gains for each sell using FIFO
    gains = []
    total_stcg = 0  # Short term (<1 year for equity, <3 years for others)
    total_ltcg = 0  # Long term
    
    for sell in sells:
        key = sell["key"]
        sell_date = datetime.strptime(sell["date"], "%Y-%m-%d") if sell["date"] else datetime.now(timezone.utc)
        sell_amount = sell["amount"]
        sell_units = sell["units"] or 1
        
        cost_basis = 0
        is_long_term = False
        holding_days = 0
        
        # FIFO matching with buys
        if key in buys and buys[key]:
            buy_entry = buys[key][0]  # Get oldest buy
            buy_date = datetime.strptime(buy_entry["date"], "%Y-%m-%d") if buy_entry["date"] else sell_date
            holding_days = (sell_date - buy_date).days
            
            # Equity: 1 year = LTCG, Others: 2 years = LTCG
            is_equity = sell["category"] in ("Stocks", "Stock", "Mutual Funds", "ETF", "ELSS", "SIP")
            ltcg_threshold = 365 if is_equity else 730
            is_long_term = holding_days >= ltcg_threshold
            
            # Calculate cost basis
            if sell_units > 0 and buy_entry.get("price_per_unit", 0) > 0:
                cost_basis = sell_units * buy_entry["price_per_unit"]
            else:
                cost_basis = buy_entry.get("amount", sell_amount)
        else:
            # No matching buy found, assume short-term with zero cost basis
            cost_basis = 0
        
        gain = sell_amount - cost_basis
        gain_pct = (gain / cost_basis * 100) if cost_basis > 0 else 0
        
        # Tax calculation based on Indian tax rules (FY 2025-26)
        # STCG on equity: 20%, LTCG on equity: 12.5% (above ₹1.25L exemption)
        # STCG on others: As per slab, LTCG on others: 20% with indexation
        is_equity = sell["category"] in ("Stocks", "Stock", "Mutual Funds", "ETF", "ELSS", "SIP")
        
        if is_long_term:
            if is_equity:
                # LTCG on equity: 12.5% above ₹1.25L exemption
                taxable_gain = max(gain, 0)
                tax_rate = 0.125
                total_ltcg += taxable_gain
            else:
                # LTCG on debt: 12.5% without indexation (new rules)
                taxable_gain = max(gain, 0)
                tax_rate = 0.125
                total_ltcg += taxable_gain
        else:
            if is_equity:
                # STCG on equity: 20%
                taxable_gain = max(gain, 0)
                tax_rate = 0.20
                total_stcg += taxable_gain
            else:
                # STCG on debt: As per slab (assume 30%)
                taxable_gain = max(gain, 0)
                tax_rate = 0.30
                total_stcg += taxable_gain
        
        tax_liability = taxable_gain * tax_rate
        
        gains.append({
            "description": sell["description"] or sell["category"],
            "category": sell["category"],
            "sell_date": sell["date"],
            "sell_amount": round(sell_amount, 2),
            "cost_basis": round(cost_basis, 2),
            "gain_loss": round(gain, 2),
            "gain_loss_pct": round(gain_pct, 2),
            "holding_days": holding_days,
            "is_long_term": is_long_term,
            "tax_rate": tax_rate,
            "tax_liability": round(tax_liability, 2),
        })
    
    # LTCG exemption for equity: ₹1.25L
    ltcg_exemption = 125000
    ltcg_taxable = max(total_ltcg - ltcg_exemption, 0)
    ltcg_tax = ltcg_taxable * 0.125 if ltcg_taxable > 0 else 0
    stcg_tax = total_stcg * 0.20  # Simplified, assuming equity STCG
    
    return {
        "gains": gains,
        "summary": {
            "total_stcg": round(total_stcg, 2),
            "total_ltcg": round(total_ltcg, 2),
            "ltcg_exemption": ltcg_exemption,
            "ltcg_taxable": round(ltcg_taxable, 2),
            "estimated_stcg_tax": round(stcg_tax, 2),
            "estimated_ltcg_tax": round(ltcg_tax, 2),
            "total_estimated_tax": round(stcg_tax + ltcg_tax, 2),
        },
        "notes": [
            "STCG on equity: 20% (holding < 1 year)",
            "LTCG on equity: 12.5% above ₹1.25L exemption (holding ≥ 1 year)",
            "Debt funds: 12.5% LTCG (≥2 years), slab rate STCG",
        ],
        "fy": "2025-26",
    }


# ══════════════════════════════════════
#  PORTFOLIO REBALANCING
# ══════════════════════════════════════

STRATEGY_ALLOCATIONS = {
    "Conservative": {"Equity": 25, "Debt": 60, "Gold": 15},
    "Moderate": {"Equity": 40, "Debt": 30, "Gold": 15, "Alt": 15},
    "Aggressive": {"Equity": 70, "Debt": 10, "Gold": 5, "Alt": 15},
}

# Map portfolio categories to allocation buckets
CATEGORY_TO_BUCKET = {
    "Stock": "Equity", "Stocks": "Equity", "Mutual Fund": "Equity", "ELSS": "Equity", "ETF": "Equity",
    "SIP": "Equity",
    "PPF": "Debt", "FD": "Debt", "Fixed Deposit": "Debt", "EPF": "Debt", "NPS": "Debt",
    "NSC": "Debt", "Bonds": "Debt",
    "Gold": "Gold", "Sovereign Gold Bond": "Gold", "Silver": "Gold",
    "Crypto": "Alt", "Real Estate": "Alt", "REIT": "Alt",
}

@api_router.get("/portfolio-rebalancing")
async def get_rebalancing(user=Depends(get_current_user)):
    # Get risk profile
    risk_doc = await db.risk_profiles.find_one({"user_id": user["id"]}, {"_id": 0})
    profile = risk_doc.get("profile", "Moderate") if risk_doc else "Moderate"
    target = STRATEGY_ALLOCATIONS.get(profile, STRATEGY_ALLOCATIONS["Moderate"])

    # Get portfolio overview for actual allocation
    txns = await db.transactions.find(
        {"user_id": user["id"], "type": "investment"}, {"_id": 0}
    ).to_list(1000)
    holdings_list = []
    async for doc in db.holdings.find({"user_id": user["id"]}):
        doc["id"] = str(doc.pop("_id"))
        holdings_list.append(doc)

    # Calculate actual allocation by bucket
    actual_amounts: dict = {}
    for t in txns:
        cat = t.get("category", "")
        bucket = CATEGORY_TO_BUCKET.get(cat, "Alt")
        actual_amounts[bucket] = actual_amounts.get(bucket, 0) + t.get("amount", 0)

    for h in holdings_list:
        cat = h.get("category", "")
        bucket = CATEGORY_TO_BUCKET.get(cat, "Alt")
        invested = h.get("quantity", 0) * h.get("buy_price", 0)
        actual_amounts[bucket] = actual_amounts.get(bucket, 0) + invested

    total = sum(actual_amounts.values())
    if total == 0:
        return {"profile": profile, "total": 0, "actions": [], "actual": {}, "target": target}

    actual_pct: dict = {}
    for bucket in set(list(target.keys()) + list(actual_amounts.keys())):
        actual_pct[bucket] = round((actual_amounts.get(bucket, 0) / total) * 100, 1)

    # Generate rebalancing actions
    actions = []
    for bucket in sorted(set(list(target.keys()) + list(actual_pct.keys()))):
        t_pct = target.get(bucket, 0)
        a_pct = actual_pct.get(bucket, 0)
        diff_pct = round(a_pct - t_pct, 1)
        diff_amount = round(total * diff_pct / 100, 0)
        if abs(diff_pct) >= 2:  # Only flag if >2% off
            actions.append({
                "bucket": bucket,
                "target_pct": t_pct,
                "actual_pct": a_pct,
                "diff_pct": diff_pct,
                "diff_amount": diff_amount,
                "action": "reduce" if diff_pct > 0 else "increase",
                "suggestion": f"{'Reduce' if diff_pct > 0 else 'Increase'} {bucket} by {abs(diff_pct):.1f}% ({formatINR_py(abs(diff_amount))})"
            })

    return {
        "profile": profile,
        "strategy_name": {"Conservative": "Safe Harbor", "Moderate": "Balanced Growth", "Aggressive": "High Growth"}.get(profile, profile),
        "total": round(total, 2),
        "actual": actual_pct,
        "target": target,
        "actions": sorted(actions, key=lambda x: abs(x["diff_pct"]), reverse=True),
    }

def formatINR_py(amount: float) -> str:
    """Format amount in Indian Rupee style"""
    if abs(amount) >= 100000:
        return f"₹{amount/100000:.1f}L"
    elif abs(amount) >= 1000:
        return f"₹{amount/1000:.1f}K"
    return f"₹{amount:,.0f}"

# ══════════════════════════════════════
#  RECURRING TRANSACTIONS (SIPs)
# ══════════════════════════════════════

class RecurringCreate(BaseModel):
    name: str
    amount: float
    frequency: str  # daily, weekly, monthly, quarterly, yearly
    category: str  # SIP, PPF, NPS, Insurance, etc.
    start_date: str
    end_date: Optional[str] = None
    day_of_month: int = 1  # For monthly: which day (1-28)
    notes: Optional[str] = None
    is_active: bool = True

class RecurringUpdate(BaseModel):
    name: Optional[str] = None
    amount: Optional[float] = None
    frequency: Optional[str] = None
    category: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    day_of_month: Optional[int] = None
    notes: Optional[str] = None
    is_active: Optional[bool] = None

def calculate_next_execution(frequency: str, last_date: str, day_of_month: int = 1) -> str:
    """Calculate the next execution date based on frequency."""
    last = datetime.strptime(last_date, "%Y-%m-%d")
    if frequency == "daily":
        next_date = last + timedelta(days=1)
    elif frequency == "weekly":
        next_date = last + timedelta(weeks=1)
    elif frequency == "monthly":
        # Move to next month, keeping day_of_month
        if last.month == 12:
            next_date = datetime(last.year + 1, 1, min(day_of_month, 28))
        else:
            next_date = datetime(last.year, last.month + 1, min(day_of_month, 28))
    elif frequency == "quarterly":
        # Move 3 months ahead
        month = last.month + 3
        year = last.year
        if month > 12:
            month -= 12
            year += 1
        next_date = datetime(year, month, min(day_of_month, 28))
    elif frequency == "yearly":
        next_date = datetime(last.year + 1, last.month, min(day_of_month, 28))
    else:
        next_date = last + timedelta(days=30)  # Default monthly
    return next_date.strftime("%Y-%m-%d")

def get_upcoming_executions(recurring: dict, count: int = 6) -> list:
    """Generate upcoming execution dates for a recurring transaction."""
    executions = []
    current_date = recurring.get("next_execution") or recurring["start_date"]
    end_date = recurring.get("end_date")
    
    for _ in range(count):
        if end_date and current_date > end_date:
            break
        executions.append({
            "date": current_date,
            "amount": recurring["amount"],
            "status": "scheduled"
        })
        current_date = calculate_next_execution(
            recurring["frequency"], 
            current_date, 
            recurring.get("day_of_month", 1)
        )
    return executions

@api_router.get("/recurring")
async def get_recurring_transactions(user=Depends(get_current_user)):
    """Get all recurring transactions (SIPs) for the user."""
    recurring_list = []
    cursor = db.recurring_transactions.find({"user_id": user["id"]})
    async for doc in cursor:
        doc["id"] = str(doc.pop("_id"))
        # Calculate upcoming executions
        doc["upcoming"] = get_upcoming_executions(doc, 3)
        recurring_list.append(doc)
    
    # Calculate summary stats
    active = [r for r in recurring_list if r.get("is_active", True)]
    monthly_commitment = sum(
        r["amount"] * (12 if r["frequency"] == "yearly" else 
                       4 if r["frequency"] == "quarterly" else
                       1 if r["frequency"] == "monthly" else
                       4.33 if r["frequency"] == "weekly" else 30)
        for r in active
    ) / 12  # Normalize to monthly
    
    return {
        "recurring": recurring_list,
        "summary": {
            "total_count": len(recurring_list),
            "active_count": len(active),
            "monthly_commitment": round(monthly_commitment, 2),
            "categories": list(set(r["category"] for r in recurring_list)),
        }
    }

@api_router.post("/recurring")
async def create_recurring_transaction(data: RecurringCreate, user=Depends(get_current_user)):
    """Create a new recurring transaction (SIP)."""
    now = datetime.now(timezone.utc).isoformat()
    
    # Calculate first execution date
    start = datetime.strptime(data.start_date, "%Y-%m-%d")
    today = datetime.now()
    
    # If start date is in the past, calculate next valid execution
    if start < today:
        next_exec = data.start_date
        while datetime.strptime(next_exec, "%Y-%m-%d") < today:
            next_exec = calculate_next_execution(data.frequency, next_exec, data.day_of_month)
    else:
        next_exec = data.start_date
    
    doc = {
        "user_id": user["id"],
        "name": data.name,
        "amount": data.amount,
        "frequency": data.frequency,
        "category": data.category,
        "start_date": data.start_date,
        "end_date": data.end_date,
        "day_of_month": data.day_of_month,
        "notes": data.notes,
        "is_active": data.is_active,
        "next_execution": next_exec,
        "total_invested": 0,
        "execution_count": 0,
        "created_at": now,
    }
    result = await db.recurring_transactions.insert_one(doc)
    doc["id"] = str(result.inserted_id)
    doc.pop("_id", None)
    doc["upcoming"] = get_upcoming_executions(doc, 3)
    return doc

@api_router.put("/recurring/{recurring_id}")
async def update_recurring_transaction(recurring_id: str, data: RecurringUpdate, user=Depends(get_current_user)):
    """Update a recurring transaction."""
    from bson import ObjectId
    
    existing = await db.recurring_transactions.find_one(
        {"_id": ObjectId(recurring_id), "user_id": user["id"]}
    )
    if not existing:
        raise HTTPException(404, "Recurring transaction not found")
    
    update_data = {k: v for k, v in data.dict().items() if v is not None}
    
    # Recalculate next execution if frequency or day changes
    if "frequency" in update_data or "day_of_month" in update_data:
        freq = update_data.get("frequency", existing["frequency"])
        day = update_data.get("day_of_month", existing.get("day_of_month", 1))
        today_str = datetime.now().strftime("%Y-%m-%d")
        update_data["next_execution"] = calculate_next_execution(freq, today_str, day)
    
    if update_data:
        await db.recurring_transactions.update_one(
            {"_id": ObjectId(recurring_id)}, {"$set": update_data}
        )
    
    updated = await db.recurring_transactions.find_one({"_id": ObjectId(recurring_id)})
    updated["id"] = str(updated.pop("_id"))
    updated["upcoming"] = get_upcoming_executions(updated, 3)
    return updated

@api_router.delete("/recurring/{recurring_id}")
async def delete_recurring_transaction(recurring_id: str, user=Depends(get_current_user)):
    """Delete a recurring transaction."""
    from bson import ObjectId
    
    result = await db.recurring_transactions.delete_one(
        {"_id": ObjectId(recurring_id), "user_id": user["id"]}
    )
    if result.deleted_count == 0:
        raise HTTPException(404, "Recurring transaction not found")
    return {"message": "Deleted"}

@api_router.post("/recurring/{recurring_id}/execute")
async def execute_recurring_transaction(recurring_id: str, user=Depends(get_current_user)):
    """Manually execute a recurring transaction (create actual transaction)."""
    from bson import ObjectId
    
    recurring = await db.recurring_transactions.find_one(
        {"_id": ObjectId(recurring_id), "user_id": user["id"]}
    )
    if not recurring:
        raise HTTPException(404, "Recurring transaction not found")
    
    # Create actual transaction
    txn_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    txn_doc = {
        "id": txn_id,
        "user_id": user["id"],
        "type": "investment",
        "amount": recurring["amount"],
        "category": recurring["category"],
        "description": f"{recurring['name']} - Auto SIP",
        "date": now.strftime("%Y-%m-%d"),
        "is_recurring": True,
        "recurring_frequency": recurring["frequency"],
        "recurring_id": recurring_id,
        "notes": recurring.get("notes"),
        "created_at": now.isoformat(),
    }
    await db.transactions.insert_one(txn_doc)
    
    # Update recurring transaction stats
    next_exec = calculate_next_execution(
        recurring["frequency"],
        recurring["next_execution"],
        recurring.get("day_of_month", 1)
    )
    await db.recurring_transactions.update_one(
        {"_id": ObjectId(recurring_id)},
        {
            "$set": {
                "next_execution": next_exec,
                "last_execution": now.strftime("%Y-%m-%d"),
            },
            "$inc": {
                "total_invested": recurring["amount"],
                "execution_count": 1,
            }
        }
    )
    
    return {
        "message": "Transaction executed",
        "transaction_id": txn_id,
        "next_execution": next_exec,
    }

@api_router.post("/recurring/{recurring_id}/pause")
async def pause_recurring_transaction(recurring_id: str, user=Depends(get_current_user)):
    """Pause/Resume a recurring transaction."""
    from bson import ObjectId
    
    recurring = await db.recurring_transactions.find_one(
        {"_id": ObjectId(recurring_id), "user_id": user["id"]}
    )
    if not recurring:
        raise HTTPException(404, "Recurring transaction not found")
    
    new_status = not recurring.get("is_active", True)
    await db.recurring_transactions.update_one(
        {"_id": ObjectId(recurring_id)},
        {"$set": {"is_active": new_status}}
    )
    
    return {"message": f"{'Resumed' if new_status else 'Paused'}", "is_active": new_status}

# ══════════════════════════════════════
#  FIXED ASSETS ENDPOINTS
# ══════════════════════════════════════

@api_router.get("/assets", response_model=List[FixedAssetResponse])
async def get_fixed_assets(user=Depends(get_current_user)):
    assets = await db.fixed_assets.find({"user_id": user["id"]}, {"_id": 0}).to_list(100)
    # Calculate accumulated depreciation for each asset
    result = []
    for asset in assets:
        purchase_date = datetime.fromisoformat(asset["purchase_date"].replace("Z", "+00:00")) if "T" in asset["purchase_date"] else datetime.strptime(asset["purchase_date"], "%Y-%m-%d")
        years_held = (datetime.now(timezone.utc) - purchase_date.replace(tzinfo=timezone.utc)).days / 365.25
        acc_dep = min(asset["purchase_value"], asset["purchase_value"] * (asset.get("depreciation_rate", 10) / 100) * years_held)
        asset["accumulated_depreciation"] = round(acc_dep, 2)
        result.append(FixedAssetResponse(**asset))
    return result

@api_router.post("/assets", response_model=FixedAssetResponse)
async def create_fixed_asset(asset: FixedAssetCreate, user=Depends(get_current_user)):
    asset_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    asset_doc = {
        "id": asset_id,
        "user_id": user["id"],
        "name": asset.name,
        "category": asset.category,
        "purchase_date": asset.purchase_date,
        "purchase_value": asset.purchase_value,
        "current_value": asset.current_value,
        "depreciation_rate": asset.depreciation_rate,
        "notes": asset.notes,
        "created_at": now,
    }
    await db.fixed_assets.insert_one(asset_doc)
    asset_doc["accumulated_depreciation"] = 0
    return FixedAssetResponse(**asset_doc)

@api_router.put("/assets/{asset_id}", response_model=FixedAssetResponse)
async def update_fixed_asset(asset_id: str, asset_update: FixedAssetUpdate, user=Depends(get_current_user)):
    existing = await db.fixed_assets.find_one({"id": asset_id, "user_id": user["id"]}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Asset not found")
    
    update_data = {k: v for k, v in asset_update.dict().items() if v is not None}
    if update_data:
        await db.fixed_assets.update_one({"id": asset_id}, {"$set": update_data})
    
    updated = await db.fixed_assets.find_one({"id": asset_id}, {"_id": 0})
    purchase_date = datetime.fromisoformat(updated["purchase_date"].replace("Z", "+00:00")) if "T" in updated["purchase_date"] else datetime.strptime(updated["purchase_date"], "%Y-%m-%d")
    years_held = (datetime.now(timezone.utc) - purchase_date.replace(tzinfo=timezone.utc)).days / 365.25
    acc_dep = min(updated["purchase_value"], updated["purchase_value"] * (updated.get("depreciation_rate", 10) / 100) * years_held)
    updated["accumulated_depreciation"] = round(acc_dep, 2)
    return FixedAssetResponse(**updated)

@api_router.delete("/assets/{asset_id}")
async def delete_fixed_asset(asset_id: str, user=Depends(get_current_user)):
    result = await db.fixed_assets.delete_one({"id": asset_id, "user_id": user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Asset not found")
    return {"message": "Asset deleted"}

# ══════════════════════════════════════
#  LOAN/LIABILITY ENDPOINTS
# ══════════════════════════════════════

def calculate_emi(principal: float, annual_rate: float, tenure_months: int) -> float:
    """Calculate EMI using standard formula"""
    if annual_rate == 0:
        return principal / tenure_months
    monthly_rate = annual_rate / 12 / 100
    emi = principal * monthly_rate * ((1 + monthly_rate) ** tenure_months) / (((1 + monthly_rate) ** tenure_months) - 1)
    return round(emi, 2)

def generate_emi_schedule(principal: float, annual_rate: float, tenure_months: int, start_date: str, emi: float) -> List[dict]:
    """Generate complete EMI amortization schedule"""
    schedule = []
    balance = principal
    monthly_rate = annual_rate / 12 / 100 if annual_rate > 0 else 0
    
    start = datetime.strptime(start_date, "%Y-%m-%d")
    today = datetime.now()
    
    for month in range(1, tenure_months + 1):
        payment_date = start + timedelta(days=30 * month)
        
        interest = balance * monthly_rate
        principal_component = emi - interest
        closing_balance = max(0, balance - principal_component)
        
        # Determine status
        if payment_date < today:
            status = "paid"
        elif payment_date.month == today.month and payment_date.year == today.year:
            status = "current"
        else:
            status = "upcoming"
        
        schedule.append({
            "month": month,
            "date": payment_date.strftime("%Y-%m-%d"),
            "opening_balance": round(balance, 2),
            "emi": emi,
            "principal": round(principal_component, 2),
            "interest": round(interest, 2),
            "closing_balance": round(closing_balance, 2),
            "status": status,
        })
        
        balance = closing_balance
        if balance <= 0:
            break
    
    return schedule

@api_router.get("/loans")
async def get_loans(user=Depends(get_current_user)):
    """Get all loans with calculated outstanding amounts"""
    loans = await db.loans.find({"user_id": user["id"]}, {"_id": 0}).to_list(100)
    result = []
    dek = user.get("encryption_key", "")
    
    for loan in loans:
        # Decrypt sensitive fields
        if dek:
            decrypt_sensitive_fields(loan, dek, LOAN_SENSITIVE_FIELDS)
        
        emi = loan.get("emi_amount") or calculate_emi(loan["principal_amount"], loan["interest_rate"], loan["tenure_months"])
        schedule = generate_emi_schedule(
            loan["principal_amount"],
            loan["interest_rate"],
            loan["tenure_months"],
            loan["start_date"],
            emi
        )
        
        # Calculate paid amounts
        paid_emis = [s for s in schedule if s["status"] == "paid"]
        total_principal_paid = sum(s["principal"] for s in paid_emis)
        total_interest_paid = sum(s["interest"] for s in paid_emis)
        outstanding = loan["principal_amount"] - total_principal_paid
        remaining_emis = loan["tenure_months"] - len(paid_emis)
        
        result.append({
            **loan,
            "emi_amount": emi,
            "outstanding_principal": round(outstanding, 2),
            "total_principal_paid": round(total_principal_paid, 2),
            "total_interest_paid": round(total_interest_paid, 2),
            "remaining_emis": remaining_emis,
        })
    
    return result

@api_router.post("/loans")
async def create_loan(loan: LoanCreate, user=Depends(get_current_user)):
    """Create a new loan"""
    loan_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    dek = user.get("encryption_key", "")
    
    emi = loan.emi_amount or calculate_emi(loan.principal_amount, loan.interest_rate, loan.tenure_months)
    
    # Encrypt sensitive fields before storing
    account_number_raw = loan.account_number
    account_number_enc = encrypt_field(account_number_raw, dek) if dek and account_number_raw else account_number_raw
    
    loan_doc = {
        "id": loan_id,
        "user_id": user["id"],
        "name": loan.name,
        "loan_type": loan.loan_type,
        "principal_amount": loan.principal_amount,
        "interest_rate": loan.interest_rate,
        "tenure_months": loan.tenure_months,
        "start_date": loan.start_date,
        "emi_amount": emi,
        "lender": loan.lender,
        "account_number": account_number_enc,
        "notes": loan.notes,
        "created_at": now,
    }
    await db.loans.insert_one(loan_doc)
    
    # Return plaintext to the caller
    return {
        "id": loan_id,
        "user_id": user["id"],
        "name": loan.name,
        "loan_type": loan.loan_type,
        "principal_amount": loan.principal_amount,
        "interest_rate": loan.interest_rate,
        "tenure_months": loan.tenure_months,
        "start_date": loan.start_date,
        "emi_amount": emi,
        "lender": loan.lender,
        "account_number": account_number_raw,
        "notes": loan.notes,
        "created_at": now,
        "outstanding_principal": loan.principal_amount,
        "total_principal_paid": 0,
        "total_interest_paid": 0,
        "remaining_emis": loan.tenure_months,
    }

@api_router.get("/loans/{loan_id}")
async def get_loan_detail(loan_id: str, user=Depends(get_current_user)):
    """Get loan details with EMI schedule"""
    loan = await db.loans.find_one({"id": loan_id, "user_id": user["id"]}, {"_id": 0})
    if not loan:
        raise HTTPException(status_code=404, detail="Loan not found")
    
    # Decrypt sensitive fields
    dek = user.get("encryption_key", "")
    if dek:
        decrypt_sensitive_fields(loan, dek, LOAN_SENSITIVE_FIELDS)
    
    emi = loan.get("emi_amount") or calculate_emi(loan["principal_amount"], loan["interest_rate"], loan["tenure_months"])
    schedule = generate_emi_schedule(
        loan["principal_amount"],
        loan["interest_rate"],
        loan["tenure_months"],
        loan["start_date"],
        emi
    )
    
    paid_emis = [s for s in schedule if s["status"] == "paid"]
    total_principal_paid = sum(s["principal"] for s in paid_emis)
    total_interest_paid = sum(s["interest"] for s in paid_emis)
    outstanding = loan["principal_amount"] - total_principal_paid
    
    return {
        **loan,
        "emi_amount": emi,
        "outstanding_principal": round(outstanding, 2),
        "total_principal_paid": round(total_principal_paid, 2),
        "total_interest_paid": round(total_interest_paid, 2),
        "remaining_emis": loan["tenure_months"] - len(paid_emis),
        "emi_schedule": schedule,
    }

@api_router.put("/loans/{loan_id}")
async def update_loan(loan_id: str, loan_update: LoanUpdate, user=Depends(get_current_user)):
    """Update a loan"""
    existing = await db.loans.find_one({"id": loan_id, "user_id": user["id"]}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Loan not found")
    
    update_data = {k: v for k, v in loan_update.dict().items() if v is not None}
    
    # Encrypt account_number if being updated
    dek = user.get("encryption_key", "")
    if dek and "account_number" in update_data and update_data["account_number"]:
        update_data["account_number"] = encrypt_field(update_data["account_number"], dek)
    
    if update_data:
        await db.loans.update_one({"id": loan_id}, {"$set": update_data})
    
    updated = await db.loans.find_one({"id": loan_id}, {"_id": 0})
    if dek:
        decrypt_sensitive_fields(updated, dek, LOAN_SENSITIVE_FIELDS)
    return updated

@api_router.delete("/loans/{loan_id}")
async def delete_loan(loan_id: str, user=Depends(get_current_user)):
    """Delete a loan"""
    result = await db.loans.delete_one({"id": loan_id, "user_id": user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Loan not found")
    return {"message": "Loan deleted"}

# ══════════════════════════════════════
#  BOOKKEEPING / LEDGER ENDPOINTS
# ══════════════════════════════════════

def get_indian_fy_dates(fy_year: int = None):
    """Get Indian Financial Year dates (April 1 to March 31)"""
    now = datetime.now(timezone.utc)
    if fy_year is None:
        # Current FY: if we're in Jan-Mar, FY started previous year
        if now.month < 4:
            fy_year = now.year - 1
        else:
            fy_year = now.year
    
    fy_start = datetime(fy_year, 4, 1, tzinfo=timezone.utc)
    fy_end = datetime(fy_year + 1, 3, 31, 23, 59, 59, tzinfo=timezone.utc)
    return fy_start, fy_end

@api_router.get("/books/ledger")
async def get_ledger(
    account_group: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user=Depends(get_current_user)
):
    """Get general ledger with double-entry bookkeeping entries"""
    user_id = user["id"]
    
    # Default to current FY if no dates provided
    if not start_date or not end_date:
        fy_start, fy_end = get_indian_fy_dates()
        start_date = start_date or fy_start.strftime("%Y-%m-%d")
        end_date = end_date or fy_end.strftime("%Y-%m-%d")
    
    # Fetch all transactions in date range
    query = {
        "user_id": user_id,
        "date": {"$gte": start_date, "$lte": end_date}
    }
    txns = await db.transactions.find(query, {"_id": 0}).sort("date", 1).to_list(2000)
    
    # Map transaction types/categories to account groups
    account_mapping = {
        # Assets
        "Salary": "Cash & Bank Balances",
        "Freelance": "Cash & Bank Balances",
        "Bonus": "Cash & Bank Balances",
        "Interest": "Cash & Bank Balances",
        "Dividend": "Cash & Bank Balances",
        # Investments
        "SIP": "Investments",
        "PPF": "Investments",
        "Stocks": "Investments",
        "Mutual Funds": "Investments",
        "FD": "Investments",
        "Gold": "Investments",
        "NPS": "Investments",
        # Expenses
        "Rent": "Rent & Housing",
        "Groceries": "Food & Dining",
        "Food": "Food & Dining",
        "Transport": "Transport & Fuel",
        "Shopping": "Shopping & Personal",
        "Utilities": "Utilities & Bills",
        "Entertainment": "Entertainment & Leisure",
        "Health": "Health & Medical",
        "EMI": "Financial Obligations",
        "Education": "Education & Learning",
        "Insurance": "Insurance Premiums",
    }
    
    # Build ledger entries with double-entry
    ledger_entries = []
    running_balances = {}
    
    for txn in txns:
        cat = txn.get("category", "Other")
        mapped_group = account_mapping.get(cat, "Other")
        
        if account_group and mapped_group != account_group:
            continue
        
        if txn["type"] == "income":
            # Debit: Cash & Bank, Credit: Income account
            ledger_entries.append({
                "date": txn["date"],
                "particulars": txn["description"],
                "voucher_ref": f"TXN-{txn['id'][:8].upper()}",
                "account": "Cash & Bank Balances",
                "debit": txn["amount"],
                "credit": 0,
                "transaction_id": txn["id"],
            })
            ledger_entries.append({
                "date": txn["date"],
                "particulars": txn["description"],
                "voucher_ref": f"TXN-{txn['id'][:8].upper()}",
                "account": f"{cat} Income",
                "debit": 0,
                "credit": txn["amount"],
                "transaction_id": txn["id"],
            })
        elif txn["type"] == "expense":
            # Debit: Expense account, Credit: Cash & Bank
            ledger_entries.append({
                "date": txn["date"],
                "particulars": txn["description"],
                "voucher_ref": f"TXN-{txn['id'][:8].upper()}",
                "account": mapped_group,
                "debit": txn["amount"],
                "credit": 0,
                "transaction_id": txn["id"],
            })
            ledger_entries.append({
                "date": txn["date"],
                "particulars": txn["description"],
                "voucher_ref": f"TXN-{txn['id'][:8].upper()}",
                "account": "Cash & Bank Balances",
                "debit": 0,
                "credit": txn["amount"],
                "transaction_id": txn["id"],
            })
        elif txn["type"] == "investment":
            # Debit: Investment account, Credit: Cash & Bank
            ledger_entries.append({
                "date": txn["date"],
                "particulars": txn["description"],
                "voucher_ref": f"TXN-{txn['id'][:8].upper()}",
                "account": f"{cat} Investment",
                "debit": txn["amount"],
                "credit": 0,
                "transaction_id": txn["id"],
            })
            ledger_entries.append({
                "date": txn["date"],
                "particulars": txn["description"],
                "voucher_ref": f"TXN-{txn['id'][:8].upper()}",
                "account": "Cash & Bank Balances",
                "debit": 0,
                "credit": txn["amount"],
                "transaction_id": txn["id"],
            })
    
    # Group by account and calculate running balances
    accounts = {}
    for entry in ledger_entries:
        acc = entry["account"]
        if acc not in accounts:
            accounts[acc] = {"entries": [], "total_debit": 0, "total_credit": 0}
        accounts[acc]["entries"].append(entry)
        accounts[acc]["total_debit"] += entry["debit"]
        accounts[acc]["total_credit"] += entry["credit"]
    
    # Calculate running balance for each account
    for acc, data in accounts.items():
        running = 0
        for entry in data["entries"]:
            running += entry["debit"] - entry["credit"]
            entry["balance"] = running
        data["closing_balance"] = running
    
    return {
        "fy_start": start_date,
        "fy_end": end_date,
        "accounts": accounts,
        "entry_count": len(ledger_entries),
    }

@api_router.get("/books/pnl")
async def get_profit_loss(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user=Depends(get_current_user)
):
    """Get Profit & Loss (Income & Expenditure) Statement"""
    user_id = user["id"]
    
    # Default to current FY if no dates provided
    if not start_date or not end_date:
        fy_start, fy_end = get_indian_fy_dates()
        start_date = start_date or fy_start.strftime("%Y-%m-%d")
        end_date = end_date or fy_end.strftime("%Y-%m-%d")
    
    query = {
        "user_id": user_id,
        "date": {"$gte": start_date, "$lte": end_date}
    }
    txns = await db.transactions.find(query, {"_id": 0}).to_list(2000)
    
    # Income categories mapping
    income_groups = {
        "A": {"name": "Revenue from Employment", "categories": ["Salary", "Bonus"], "items": {}},
        "B": {"name": "Income from Profession/Freelance", "categories": ["Freelance", "Consulting"], "items": {}},
        "C": {"name": "Income from Investments", "categories": ["Interest", "Dividend", "Capital Gains"], "items": {}},
        "D": {"name": "Other Income", "categories": ["Rental", "Other"], "items": {}},
    }
    
    # Expense categories mapping
    expense_groups = {
        "E": {"name": "Living & Household Expenses", "categories": ["Food", "Groceries", "Utilities", "Rent", "Transport", "Shopping", "Entertainment", "Health"], "items": {}},
        "F": {"name": "Financial Obligations", "categories": ["EMI", "Insurance"], "items": {}},
        "G": {"name": "Taxes & Statutory Deductions", "categories": ["Tax", "TDS"], "items": {}},
        "H": {"name": "Education & Development", "categories": ["Education"], "items": {}},
        "I": {"name": "Other Expenses", "categories": ["Bank Charges", "Depreciation", "Other"], "items": {}},
    }
    
    # Process transactions
    total_income = 0
    total_expenses = 0
    total_investments = 0
    
    for txn in txns:
        cat = txn.get("category", "Other")
        amount = txn["amount"]
        
        if txn["type"] == "income":
            total_income += amount
            # Find appropriate income group
            placed = False
            for group_id, group in income_groups.items():
                if cat in group["categories"] or any(c.lower() in cat.lower() for c in group["categories"]):
                    if cat not in group["items"]:
                        group["items"][cat] = 0
                    group["items"][cat] += amount
                    placed = True
                    break
            if not placed:
                income_groups["D"]["items"][cat] = income_groups["D"]["items"].get(cat, 0) + amount
                
        elif txn["type"] == "expense":
            total_expenses += amount
            # Find appropriate expense group
            placed = False
            for group_id, group in expense_groups.items():
                if cat in group["categories"] or any(c.lower() in cat.lower() for c in group["categories"]):
                    if cat not in group["items"]:
                        group["items"][cat] = 0
                    group["items"][cat] += amount
                    placed = True
                    break
            if not placed:
                expense_groups["I"]["items"][cat] = expense_groups["I"]["items"].get(cat, 0) + amount
                
        elif txn["type"] == "investment":
            total_investments += amount
    
    # Calculate subtotals
    income_sections = []
    for group_id, group in income_groups.items():
        subtotal = sum(group["items"].values())
        if subtotal > 0 or group["items"]:
            income_sections.append({
                "id": group_id,
                "name": group["name"],
                "items": [{"category": k, "amount": v} for k, v in sorted(group["items"].items(), key=lambda x: -x[1])],
                "subtotal": subtotal,
            })
    
    expense_sections = []
    for group_id, group in expense_groups.items():
        subtotal = sum(group["items"].values())
        if subtotal > 0 or group["items"]:
            expense_sections.append({
                "id": group_id,
                "name": group["name"],
                "items": [{"category": k, "amount": v} for k, v in sorted(group["items"].items(), key=lambda x: -x[1])],
                "subtotal": subtotal,
            })
    
    surplus = total_income - total_expenses
    
    return {
        "period_start": start_date,
        "period_end": end_date,
        "income_sections": income_sections,
        "expense_sections": expense_sections,
        "total_income": total_income,
        "total_expenses": total_expenses,
        "total_investments": total_investments,
        "surplus_deficit": surplus,
        "allocation": {
            "to_savings": max(0, surplus * 0.4),
            "to_investments": total_investments,
            "retained": max(0, surplus - total_investments - (surplus * 0.4)),
        }
    }

@api_router.get("/books/balance-sheet")
async def get_balance_sheet(
    as_of_date: Optional[str] = None,
    user=Depends(get_current_user)
):
    """Get Balance Sheet as of a specific date"""
    user_id = user["id"]
    
    if not as_of_date:
        as_of_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    # Get all transactions up to the date
    query = {"user_id": user_id, "date": {"$lte": as_of_date}}
    txns = await db.transactions.find(query, {"_id": 0}).to_list(5000)
    
    # Get fixed assets
    assets = await db.fixed_assets.find({"user_id": user_id}, {"_id": 0}).to_list(100)
    
    # Get goals (as receivables/planned savings)
    goals = await db.goals.find({"user_id": user_id}, {"_id": 0}).to_list(100)
    
    # Calculate totals
    total_income = sum(t["amount"] for t in txns if t["type"] == "income")
    total_expenses = sum(t["amount"] for t in txns if t["type"] == "expense")
    total_investments = sum(t["amount"] for t in txns if t["type"] == "investment")
    
    # Investment breakdown
    invest_by_cat = {}
    for t in txns:
        if t["type"] == "investment":
            cat = t.get("category", "Other")
            invest_by_cat[cat] = invest_by_cat.get(cat, 0) + t["amount"]
    
    # Calculate fixed assets with depreciation
    fixed_asset_items = []
    total_fixed_assets = 0
    total_depreciation = 0
    for asset in assets:
        purchase_date = datetime.fromisoformat(asset["purchase_date"].replace("Z", "+00:00")) if "T" in asset["purchase_date"] else datetime.strptime(asset["purchase_date"], "%Y-%m-%d")
        years_held = (datetime.now(timezone.utc) - purchase_date.replace(tzinfo=timezone.utc)).days / 365.25
        acc_dep = min(asset["purchase_value"], asset["purchase_value"] * (asset.get("depreciation_rate", 10) / 100) * years_held)
        net_value = asset["purchase_value"] - acc_dep
        fixed_asset_items.append({
            "name": asset["name"],
            "category": asset["category"],
            "purchase_value": asset["purchase_value"],
            "accumulated_depreciation": round(acc_dep, 2),
            "net_value": round(net_value, 2),
        })
        total_fixed_assets += asset["purchase_value"]
        total_depreciation += acc_dep
    
    # Build balance sheet structure
    cash_balance = total_income - total_expenses - total_investments
    
    # Non-current assets
    non_current_assets = {
        "fixed_assets": {
            "items": fixed_asset_items,
            "gross_value": total_fixed_assets,
            "depreciation": round(total_depreciation, 2),
            "net_value": round(total_fixed_assets - total_depreciation, 2),
        },
        "long_term_investments": {
            "items": [
                {"name": "PPF", "amount": invest_by_cat.get("PPF", 0)},
                {"name": "NPS", "amount": invest_by_cat.get("NPS", 0)},
                {"name": "Long-term MF", "amount": invest_by_cat.get("Mutual Funds", 0) * 0.5},
            ],
            "total": invest_by_cat.get("PPF", 0) + invest_by_cat.get("NPS", 0) + invest_by_cat.get("Mutual Funds", 0) * 0.5,
        },
        "total": round(total_fixed_assets - total_depreciation + invest_by_cat.get("PPF", 0) + invest_by_cat.get("NPS", 0) + invest_by_cat.get("Mutual Funds", 0) * 0.5, 2),
    }
    
    # Current assets
    current_assets = {
        "short_term_investments": {
            "items": [
                {"name": "Fixed Deposits", "amount": invest_by_cat.get("FD", 0)},
                {"name": "Liquid MF", "amount": invest_by_cat.get("Mutual Funds", 0) * 0.5},
                {"name": "Stocks", "amount": invest_by_cat.get("Stocks", 0)},
                {"name": "Gold", "amount": invest_by_cat.get("Gold", 0)},
                {"name": "SIP Accumulation", "amount": invest_by_cat.get("SIP", 0)},
            ],
            "total": invest_by_cat.get("FD", 0) + invest_by_cat.get("Mutual Funds", 0) * 0.5 + invest_by_cat.get("Stocks", 0) + invest_by_cat.get("Gold", 0) + invest_by_cat.get("SIP", 0),
        },
        "cash_bank": {
            "items": [
                {"name": "Bank Balances", "amount": max(0, cash_balance)},
            ],
            "total": max(0, cash_balance),
        },
        "receivables": {
            "items": [],
            "total": 0,
        },
        "total": round(invest_by_cat.get("FD", 0) + invest_by_cat.get("Mutual Funds", 0) * 0.5 + invest_by_cat.get("Stocks", 0) + invest_by_cat.get("Gold", 0) + invest_by_cat.get("SIP", 0) + max(0, cash_balance), 2),
    }
    
    total_assets = non_current_assets["total"] + current_assets["total"]
    
    # Get loans for liabilities
    loans = await db.loans.find({"user_id": user_id}, {"_id": 0}).to_list(100)
    
    # Calculate loan liabilities
    long_term_loan_items = []
    short_term_loan_items = []
    total_long_term = 0
    total_short_term = 0
    
    for loan in loans:
        emi = loan.get("emi_amount") or calculate_emi(loan["principal_amount"], loan["interest_rate"], loan["tenure_months"])
        schedule = generate_emi_schedule(
            loan["principal_amount"],
            loan["interest_rate"],
            loan["tenure_months"],
            loan["start_date"],
            emi
        )
        
        paid_emis = [s for s in schedule if s["status"] == "paid"]
        total_principal_paid = sum(s["principal"] for s in paid_emis)
        outstanding = loan["principal_amount"] - total_principal_paid
        remaining_emis = loan["tenure_months"] - len(paid_emis)
        
        # Principal due within 12 months = short-term, rest = long-term
        short_term_portion = min(outstanding, emi * min(12, remaining_emis) * (loan["principal_amount"] / (loan["principal_amount"] + loan["interest_rate"] * loan["tenure_months"] / 1200)))
        long_term_portion = outstanding - short_term_portion
        
        if long_term_portion > 0:
            long_term_loan_items.append({
                "name": f"{loan['name']} ({loan['loan_type']})",
                "amount": round(long_term_portion, 2),
                "lender": loan.get("lender"),
            })
            total_long_term += long_term_portion
        
        if short_term_portion > 0:
            short_term_loan_items.append({
                "name": f"{loan['name']} (Current Portion)",
                "amount": round(short_term_portion, 2),
            })
            total_short_term += short_term_portion
    
    # Liabilities
    non_current_liabilities = {
        "long_term_borrowings": {
            "items": long_term_loan_items,
            "total": round(total_long_term, 2),
        },
        "total": round(total_long_term, 2),
    }
    
    current_liabilities = {
        "short_term_borrowings": {
            "items": short_term_loan_items,
            "total": round(total_short_term, 2),
        },
        "payables": {
            "items": [],
            "total": 0,
        },
        "total": round(total_short_term, 2),
    }
    
    total_liabilities = non_current_liabilities["total"] + current_liabilities["total"]
    
    # Net Worth
    net_worth = total_assets - total_liabilities
    
    # Check if balanced
    is_balanced = abs(total_assets - (total_liabilities + net_worth)) < 0.01
    
    return {
        "as_of_date": as_of_date,
        "assets": {
            "non_current": non_current_assets,
            "current": current_assets,
            "total": round(total_assets, 2),
        },
        "liabilities": {
            "non_current": non_current_liabilities,
            "current": current_liabilities,
            "total": round(total_liabilities, 2),
        },
        "net_worth": {
            "opening": 0,
            "surplus_for_period": round(total_income - total_expenses, 2),
            "closing": round(net_worth, 2),
        },
        "total_liabilities_and_net_worth": round(total_liabilities + net_worth, 2),
        "is_balanced": is_balanced,
    }

# ══════════════════════════════════════
#  MARKET DATA (Yahoo Finance — Live)
# ══════════════════════════════════════

import asyncio
import requests
import yfinance as yf
from concurrent.futures import ThreadPoolExecutor

REFRESH_TIMES_IST = ["09:25", "11:30", "12:30", "15:15"]
TROY_OZ_TO_GRAMS = 31.1035
# Domestic premium: international spot → Indian retail (import duty 6% + GST 3% + margin)
GOLD_DOMESTIC_PREMIUM = 1.075
SILVER_DOMESTIC_PREMIUM = 1.155
GOLDAPI_KEY = os.environ.get("GOLDAPI_KEY", "")

_yf_executor = ThreadPoolExecutor(max_workers=2)


def _fetch_goldapi_prices() -> dict:
    """Fetch gold & silver spot prices in INR from GoldAPI.io (most accurate)."""
    result = {}
    if not GOLDAPI_KEY:
        return result
    headers = {"x-access-token": GOLDAPI_KEY, "Content-Type": "application/json"}

    try:
        r = requests.get("https://www.goldapi.io/api/XAU/INR", headers=headers, timeout=10)
        if r.status_code == 200:
            d = r.json()
            gram_24k = d.get("price_gram_24k", 0)
            prev_gram = d.get("prev_close_price", 0) / TROY_OZ_TO_GRAMS if d.get("prev_close_price") else 0
            if gram_24k > 0:
                price_10g = round(gram_24k * 10 * GOLD_DOMESTIC_PREMIUM, 0)
                prev_10g = round(prev_gram * 10 * GOLD_DOMESTIC_PREMIUM, 0) if prev_gram > 0 else price_10g
                change = round(price_10g - prev_10g, 0)
                change_pct = round(change / prev_10g * 100, 2) if prev_10g > 0 else 0
                result["gold"] = {
                    "price": price_10g, "change": change, "change_percent": change_pct,
                    "prev_close": prev_10g, "source": "goldapi.io"
                }
                logger.info(f"GoldAPI: Gold 10g = ₹{price_10g:,.0f}")
    except Exception as e:
        logger.warning(f"GoldAPI gold fetch failed: {e}")

    try:
        r = requests.get("https://www.goldapi.io/api/XAG/INR", headers=headers, timeout=10)
        if r.status_code == 200:
            d = r.json()
            price_oz = d.get("price", 0)
            prev_oz = d.get("prev_close_price", 0)
            if price_oz > 0:
                price_per_gram = price_oz / TROY_OZ_TO_GRAMS
                price_kg = round(price_per_gram * 1000 * SILVER_DOMESTIC_PREMIUM, 0)
                prev_per_gram = prev_oz / TROY_OZ_TO_GRAMS if prev_oz > 0 else 0
                prev_kg = round(prev_per_gram * 1000 * SILVER_DOMESTIC_PREMIUM, 0) if prev_per_gram > 0 else price_kg
                change = round(price_kg - prev_kg, 0)
                change_pct = round(change / prev_kg * 100, 2) if prev_kg > 0 else 0
                result["silver"] = {
                    "price": price_kg, "change": change, "change_percent": change_pct,
                    "prev_close": prev_kg, "source": "goldapi.io"
                }
                logger.info(f"GoldAPI: Silver 1Kg = ₹{price_kg:,.0f}")
    except Exception as e:
        logger.warning(f"GoldAPI silver fetch failed: {e}")

    return result


def _fetch_yfinance_data() -> list:
    """Synchronous yfinance fetch — runs in thread executor. Uses GoldAPI for gold/silver when available."""
    results = []

    # Try GoldAPI first for more accurate gold/silver prices
    goldapi_data = _fetch_goldapi_prices()

    try:
        tickers = yf.Tickers("^NSEI ^BSESN ^NSEBANK GC=F SI=F INR=X")
        usd_inr = 87.0
        try:
            fx = tickers.tickers["INR=X"].fast_info
            usd_inr = float(fx.last_price) if fx.last_price else 87.0
        except Exception:
            pass

        configs = [
            {"key": "nifty_50", "name": "Nifty 50", "yf": "^NSEI", "type": "index"},
            {"key": "sensex", "name": "SENSEX", "yf": "^BSESN", "type": "index"},
            {"key": "nifty_bank", "name": "Nifty Bank", "yf": "^NSEBANK", "type": "index"},
            {"key": "gold_10g", "name": "Gold (10g)", "yf": "GC=F", "type": "gold"},
            {"key": "silver_1kg", "name": "Silver (1Kg)", "yf": "SI=F", "type": "silver"},
        ]

        for cfg in configs:
            try:
                # Use GoldAPI data for gold/silver if available
                if cfg["type"] == "gold" and "gold" in goldapi_data:
                    gd = goldapi_data["gold"]
                    results.append({"key": cfg["key"], "name": cfg["name"], "price": gd["price"],
                                    "change": gd["change"], "change_percent": gd["change_percent"],
                                    "prev_close": gd["prev_close"]})
                    continue
                if cfg["type"] == "silver" and "silver" in goldapi_data:
                    sd = goldapi_data["silver"]
                    results.append({"key": cfg["key"], "name": cfg["name"], "price": sd["price"],
                                    "change": sd["change"], "change_percent": sd["change_percent"],
                                    "prev_close": sd["prev_close"]})
                    continue

                # Fallback to yfinance for gold/silver or always for indices
                t = tickers.tickers[cfg["yf"]]
                info = t.fast_info
                last = float(info.last_price) if info.last_price else 0
                prev = float(info.previous_close) if info.previous_close else 0

                if cfg["type"] == "gold":
                    price_per_gram = (last / TROY_OZ_TO_GRAMS) * usd_inr * GOLD_DOMESTIC_PREMIUM
                    price = round(price_per_gram * 10, 0)
                    prev_price_per_gram = (prev / TROY_OZ_TO_GRAMS) * usd_inr * GOLD_DOMESTIC_PREMIUM
                    prev_price = round(prev_price_per_gram * 10, 0)
                    change = round(price - prev_price, 0)
                    change_pct = round((change / prev_price * 100), 2) if prev_price else 0
                    results.append({"key": cfg["key"], "name": cfg["name"], "price": price, "change": change, "change_percent": change_pct, "prev_close": prev_price})
                elif cfg["type"] == "silver":
                    price_per_kg = (last * (1000 / TROY_OZ_TO_GRAMS)) * usd_inr * SILVER_DOMESTIC_PREMIUM
                    price = round(price_per_kg, 0)
                    prev_per_kg = (prev * (1000 / TROY_OZ_TO_GRAMS)) * usd_inr * SILVER_DOMESTIC_PREMIUM
                    prev_price = round(prev_per_kg, 0)
                    change = round(price - prev_price, 0)
                    change_pct = round((change / prev_price * 100), 2) if prev_price else 0
                    results.append({"key": cfg["key"], "name": cfg["name"], "price": price, "change": change, "change_percent": change_pct, "prev_close": prev_price})
                else:
                    price = round(last, 2)
                    change = round(last - prev, 2)
                    change_pct = round((change / prev * 100), 2) if prev else 0
                    results.append({"key": cfg["key"], "name": cfg["name"], "price": price, "change": change, "change_percent": change_pct, "prev_close": round(prev, 2)})
            except Exception as e:
                logger.error(f"yfinance fetch error for {cfg['key']}: {e}")
    except Exception as e:
        logger.error(f"yfinance batch fetch error: {e}")
    return results

async def refresh_all_market_data():
    loop = asyncio.get_event_loop()
    results = await loop.run_in_executor(_yf_executor, _fetch_yfinance_data)
    now = datetime.now(timezone.utc).isoformat()
    for item in results:
        if item.get("price"):
            item["last_updated"] = now
            await db.market_data.update_one({"key": item["key"]}, {"$set": item}, upsert=True)
            logger.info(f"Market LIVE: {item['name']} = {item['price']}")
    return results

async def market_data_scheduler():
    last_refresh = None
    while True:
        utc_now = datetime.now(timezone.utc)
        ist_now = utc_now + timedelta(hours=5, minutes=30)
        ist_hhmm = ist_now.strftime("%H:%M")
        if ist_hhmm in REFRESH_TIMES_IST and last_refresh != ist_hhmm:
            logger.info(f"Scheduled market refresh at IST {ist_hhmm}")
            last_refresh = ist_hhmm
            await refresh_all_market_data()
        await asyncio.sleep(30)

async def seed_market_data():
    """Seed with live data from yfinance, or use accurate fallback values."""
    await db.market_data.delete_many({})
    logger.info("Fetching live market data for seed...")
    results = await refresh_all_market_data()
    if not results or len(results) < 5:
        logger.warning("Live fetch incomplete, using accurate fallback seed data")
        now = datetime.now(timezone.utc).isoformat()
        fallback = [
            {"key": "nifty_50", "name": "Nifty 50", "price": 25734.20, "change": 51.75, "change_percent": 0.20, "prev_close": 25682.45, "last_updated": now},
            {"key": "sensex", "name": "SENSEX", "price": 83389.45, "change": 112.39, "change_percent": 0.13, "prev_close": 83277.06, "last_updated": now},
            {"key": "nifty_bank", "name": "Nifty Bank", "price": 61126.95, "change": 133.10, "change_percent": 0.22, "prev_close": 60993.85, "last_updated": now},
            {"key": "gold_10g", "name": "Gold (10g)", "price": 154910.0, "change": -1530.0, "change_percent": -0.98, "prev_close": 156440.0, "last_updated": now},
            {"key": "silver_1kg", "name": "Silver (1Kg)", "price": 260000.0, "change": -8000.0, "change_percent": -2.99, "prev_close": 268000.0, "last_updated": now},
        ]
        for s in fallback:
            await db.market_data.update_one({"key": s["key"]}, {"$set": s}, upsert=True)
    logger.info("Market data seeded successfully")

@api_router.get("/market-data")
async def get_market_data():
    """Return market data, refreshing from yfinance if stale (>2 min)."""
    data = await db.market_data.find({}, {"_id": 0}).to_list(10)

    # Check if data is stale
    is_stale = True
    if data:
        last_updated = data[0].get("last_updated", "")
        if last_updated:
            try:
                from dateutil.parser import parse as parse_date
                updated_dt = parse_date(last_updated)
                if updated_dt.tzinfo is None:
                    updated_dt = updated_dt.replace(tzinfo=timezone.utc)
                age_seconds = (datetime.now(timezone.utc) - updated_dt).total_seconds()
                is_stale = age_seconds > 120  # 2 minutes
            except Exception:
                pass

    if is_stale:
        logger.info("Market data stale, fetching live prices...")
        fresh = await refresh_all_market_data()
        if fresh:
            data = await db.market_data.find({}, {"_id": 0}).to_list(10)

    return data

@api_router.post("/market-data/refresh")
async def trigger_market_refresh(user=Depends(get_current_user)):
    asyncio.create_task(refresh_all_market_data())
    return {"message": "Market data refresh triggered"}

# ══════════════════════════════════════
#  PORTFOLIO OVERVIEW
# ══════════════════════════════════════

# Annual return assumptions by category
CATEGORY_RETURNS = {
    "Stocks": "nifty", "SIP": "nifty", "Mutual Funds": "nifty", "ETFs": "nifty", "ELSS": "nifty",
    "Gold": "gold", "Sovereign Gold Bond": "gold",
    "Silver": "silver",
    "PPF": 0.071, "EPF": 0.0815, "NPS": 0.10, "FD": 0.07, "Fixed Deposit": 0.07,
    "Bonds": 0.075, "ULIP": 0.08, "Real Estate": 0.08, "Crypto": 0.0,
}

def _compute_portfolio_values(transactions: list, market_data: dict) -> dict:
    """Compute estimated current value for each investment category."""
    from datetime import date as dt_date
    today = dt_date.today()
    categories = {}
    for txn in transactions:
        cat = txn["category"]
        if cat not in categories:
            categories[cat] = {"invested": 0, "current_value": 0, "transactions": 0}
        amount = txn["amount"]
        categories[cat]["invested"] += amount
        categories[cat]["transactions"] += 1
        try:
            parts = txn["date"].split("-")
            txn_date = dt_date(int(parts[0]), int(parts[1]), int(parts[2]))
        except Exception:
            txn_date = today
        days_held = max((today - txn_date).days, 0)
        return_type = CATEGORY_RETURNS.get(cat, 0.0)
        if return_type == "nifty":
            nifty_now = market_data.get("nifty_50", {}).get("price", 0)
            nifty_prev = market_data.get("nifty_50", {}).get("prev_close", nifty_now)
            if nifty_now and nifty_prev:
                daily_return = (nifty_now / nifty_prev - 1) if nifty_prev else 0
                estimated_return = daily_return * days_held * 0.6
                categories[cat]["current_value"] += amount * (1 + estimated_return)
            else:
                categories[cat]["current_value"] += amount
        elif return_type == "gold":
            gold_now = market_data.get("gold_10g", {}).get("price", 0)
            gold_prev = market_data.get("gold_10g", {}).get("prev_close", gold_now)
            if gold_now and gold_prev:
                daily_return = (gold_now / gold_prev - 1) if gold_prev else 0
                estimated_return = daily_return * days_held * 0.5
                categories[cat]["current_value"] += amount * (1 + estimated_return)
            else:
                categories[cat]["current_value"] += amount
        elif return_type == "silver":
            silver_now = market_data.get("silver_1kg", {}).get("price", 0)
            silver_prev = market_data.get("silver_1kg", {}).get("prev_close", silver_now)
            if silver_now and silver_prev:
                daily_return = (silver_now / silver_prev - 1) if silver_prev else 0
                estimated_return = daily_return * days_held * 0.5
                categories[cat]["current_value"] += amount * (1 + estimated_return)
            else:
                categories[cat]["current_value"] += amount
        elif isinstance(return_type, (int, float)):
            annual_rate = return_type
            pro_rated = annual_rate * (days_held / 365)
            categories[cat]["current_value"] += amount * (1 + pro_rated)
        else:
            categories[cat]["current_value"] += amount
    for cat in categories:
        categories[cat]["invested"] = round(categories[cat]["invested"], 2)
        categories[cat]["current_value"] = round(categories[cat]["current_value"], 2)
        inv = categories[cat]["invested"]
        cur = categories[cat]["current_value"]
        categories[cat]["gain_loss"] = round(cur - inv, 2)
        categories[cat]["gain_loss_pct"] = round(((cur - inv) / inv * 100), 2) if inv else 0
    return categories

@api_router.get("/portfolio-overview")
async def get_portfolio_overview(user=Depends(get_current_user)):
    transactions = await db.transactions.find(
        {"user_id": user["id"], "type": "investment"}, {"_id": 0}
    ).to_list(1000)
    mkt_list = await db.market_data.find({}, {"_id": 0}).to_list(10)
    market_data = {m["key"]: m for m in mkt_list}
    loop = asyncio.get_event_loop()
    categories = await loop.run_in_executor(
        _yf_executor, _compute_portfolio_values, transactions, market_data
    )

    # Merge holdings data into portfolio categories
    # Normalize category names to avoid duplicates
    CATEGORY_NORMALIZE = {"Stocks": "Stock", "Fixed Deposit": "FD", "Mutual Funds": "Mutual Fund"}
    normalized_categories = {}
    for cat_key, data in categories.items():
        norm_key = CATEGORY_NORMALIZE.get(cat_key, cat_key)
        if norm_key not in normalized_categories:
            normalized_categories[norm_key] = {"invested": 0, "current_value": 0, "gain_loss": 0, "gain_loss_pct": 0, "transactions": 0}
        normalized_categories[norm_key]["invested"] += data["invested"]
        normalized_categories[norm_key]["current_value"] += data["current_value"]
        normalized_categories[norm_key]["transactions"] += data["transactions"]
    categories = normalized_categories

    holdings_cursor = db.holdings.find({"user_id": user["id"]})
    holdings_list = []
    async for doc in holdings_cursor:
        doc["id"] = str(doc.pop("_id"))
        holdings_list.append(doc)
    if holdings_list:
        tickers = list(set(h["ticker"] for h in holdings_list if h.get("ticker")))
        prices = await loop.run_in_executor(_yf_executor, _fetch_live_prices, tickers)
        for h in holdings_list:
            cat = h.get("category", "Stock")
            # Map holding categories to portfolio categories
            cat_key = cat
            if cat in ("Mutual Fund", "ETF"):
                cat_key = cat
            
            # Use stored invested_value and current_value from CAS if available
            stored_invested = h.get("invested_value", 0)
            stored_current = h.get("current_value", 0)
            
            if stored_invested > 0 and stored_current > 0:
                # Use eCAS values directly
                invested = stored_invested
                current_value = stored_current
            else:
                # Calculate from buy_price for manual entries
                invested = h["quantity"] * h["buy_price"]
                current_price = h["buy_price"]
                if h.get("ticker") and h["ticker"] in prices:
                    current_price = prices[h["ticker"]]["price"]
                current_value = h["quantity"] * current_price
            
            if cat_key not in categories:
                categories[cat_key] = {"invested": 0, "current_value": 0, "gain_loss": 0, "gain_loss_pct": 0, "transactions": 0}
            categories[cat_key]["invested"] = round(categories[cat_key]["invested"] + invested, 2)
            categories[cat_key]["current_value"] = round(categories[cat_key]["current_value"] + current_value, 2)
            categories[cat_key]["transactions"] += 1
        # Recalculate gain/loss for each category
        for cat_key in categories:
            inv = categories[cat_key]["invested"]
            cur = categories[cat_key]["current_value"]
            categories[cat_key]["gain_loss"] = round(cur - inv, 2)
            categories[cat_key]["gain_loss_pct"] = round((cur - inv) / inv * 100, 2) if inv else 0

    total_invested = sum(c["invested"] for c in categories.values())
    total_current = sum(c["current_value"] for c in categories.values())
    total_gain = round(total_current - total_invested, 2)
    total_gain_pct = round((total_gain / total_invested * 100), 2) if total_invested else 0
    breakdown = []
    for cat, data in sorted(categories.items(), key=lambda x: x[1]["invested"], reverse=True):
        breakdown.append({
            "category": cat,
            "invested": data["invested"],
            "current_value": data["current_value"],
            "gain_loss": data["gain_loss"],
            "gain_loss_pct": data["gain_loss_pct"],
            "transactions": data["transactions"],
        })
    return {
        "total_invested": total_invested,
        "total_current_value": round(total_current, 2),
        "total_gain_loss": total_gain,
        "total_gain_loss_pct": total_gain_pct,
        "categories": breakdown,
        "last_updated": datetime.now(timezone.utc).isoformat(),
    }

# ══════════════════════════════════════
#  HOLDINGS (Manual + CAS Upload)
# ══════════════════════════════════════

from fastapi import File, UploadFile, Form
import pdfplumber
import re
import io
from bson import ObjectId

class HoldingCreate(BaseModel):
    name: str
    ticker: str = ""
    isin: str = ""
    category: str = "Stock"
    quantity: float
    buy_price: float
    buy_date: str = ""

def _fetch_live_prices(tickers: list[str]) -> dict:
    """Fetch live prices for a list of tickers via yfinance."""
    result = {}
    if not tickers:
        return result
    try:
        batch = yf.Tickers(" ".join(tickers))
        for t_str in tickers:
            try:
                info = batch.tickers[t_str].fast_info
                price = float(info.last_price) if info.last_price else 0
                prev = float(info.previous_close) if info.previous_close else 0
                result[t_str] = {"price": price, "prev_close": prev}
            except Exception:
                pass
    except Exception as e:
        logger.error(f"Batch price fetch error: {e}")
    return result

def _search_ticker(name: str, category: str) -> str:
    """Try to find a Yahoo Finance ticker for a given name."""
    try:
        suffix = ".NS"
        clean = re.sub(r'[^a-zA-Z0-9 ]', '', name).strip()
        search = yf.Ticker(clean.split()[0].upper() + suffix)
        if search.fast_info.last_price:
            return clean.split()[0].upper() + suffix
    except Exception:
        pass
    return ""

@api_router.post("/holdings")
async def add_holding(holding: HoldingCreate, user=Depends(get_current_user)):
    doc = {
        "user_id": user["id"],
        "name": holding.name,
        "ticker": holding.ticker,
        "isin": holding.isin,
        "category": holding.category,
        "quantity": holding.quantity,
        "buy_price": holding.buy_price,
        "buy_date": holding.buy_date or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "source": "manual",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    result = await db.holdings.insert_one(doc)
    doc["id"] = str(result.inserted_id)
    doc.pop("_id", None)
    return doc

@api_router.get("/holdings")
async def get_holdings(user=Depends(get_current_user)):
    holdings = []
    cursor = db.holdings.find({"user_id": user["id"]})
    async for doc in cursor:
        doc["id"] = str(doc.pop("_id"))
        holdings.append(doc)
    return holdings

@api_router.get("/holdings/live")
async def get_holdings_live(user=Depends(get_current_user)):
    holdings = []
    cursor = db.holdings.find({"user_id": user["id"]})
    async for doc in cursor:
        doc["id"] = str(doc.pop("_id"))
        holdings.append(doc)
    tickers = list(set(h["ticker"] for h in holdings if h.get("ticker")))
    loop = asyncio.get_event_loop()
    prices = await loop.run_in_executor(_yf_executor, _fetch_live_prices, tickers)
    result = []
    for h in holdings:
        # For CAS-uploaded holdings, use stored invested_value and current_value
        # For manual entries without these values, calculate from buy_price
        stored_invested = h.get("invested_value", 0)
        stored_current = h.get("current_value", 0)
        
        if stored_invested > 0 and stored_current > 0:
            # Use eCAS values directly
            invested = stored_invested
            current_value = stored_current
            current_price = current_value / h["quantity"] if h["quantity"] > 0 else 0
        else:
            # Calculate from buy_price (manual entries)
            invested = h["quantity"] * h["buy_price"]
            current_price = 0
            if h.get("ticker") and h["ticker"] in prices:
                current_price = prices[h["ticker"]]["price"]
            else:
                current_price = h["buy_price"]
            current_value = h["quantity"] * current_price
        
        gain = current_value - invested
        gain_pct = round((gain / invested * 100), 2) if invested else 0
        result.append({
            **{k: v for k, v in h.items()},
            "current_price": round(current_price, 2),
            "invested_value": round(invested, 2),
            "current_value": round(current_value, 2),
            "gain_loss": round(gain, 2),
            "gain_loss_pct": gain_pct,
        })
    total_invested = sum(r["invested_value"] for r in result)
    total_current = sum(r["current_value"] for r in result)
    total_gain = round(total_current - total_invested, 2)
    return {
        "holdings": result,
        "summary": {
            "total_invested": round(total_invested, 2),
            "total_current": round(total_current, 2),
            "total_gain_loss": total_gain,
            "total_gain_loss_pct": round((total_gain / total_invested * 100), 2) if total_invested else 0,
            "count": len(result),
        },
    }

@api_router.delete("/holdings/clear-all")
async def clear_all_holdings(user=Depends(get_current_user)):
    """Clear all holdings for the user. Must be defined before /holdings/{holding_id}."""
    result = await db.holdings.delete_many({"user_id": user["id"]})
    return {"message": f"Deleted {result.deleted_count} holdings"}

@api_router.put("/holdings/{holding_id}")
async def update_holding(holding_id: str, holding: HoldingCreate, user=Depends(get_current_user)):
    update = {
        "name": holding.name, "ticker": holding.ticker, "isin": holding.isin,
        "category": holding.category, "quantity": holding.quantity,
        "buy_price": holding.buy_price, "buy_date": holding.buy_date,
    }
    result = await db.holdings.update_one(
        {"_id": ObjectId(holding_id), "user_id": user["id"]}, {"$set": update}
    )
    if result.matched_count == 0:
        raise HTTPException(404, "Holding not found")
    return {"message": "Updated"}

@api_router.delete("/holdings/{holding_id}")
async def delete_holding(holding_id: str, user=Depends(get_current_user)):
    result = await db.holdings.delete_one({"_id": ObjectId(holding_id), "user_id": user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(404, "Holding not found")
    return {"message": "Deleted"}

def _parse_cas_pdf(pdf_bytes: bytes, password: str = "") -> list:
    """Parse NSDL/CDSL CAS PDF to extract holdings with proper invested amount and current value."""
    import pikepdf
    holdings = []
    decrypted_bytes = pdf_bytes
    
    # First, try to decrypt using pikepdf if password-protected
    try:
        pdf_stream = io.BytesIO(pdf_bytes)
        with pikepdf.open(pdf_stream, password=password or "") as pike_pdf:
            decrypted_stream = io.BytesIO()
            pike_pdf.save(decrypted_stream)
            decrypted_bytes = decrypted_stream.getvalue()
            logger.info(f"PDF decrypted successfully, {len(pike_pdf.pages)} pages")
    except pikepdf.PasswordError:
        logger.error("PDF password is incorrect")
        raise HTTPException(400, "Incorrect PDF password. Please check and try again.")
    except Exception as e:
        logger.warning(f"pikepdf decryption warning: {e}")
    
    try:
        with pdfplumber.open(io.BytesIO(decrypted_bytes)) as pdf:
            logger.info(f"Parsing PDF with {len(pdf.pages)} pages")
            all_text = ""
            
            for page_num, page in enumerate(pdf.pages):
                page_text = page.extract_text() or ""
                all_text += page_text + "\n"
                logger.info(f"Page {page_num + 1} text length: {len(page_text)}")
            
            # Parse the text line by line to find holdings
            lines = all_text.split("\n")
            
            # ISIN pattern for Indian securities
            isin_pattern = re.compile(r'(INE[A-Z0-9]{7,10}|INF[A-Z0-9]{7,10})')
            
            # Pattern to extract financial data from CDSL eCAS:
            # cost_value units_balance DD-Mon-YYYY nav market_value
            # Example: 42,000.000 272.234 16-Feb-2026 194.729 53,011.85
            financial_pattern = re.compile(
                r'([\d,]+\.\d+)\s+([\d,]+\.\d+)\s+\d{1,2}-[A-Za-z]{3}-\d{4}\s+([\d,]+\.\d+)\s+([\d,]+\.\d+)'
            )
            
            for line in lines:
                isin_match = isin_pattern.search(line)
                if not isin_match:
                    continue
                    
                isin = isin_match.group(1)
                logger.info(f"Processing line with ISIN {isin}")
                
                # Extract financial data using the precise pattern
                fin_match = financial_pattern.search(line)
                if not fin_match:
                    logger.warning(f"Could not extract financial data for {isin}")
                    continue
                
                cost_value = float(fin_match.group(1).replace(",", ""))  # Invested amount
                unit_balance = float(fin_match.group(2).replace(",", ""))  # Quantity
                nav = float(fin_match.group(3).replace(",", ""))  # NAV
                market_value = float(fin_match.group(4).replace(",", ""))  # Current value
                
                # Calculate buy price (average cost per unit)
                buy_price = cost_value / unit_balance if unit_balance > 0 else 0
                
                # Extract scheme name using multiple patterns
                scheme_name = ""
                
                # Pattern 1: Text after "- " that contains common fund keywords
                name_match = re.search(r'[-–]\s*([A-Za-z][A-Za-z\s\(\)\-]+(?:Fund|Growth|Plan|Option|Cap)[^\d]*)', line)
                if name_match:
                    scheme_name = name_match.group(1).strip()
                
                # Pattern 2: For all caps names like "NIPPON INDIA SMALL CAP"
                if not scheme_name or len(scheme_name) < 5:
                    # Look for all-caps fund names after scheme code
                    caps_match = re.search(r'[A-Z0-9]{6,}\s*[-–]?\s*([A-Z][A-Z\s]+(?:FUND|CAP|GROWTH)[^\d]*)', line, re.IGNORECASE)
                    if caps_match:
                        scheme_name = caps_match.group(1).strip().title()  # Convert to title case
                
                # Pattern 3: Look for name after ISIN in the format "CODE - Name"
                if not scheme_name or len(scheme_name) < 5:
                    after_isin = line.split(isin)[-1] if isin in line else ""
                    code_name_match = re.search(r'^[\s/0-9]*[A-Z0-9]+\s*[-–]\s*([A-Za-z][A-Za-z\s]+)', after_isin)
                    if code_name_match:
                        scheme_name = code_name_match.group(1).strip()
                
                # Clean up the extracted name
                if scheme_name:
                    scheme_name = re.sub(r'\s*[-–]\s*$', '', scheme_name)  # Remove trailing dash
                    scheme_name = re.sub(r'\(Non-Demat\)', '', scheme_name).strip()
                    scheme_name = re.sub(r'\(formerly.*?\)', '', scheme_name).strip()
                    scheme_name = re.sub(r'\s+', ' ', scheme_name).strip()  # Normalize whitespace
                    # Remove any numeric suffixes
                    scheme_name = re.sub(r'\s*\d+[\.,]?\d*\s*$', '', scheme_name).strip()
                
                # Fallback to ISIN if no name found
                if not scheme_name or len(scheme_name) < 5:
                    scheme_name = f"Fund {isin}"
                
                # Determine category based on ISIN prefix
                is_mf = isin.startswith("INF")
                category = "Mutual Fund" if is_mf else "Stock"
                
                # Only add if we have meaningful data
                if unit_balance > 0 and cost_value > 0:
                    # Check for duplicates
                    existing = next((h for h in holdings if h["isin"] == isin), None)
                    if not existing:
                        holding = {
                            "name": scheme_name[:100],
                            "ticker": "",
                            "isin": isin,
                            "category": category,
                            "quantity": round(unit_balance, 4),
                            "buy_price": round(buy_price, 4),
                            "invested_value": round(cost_value, 2),
                            "current_value": round(market_value, 2),
                            "nav": round(nav, 4),
                        }
                        holdings.append(holding)
                        logger.info(f"✓ Parsed: {scheme_name[:40]} | Invested: ₹{cost_value:,.2f} | Current: ₹{market_value:,.2f} | Qty: {unit_balance:.3f}")
            
            # Log summary
            logger.info(f"CAS parsing complete: Found {len(holdings)} holdings")
            if len(holdings) == 0:
                logger.warning(f"No holdings found. Sample text: {all_text[:1000]}")
            else:
                total_invested = sum(h.get("invested_value", 0) for h in holdings)
                total_current = sum(h.get("current_value", 0) for h in holdings)
                gain = total_current - total_invested
                gain_pct = (gain / total_invested * 100) if total_invested > 0 else 0
                logger.info(f"Total Invested: ₹{total_invested:,.2f} | Total Current: ₹{total_current:,.2f} | Gain: ₹{gain:,.2f} ({gain_pct:+.2f}%)")
                
    except Exception as e:
        error_msg = str(e) if str(e) else "Unknown PDF parsing error"
        logger.error(f"CAS PDF parse error: {error_msg}")
        if "password" in error_msg.lower() or "encrypted" in error_msg.lower() or "decrypt" in error_msg.lower():
            raise HTTPException(400, "PDF is password-protected. Please enter the correct password.")
        if "invalid" in error_msg.lower() or "corrupt" in error_msg.lower():
            raise HTTPException(400, "PDF file appears to be corrupted or invalid.")
        raise HTTPException(400, f"Failed to parse PDF: {error_msg}")
    return holdings

@api_router.post("/holdings/upload-cas")
async def upload_cas(
    file: UploadFile = File(...),
    password: str = Form(""),
    user=Depends(get_current_user),
):
    if not file.filename or not file.filename.lower().endswith(".pdf"):
        raise HTTPException(400, "Please upload a PDF file")
    pdf_bytes = await file.read()
    if len(pdf_bytes) > 10 * 1024 * 1024:
        raise HTTPException(400, "File too large (max 10MB)")
    parsed = _parse_cas_pdf(pdf_bytes, password)
    if not parsed:
        raise HTTPException(400, "No holdings found in the PDF. Please check the file format or password.")
    # ALWAYS delete existing CAS-uploaded holdings for this user before importing
    # This prevents duplicates when re-uploading the same CAS statement
    deleted = await db.holdings.delete_many({
        "user_id": user["id"],
        "source": "cas_upload"
    })
    logger.info(f"Cleared {deleted.deleted_count} existing CAS holdings before import")
    
    created = []
    now = datetime.now(timezone.utc).isoformat()
    for h in parsed:
        doc = {
            "user_id": user["id"],
            "name": h["name"],
            "ticker": h.get("ticker", ""),
            "isin": h["isin"],
            "category": h["category"],
            "quantity": h["quantity"],
            "buy_price": h.get("buy_price", 0),
            "invested_value": h.get("invested_value", 0),
            "current_value": h.get("current_value", 0),
            "nav": h.get("nav", 0),
            "buy_date": "",
            "source": "cas_upload",
            "created_at": now,
        }
        result = await db.holdings.insert_one(doc)
        doc["id"] = str(result.inserted_id)
        doc.pop("_id", None)
        created.append(doc)
    
    # Calculate totals for response
    total_invested = sum(h.get("invested_value", 0) for h in parsed)
    total_current = sum(h.get("current_value", 0) for h in parsed)
    gain = total_current - total_invested
    gain_pct = (gain / total_invested * 100) if total_invested > 0 else 0
    
    return {
        "message": f"Replaced {deleted.deleted_count} → Imported {len(created)} holdings from CAS",
        "holdings": created,
        "summary": {
            "total_invested": round(total_invested, 2),
            "total_current": round(total_current, 2),
            "gain_loss": round(gain, 2),
            "gain_loss_pct": round(gain_pct, 2),
        }
    }

async def seed_demo_data():
    # Check if demo users already exist
    demo1 = await db.users.find_one({"email": "rajesh@visor.demo"}, {"_id": 0})
    if demo1:
        # Ensure existing demo users have encryption keys
        if not demo1.get("encryption_key"):
            dek = generate_user_dek()
            pan_enc = encrypt_field(demo1.get("pan", ""), dek) if demo1.get("pan") and not demo1.get("pan", "").startswith("ENC:") else demo1.get("pan", "")
            aadhaar_enc = encrypt_field(demo1.get("aadhaar", ""), dek) if demo1.get("aadhaar") and not demo1.get("aadhaar", "").startswith("ENC:") else demo1.get("aadhaar", "")
            await db.users.update_one({"email": "rajesh@visor.demo"}, {"$set": {"encryption_key": dek, "pan": pan_enc, "aadhaar": aadhaar_enc}})
            logger.info("Migrated demo user rajesh with encryption key")
        demo2 = await db.users.find_one({"email": "priya@visor.demo"}, {"_id": 0})
        if demo2 and not demo2.get("encryption_key"):
            dek2 = generate_user_dek()
            pan_enc2 = encrypt_field(demo2.get("pan", ""), dek2) if demo2.get("pan") and not demo2.get("pan", "").startswith("ENC:") else demo2.get("pan", "")
            aadhaar_enc2 = encrypt_field(demo2.get("aadhaar", ""), dek2) if demo2.get("aadhaar") and not demo2.get("aadhaar", "").startswith("ENC:") else demo2.get("aadhaar", "")
            await db.users.update_one({"email": "priya@visor.demo"}, {"$set": {"encryption_key": dek2, "pan": pan_enc2, "aadhaar": aadhaar_enc2}})
            logger.info("Migrated demo user priya with encryption key")
        logger.info("Demo data already exists, skipping seed")
        return

    logger.info("Seeding demo data...")
    now = datetime.now(timezone.utc).isoformat()

    # Demo User 1: Rajesh Kumar
    user1_id = str(uuid.uuid4())
    user1_dek = generate_user_dek()
    await db.users.insert_one({
        "id": user1_id,
        "email": "rajesh@visor.demo",
        "password": hash_password("Demo@123"),
        "full_name": "Rajesh Kumar",
        "dob": "1995-05-15",
        "pan": encrypt_field("ABCDE1234F", user1_dek),
        "aadhaar": encrypt_field("123456789012", user1_dek),
        "encryption_key": user1_dek,
        "created_at": now,
    })

    # Demo User 2: Priya Sharma
    user2_id = str(uuid.uuid4())
    user2_dek = generate_user_dek()
    await db.users.insert_one({
        "id": user2_id,
        "email": "priya@visor.demo",
        "password": hash_password("Demo@456"),
        "full_name": "Priya Sharma",
        "dob": "1990-08-22",
        "pan": encrypt_field("FGHIJ5678K", user2_dek),
        "aadhaar": encrypt_field("987654321098", user2_dek),
        "encryption_key": user2_dek,
        "created_at": now,
    })

    # Transactions for Rajesh
    rajesh_txns = [
        {"type": "income", "amount": 85000, "category": "Salary", "description": "Monthly Salary - TCS", "date": "2026-02-01"},
        {"type": "income", "amount": 85000, "category": "Salary", "description": "Monthly Salary - TCS", "date": "2026-01-01"},
        {"type": "income", "amount": 12000, "category": "Freelance", "description": "UI Design Project", "date": "2026-01-15"},
        {"type": "expense", "amount": 18000, "category": "Rent", "description": "Monthly Rent - Koramangala", "date": "2026-02-01"},
        {"type": "expense", "amount": 18000, "category": "Rent", "description": "Monthly Rent - Koramangala", "date": "2026-01-01"},
        {"type": "expense", "amount": 6500, "category": "Groceries", "description": "BigBasket + D-Mart", "date": "2026-02-05"},
        {"type": "expense", "amount": 5800, "category": "Groceries", "description": "Monthly Groceries", "date": "2026-01-08"},
        {"type": "expense", "amount": 3200, "category": "Food", "description": "Swiggy & Zomato", "date": "2026-02-10"},
        {"type": "expense", "amount": 4100, "category": "Food", "description": "Dining Out + Delivery", "date": "2026-01-12"},
        {"type": "expense", "amount": 2500, "category": "Transport", "description": "Uber + Metro", "date": "2026-02-08"},
        {"type": "expense", "amount": 2200, "category": "Transport", "description": "Ola + Auto", "date": "2026-01-10"},
        {"type": "expense", "amount": 4500, "category": "Shopping", "description": "Amazon - Electronics", "date": "2026-01-20"},
        {"type": "expense", "amount": 1500, "category": "Utilities", "description": "Electricity + Internet", "date": "2026-02-03"},
        {"type": "expense", "amount": 1500, "category": "Utilities", "description": "Electricity + Internet", "date": "2026-01-03"},
        {"type": "expense", "amount": 2000, "category": "Entertainment", "description": "Netflix + Spotify + Books", "date": "2026-02-07"},
        {"type": "expense", "amount": 3000, "category": "Health", "description": "Gym + Medicines", "date": "2026-01-25"},
        {"type": "investment", "amount": 10000, "category": "SIP", "description": "Axis Bluechip SIP", "date": "2026-02-05"},
        {"type": "investment", "amount": 10000, "category": "SIP", "description": "Axis Bluechip SIP", "date": "2026-01-05"},
        {"type": "investment", "amount": 5000, "category": "PPF", "description": "PPF Contribution", "date": "2026-02-01"},
        {"type": "investment", "amount": 5000, "category": "PPF", "description": "PPF Contribution", "date": "2026-01-01"},
        {"type": "investment", "amount": 8000, "category": "Stocks", "description": "Reliance + HDFC Bank", "date": "2026-01-18"},
    ]

    for txn in rajesh_txns:
        await db.transactions.insert_one({
            "id": str(uuid.uuid4()),
            "user_id": user1_id,
            **txn,
            "created_at": now,
        })

    # Transactions for Priya
    priya_txns = [
        {"type": "income", "amount": 120000, "category": "Salary", "description": "Monthly Salary - Flipkart", "date": "2026-02-01"},
        {"type": "income", "amount": 120000, "category": "Salary", "description": "Monthly Salary - Flipkart", "date": "2026-01-01"},
        {"type": "income", "amount": 25000, "category": "Bonus", "description": "Quarterly Bonus", "date": "2026-01-15"},
        {"type": "expense", "amount": 25000, "category": "Rent", "description": "Monthly Rent - Indiranagar", "date": "2026-02-01"},
        {"type": "expense", "amount": 25000, "category": "Rent", "description": "Monthly Rent - Indiranagar", "date": "2026-01-01"},
        {"type": "expense", "amount": 8000, "category": "Groceries", "description": "Organic Store + Zepto", "date": "2026-02-04"},
        {"type": "expense", "amount": 7500, "category": "Groceries", "description": "Monthly Groceries", "date": "2026-01-06"},
        {"type": "expense", "amount": 5000, "category": "Food", "description": "Restaurants + Cafes", "date": "2026-02-09"},
        {"type": "expense", "amount": 6000, "category": "Shopping", "description": "Myntra + Nykaa", "date": "2026-01-22"},
        {"type": "expense", "amount": 3500, "category": "Transport", "description": "Uber Premier", "date": "2026-02-06"},
        {"type": "expense", "amount": 2000, "category": "Utilities", "description": "Bills", "date": "2026-02-02"},
        {"type": "expense", "amount": 15000, "category": "EMI", "description": "Car Loan EMI", "date": "2026-02-05"},
        {"type": "expense", "amount": 15000, "category": "EMI", "description": "Car Loan EMI", "date": "2026-01-05"},
        {"type": "expense", "amount": 4000, "category": "Health", "description": "Yoga + Health Insurance", "date": "2026-01-20"},
        {"type": "investment", "amount": 15000, "category": "Mutual Funds", "description": "HDFC Mid Cap SIP", "date": "2026-02-05"},
        {"type": "investment", "amount": 15000, "category": "Mutual Funds", "description": "HDFC Mid Cap SIP", "date": "2026-01-05"},
        {"type": "investment", "amount": 10000, "category": "FD", "description": "SBI FD", "date": "2026-01-10"},
        {"type": "investment", "amount": 5000, "category": "Gold", "description": "Sovereign Gold Bond", "date": "2026-01-15"},
    ]

    for txn in priya_txns:
        await db.transactions.insert_one({
            "id": str(uuid.uuid4()),
            "user_id": user2_id,
            **txn,
            "created_at": now,
        })

    # Goals for Rajesh
    rajesh_goals = [
        {"title": "Emergency Fund", "target_amount": 300000, "current_amount": 185000, "deadline": "2026-06-30", "category": "Safety"},
        {"title": "Goa Trip", "target_amount": 50000, "current_amount": 32000, "deadline": "2026-04-15", "category": "Travel"},
        {"title": "New Laptop", "target_amount": 80000, "current_amount": 45000, "deadline": "2026-08-01", "category": "Purchase"},
    ]
    for g in rajesh_goals:
        await db.goals.insert_one({"id": str(uuid.uuid4()), "user_id": user1_id, **g, "created_at": now})

    # Goals for Priya
    priya_goals = [
        {"title": "House Down Payment", "target_amount": 2000000, "current_amount": 850000, "deadline": "2027-12-31", "category": "Property"},
        {"title": "Europe Trip", "target_amount": 300000, "current_amount": 120000, "deadline": "2026-09-01", "category": "Travel"},
        {"title": "Emergency Fund", "target_amount": 500000, "current_amount": 380000, "deadline": "2026-06-30", "category": "Safety"},
    ]
    for g in priya_goals:
        await db.goals.insert_one({"id": str(uuid.uuid4()), "user_id": user2_id, **g, "created_at": now})

    # Create indexes
    await db.users.create_index("email", unique=True)
    await db.users.create_index("id", unique=True)
    await db.transactions.create_index("user_id")
    await db.goals.create_index("user_id")
    await db.chat_history.create_index("user_id")

    logger.info("Demo data seeded successfully!")

# ══════════════════════════════════════
#  AI FINANCIAL ADVISOR
# ══════════════════════════════════════

class ChatMessage(BaseModel):
    message: str
    calculator_type: Optional[str] = None  # emi, sip, portfolio, compound, etc.
    calculator_params: Optional[dict] = None

class ChatResponse(BaseModel):
    response: str
    calculator_result: Optional[dict] = None

# Financial Calculator Functions
def calculate_sip_returns(monthly_investment: float, annual_return: float, years: int) -> dict:
    """Calculate SIP returns with compound interest"""
    months = years * 12
    monthly_rate = annual_return / 12 / 100
    
    if monthly_rate == 0:
        future_value = monthly_investment * months
    else:
        future_value = monthly_investment * (((1 + monthly_rate) ** months - 1) / monthly_rate) * (1 + monthly_rate)
    
    total_invested = monthly_investment * months
    wealth_gained = future_value - total_invested
    
    return {
        "monthly_investment": monthly_investment,
        "time_period_years": years,
        "expected_return_rate": annual_return,
        "total_invested": round(total_invested, 2),
        "future_value": round(future_value, 2),
        "wealth_gained": round(wealth_gained, 2),
        "absolute_returns": round((wealth_gained / total_invested) * 100, 2),
    }

def calculate_compound_interest(principal: float, annual_rate: float, years: int, compounding: str = "yearly") -> dict:
    """Calculate compound interest"""
    n = {"yearly": 1, "half-yearly": 2, "quarterly": 4, "monthly": 12}.get(compounding, 1)
    rate = annual_rate / 100
    
    amount = principal * ((1 + rate/n) ** (n * years))
    interest = amount - principal
    
    return {
        "principal": principal,
        "annual_rate": annual_rate,
        "time_years": years,
        "compounding": compounding,
        "maturity_amount": round(amount, 2),
        "interest_earned": round(interest, 2),
        "effective_rate": round(((amount/principal) ** (1/years) - 1) * 100, 2),
    }

def calculate_loan_emi_details(principal: float, annual_rate: float, tenure_years: int) -> dict:
    """Calculate EMI with full breakdown"""
    tenure_months = tenure_years * 12
    monthly_rate = annual_rate / 12 / 100
    
    if monthly_rate == 0:
        emi = principal / tenure_months
    else:
        emi = principal * monthly_rate * ((1 + monthly_rate) ** tenure_months) / (((1 + monthly_rate) ** tenure_months) - 1)
    
    total_payment = emi * tenure_months
    total_interest = total_payment - principal
    
    return {
        "loan_amount": principal,
        "interest_rate": annual_rate,
        "tenure_years": tenure_years,
        "tenure_months": tenure_months,
        "monthly_emi": round(emi, 2),
        "total_payment": round(total_payment, 2),
        "total_interest": round(total_interest, 2),
        "interest_to_principal_ratio": round((total_interest / principal) * 100, 2),
    }

def calculate_portfolio_returns(investments: list, years: int) -> dict:
    """Calculate weighted portfolio returns"""
    total_invested = sum(inv.get("amount", 0) for inv in investments)
    weighted_return = sum(
        (inv.get("amount", 0) / total_invested) * inv.get("expected_return", 0) 
        for inv in investments if total_invested > 0
    )
    
    future_value = total_invested * ((1 + weighted_return/100) ** years)
    
    return {
        "total_invested": total_invested,
        "weighted_avg_return": round(weighted_return, 2),
        "time_years": years,
        "projected_value": round(future_value, 2),
        "projected_gain": round(future_value - total_invested, 2),
        "breakdown": investments,
    }

def calculate_tax_savings_80c(investments: dict) -> dict:
    """Calculate tax savings under Section 80C"""
    limit_80c = 150000
    total_claimed = min(sum(investments.values()), limit_80c)
    
    # Tax slabs for FY 2025-26 (New Regime)
    tax_saved_30_percent = total_claimed * 0.30  # Highest slab
    tax_saved_20_percent = total_claimed * 0.20
    
    return {
        "section_80c_limit": limit_80c,
        "investments": investments,
        "total_claimed": total_claimed,
        "remaining_limit": limit_80c - total_claimed,
        "tax_saved_30_slab": round(tax_saved_30_percent, 2),
        "tax_saved_20_slab": round(tax_saved_20_percent, 2),
        "recommendation": "Maximize 80C limit for optimal tax savings" if total_claimed < limit_80c else "80C limit fully utilized",
    }

def calculate_fire_number(monthly_expenses: float, withdrawal_rate: float = 4) -> dict:
    """Calculate FIRE (Financial Independence Retire Early) number"""
    annual_expenses = monthly_expenses * 12
    fire_number = annual_expenses * (100 / withdrawal_rate)
    
    return {
        "monthly_expenses": monthly_expenses,
        "annual_expenses": annual_expenses,
        "withdrawal_rate": withdrawal_rate,
        "fire_number": round(fire_number, 2),
        "explanation": f"You need ₹{round(fire_number/100000, 2)} lakhs to be financially independent at {withdrawal_rate}% withdrawal rate",
    }

FINANCIAL_ADVISOR_SYSTEM_PROMPT = """You are Visor, a friendly yet highly knowledgeable Indian Financial Advisor. You possess expertise equivalent to:
- Chartered Accountant (CA)
- Chartered Financial Analyst (CFA)
- Financial Risk Manager (FRM)
- ACCA qualified
- MBA in Finance from IIM

## Your Core Competencies:

### 1. Indian Tax Laws & Planning
- Income Tax Act 1961 - All sections (80C, 80D, 80CCD, 80G, 80E, 80EE, 80TTA, etc.)
- New Tax Regime vs Old Tax Regime comparison
- Capital Gains Tax (Short-term & Long-term)
- TDS provisions and compliance
- GST implications on investments
- Tax-saving investment instruments (ELSS, PPF, NPS, ULIP)
- HRA, LTA, Standard Deduction rules
- Advance Tax and TDS deadlines

### 2. Investment Knowledge
- Indian Equity Markets (NSE, BSE, NIFTY 50, Sensex)
- Mutual Funds (Equity, Debt, Hybrid, ELSS, Index, ETFs)
- Fixed Income (FDs, RDs, Bonds, NCDs, Government Securities)
- PPF, EPF, NPS, VPF
- Gold (Physical, SGB, Gold ETFs, Digital Gold)
- Silver, Copper, and other commodities (MCX)
- Real Estate investment considerations
- REITs and InvITs
- International investing (US stocks, Mutual Funds)

### 3. Financial Regulations
- SEBI regulations
- RBI guidelines
- IRDAI (Insurance)
- PFRDA (NPS/Pension)
- Banking regulations

### 4. Risk Management
- Risk profiling (Conservative, Moderate, Aggressive)
- Asset allocation strategies
- Diversification principles
- Emergency fund planning
- Insurance (Term, Health, ULIP analysis)

### 5. Financial Planning
- Goal-based investing
- Retirement planning (FIRE calculations)
- Education planning
- Marriage planning
- House purchase planning
- Estate planning basics

## Built-in Calculators You Can Use:
- **SIP Calculator**: Calculate SIP returns with compounding
- **EMI Calculator**: Loan EMI with amortization
- **Compound Interest**: FD/RD returns
- **Portfolio Returns**: Weighted average portfolio performance
- **Tax Savings (80C)**: Section 80C optimization
- **FIRE Calculator**: Financial Independence number

## User Context:
You have access to the user's financial data including:
- Income, Expenses, Investments
- Savings rate, EMI obligations
- Financial goals and progress
- Transaction history
- Age and risk profile

## Communication Style:
- Be friendly, educational, and supportive
- Explain complex concepts in simple Hindi-English (Hinglish) when appropriate
- Always provide specific numbers and calculations
- Give actionable recommendations
- Cite relevant sections/rules when discussing taxes
- Compare options with pros and cons
- Use Indian currency (₹) and Indian number system (lakhs, crores)

## Important Guidelines:
- Never provide specific stock tips or guarantee returns
- Always mention that past performance doesn't guarantee future returns
- Recommend consulting a registered financial advisor for large decisions
- Be transparent about risks involved in any investment
- Consider the user's specific situation before giving advice

When user asks for calculations, use the calculator tools provided and explain the results."""

@api_router.post("/ai-advisor/chat", response_model=ChatResponse)
async def chat_with_advisor(chat_msg: ChatMessage, user=Depends(get_current_user)):
    """Chat with the AI Financial Advisor"""
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
    except ImportError:
        raise HTTPException(status_code=500, detail="AI service not available")
    
    # Get user's financial context
    transactions = await db.transactions.find({"user_id": user["id"]}, {"_id": 0}).to_list(1000)
    goals = await db.goals.find({"user_id": user["id"]}, {"_id": 0}).to_list(100)
    loans = await db.loans.find({"user_id": user["id"]}, {"_id": 0}).to_list(100)
    assets = await db.fixed_assets.find({"user_id": user["id"]}, {"_id": 0}).to_list(100)
    
    # Calculate financial summary
    total_income = sum(t["amount"] for t in transactions if t["type"] == "income")
    total_expenses = sum(t["amount"] for t in transactions if t["type"] == "expense")
    total_investments = sum(t["amount"] for t in transactions if t["type"] == "investment")
    savings_rate = ((total_income - total_expenses) / total_income * 100) if total_income > 0 else 0
    
    # Calculate age from DOB
    user_age = None
    if user.get("dob"):
        try:
            dob = datetime.strptime(user["dob"], "%Y-%m-%d")
            user_age = (datetime.now() - dob).days // 365
        except:
            pass
    
    # Build financial context
    financial_context = f"""
## Current User Financial Profile:
- **Name**: {user.get('full_name', 'User')}
- **Age**: {user_age or 'Not provided'} years
- **Account Created**: {user.get('created_at', 'N/A')[:10] if user.get('created_at') else 'N/A'}

## Financial Summary (All Time):
- **Total Income**: ₹{total_income:,.0f}
- **Total Expenses**: ₹{total_expenses:,.0f}
- **Total Investments**: ₹{total_investments:,.0f}
- **Net Savings**: ₹{(total_income - total_expenses):,.0f}
- **Savings Rate**: {savings_rate:.1f}%
- **Investment Rate**: {(total_investments/total_income*100) if total_income > 0 else 0:.1f}%

## Investment Breakdown:
{chr(10).join([f"- {cat}: ₹{sum(t['amount'] for t in transactions if t['type']=='investment' and t['category']==cat):,.0f}" for cat in set(t['category'] for t in transactions if t['type']=='investment')])}

## Expense Categories:
{chr(10).join([f"- {cat}: ₹{sum(t['amount'] for t in transactions if t['type']=='expense' and t['category']==cat):,.0f}" for cat in set(t['category'] for t in transactions if t['type']=='expense')])}

## Active Loans:
{chr(10).join([f"- {l['name']}: ₹{l['principal_amount']:,.0f} at {l['interest_rate']}% ({l['tenure_months']} months)" for l in loans]) if loans else 'No active loans'}

## Fixed Assets:
{chr(10).join([f"- {a['name']}: ₹{a.get('current_value', a['purchase_value']):,.0f}" for a in assets]) if assets else 'No fixed assets recorded'}

## Financial Goals:
{chr(10).join([f"- {g['title']}: ₹{g['current_amount']:,.0f} / ₹{g['target_amount']:,.0f} ({(g['current_amount']/g['target_amount']*100):.0f}% complete)" for g in goals]) if goals else 'No goals set'}
"""

    # Handle calculator requests
    calculator_result = None
    if chat_msg.calculator_type and chat_msg.calculator_params:
        params = chat_msg.calculator_params
        try:
            if chat_msg.calculator_type == "sip":
                calculator_result = calculate_sip_returns(
                    params.get("monthly_investment", 10000),
                    params.get("annual_return", 12),
                    params.get("years", 10)
                )
            elif chat_msg.calculator_type == "emi":
                calculator_result = calculate_loan_emi_details(
                    params.get("principal", 5000000),
                    params.get("annual_rate", 8.5),
                    params.get("tenure_years", 20)
                )
            elif chat_msg.calculator_type == "compound":
                calculator_result = calculate_compound_interest(
                    params.get("principal", 100000),
                    params.get("annual_rate", 7),
                    params.get("years", 5),
                    params.get("compounding", "yearly")
                )
            elif chat_msg.calculator_type == "portfolio":
                calculator_result = calculate_portfolio_returns(
                    params.get("investments", []),
                    params.get("years", 10)
                )
            elif chat_msg.calculator_type == "tax_80c":
                calculator_result = calculate_tax_savings_80c(params.get("investments", {}))
            elif chat_msg.calculator_type == "fire":
                calculator_result = calculate_fire_number(
                    params.get("monthly_expenses", 50000),
                    params.get("withdrawal_rate", 4)
                )
        except Exception as e:
            logger.error(f"Calculator error: {e}")
    
    # Get or create chat session
    session_id = f"advisor_{user['id']}"
    
    # Load chat history
    history = await db.chat_history.find(
        {"user_id": user["id"], "type": "advisor"}
    ).sort("created_at", -1).limit(20).to_list(20)
    history.reverse()
    
    # Build message with context
    user_message_text = f"""
{financial_context}

---
**User Question**: {chat_msg.message}
"""
    
    if calculator_result:
        user_message_text += f"\n\n**Calculator Result ({chat_msg.calculator_type})**: {calculator_result}"
    
    try:
        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=session_id,
            system_message=FINANCIAL_ADVISOR_SYSTEM_PROMPT
        ).with_model("openai", "gpt-4o")
        
        # Note: History is already managed by session_id, we don't need to replay it
        # Just send the current message with context
        
        # Send message and get response
        response = await chat.send_message(UserMessage(text=user_message_text))
        
        # Store messages in database
        now = datetime.now(timezone.utc).isoformat()
        await db.chat_history.insert_one({
            "id": str(uuid.uuid4()),
            "user_id": user["id"],
            "type": "advisor",
            "role": "user",
            "content": chat_msg.message,
            "created_at": now,
        })
        await db.chat_history.insert_one({
            "id": str(uuid.uuid4()),
            "user_id": user["id"],
            "type": "advisor",
            "role": "assistant",
            "content": response,
            "created_at": now,
        })
        
        return ChatResponse(response=response, calculator_result=calculator_result)
        
    except Exception as e:
        logger.error(f"AI Advisor error: {e}")
        raise HTTPException(status_code=500, detail=f"AI service error: {str(e)}")

@api_router.get("/ai-advisor/history")
async def get_chat_history(user=Depends(get_current_user)):
    """Get chat history with AI advisor"""
    history = await db.chat_history.find(
        {"user_id": user["id"], "type": "advisor"},
        {"_id": 0}
    ).sort("created_at", 1).to_list(100)
    return history

@api_router.delete("/ai-advisor/history")
async def clear_chat_history(user=Depends(get_current_user)):
    """Clear chat history"""
    await db.chat_history.delete_many({"user_id": user["id"], "type": "advisor"})
    return {"message": "Chat history cleared"}

@api_router.post("/ai-advisor/calculate/{calc_type}")
async def run_calculator(calc_type: str, params: dict, user=Depends(get_current_user)):
    """Run a financial calculator"""
    try:
        if calc_type == "sip":
            return calculate_sip_returns(
                params.get("monthly_investment", 10000),
                params.get("annual_return", 12),
                params.get("years", 10)
            )
        elif calc_type == "emi":
            return calculate_loan_emi_details(
                params.get("principal", 5000000),
                params.get("annual_rate", 8.5),
                params.get("tenure_years", 20)
            )
        elif calc_type == "compound":
            return calculate_compound_interest(
                params.get("principal", 100000),
                params.get("annual_rate", 7),
                params.get("years", 5),
                params.get("compounding", "yearly")
            )
        elif calc_type == "portfolio":
            return calculate_portfolio_returns(
                params.get("investments", []),
                params.get("years", 10)
            )
        elif calc_type == "tax_80c":
            return calculate_tax_savings_80c(params.get("investments", {}))
        elif calc_type == "fire":
            return calculate_fire_number(
                params.get("monthly_expenses", 50000),
                params.get("withdrawal_rate", 4)
            )
        else:
            raise HTTPException(status_code=400, detail=f"Unknown calculator: {calc_type}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ══════════════════════════════════════
#  EXPORT ENDPOINTS (Excel & PDF)
# ══════════════════════════════════════

from fastapi.responses import StreamingResponse
import io

@api_router.get("/books/export/{report_type}/{format}")
async def export_report(
    report_type: str,  # ledger, pnl, balance
    format: str,  # excel, pdf
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user=Depends(get_current_user)
):
    """Export financial reports as Excel or PDF"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from fpdf import FPDF
    
    user_id = user["id"]
    user_name = user.get("full_name", "User")
    
    # Default to current FY if no dates provided
    if not start_date or not end_date:
        fy_start, fy_end = get_indian_fy_dates()
        start_date = start_date or fy_start.strftime("%Y-%m-%d")
        end_date = end_date or fy_end.strftime("%Y-%m-%d")
    
    if format == "excel":
        wb = Workbook()
        ws = wb.active
        
        # Styling
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        title_font = Font(bold=True, size=16)
        subtitle_font = Font(bold=True, size=11, color="666666")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        if report_type == "ledger":
            ws.title = "General Ledger"
            
            # Title
            ws.merge_cells('A1:G1')
            ws['A1'] = f"GENERAL LEDGER - {user_name}"
            ws['A1'].font = title_font
            ws['A2'] = f"Period: {start_date} to {end_date}"
            ws['A2'].font = subtitle_font
            ws['A3'] = f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}"
            ws['A3'].font = subtitle_font
            
            # Get ledger data
            query = {"user_id": user_id, "date": {"$gte": start_date, "$lte": end_date}}
            txns = await db.transactions.find(query, {"_id": 0}).sort("date", 1).to_list(2000)
            
            # Process transactions into ledger entries
            row = 5
            
            # Headers
            headers = ['Date', 'Particulars', 'Voucher Ref', 'Account', 'Debit (₹)', 'Credit (₹)', 'Balance (₹)']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            
            row += 1
            balance = 0
            
            for txn in txns:
                debit = txn['amount'] if txn['type'] == 'income' else 0
                credit = txn['amount'] if txn['type'] in ['expense', 'investment'] else 0
                balance += debit - credit
                
                ws.cell(row=row, column=1, value=txn['date']).border = border
                ws.cell(row=row, column=2, value=txn['description']).border = border
                ws.cell(row=row, column=3, value=f"TXN-{txn['id'][:8].upper()}").border = border
                ws.cell(row=row, column=4, value=txn['category']).border = border
                ws.cell(row=row, column=5, value=debit if debit > 0 else "").border = border
                ws.cell(row=row, column=6, value=credit if credit > 0 else "").border = border
                ws.cell(row=row, column=7, value=balance).border = border
                row += 1
            
            # Column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 14
            ws.column_dimensions['F'].width = 14
            ws.column_dimensions['G'].width = 14
            
        elif report_type == "pnl":
            ws.title = "Profit & Loss"
            
            # Title
            ws.merge_cells('A1:C1')
            ws['A1'] = f"INCOME & EXPENDITURE STATEMENT - {user_name}"
            ws['A1'].font = title_font
            ws['A2'] = f"Period: {start_date} to {end_date}"
            ws['A2'].font = subtitle_font
            
            # Get P&L data
            query = {"user_id": user_id, "date": {"$gte": start_date, "$lte": end_date}}
            txns = await db.transactions.find(query, {"_id": 0}).to_list(2000)
            
            total_income = sum(t["amount"] for t in txns if t["type"] == "income")
            total_expenses = sum(t["amount"] for t in txns if t["type"] == "expense")
            total_investments = sum(t["amount"] for t in txns if t["type"] == "investment")
            
            # Income by category
            income_cats = {}
            expense_cats = {}
            for t in txns:
                if t["type"] == "income":
                    income_cats[t["category"]] = income_cats.get(t["category"], 0) + t["amount"]
                elif t["type"] == "expense":
                    expense_cats[t["category"]] = expense_cats.get(t["category"], 0) + t["amount"]
            
            row = 4
            
            # Income Section
            ws.cell(row=row, column=1, value="INCOME").font = Font(bold=True, size=14, color="10B981")
            row += 1
            
            for cat, amt in sorted(income_cats.items(), key=lambda x: -x[1]):
                ws.cell(row=row, column=1, value=cat)
                ws.cell(row=row, column=3, value=amt)
                row += 1
            
            ws.cell(row=row, column=1, value="TOTAL INCOME").font = Font(bold=True)
            ws.cell(row=row, column=3, value=total_income).font = Font(bold=True, color="10B981")
            row += 2
            
            # Expenses Section
            ws.cell(row=row, column=1, value="EXPENDITURE").font = Font(bold=True, size=14, color="EF4444")
            row += 1
            
            for cat, amt in sorted(expense_cats.items(), key=lambda x: -x[1]):
                ws.cell(row=row, column=1, value=cat)
                ws.cell(row=row, column=3, value=amt)
                row += 1
            
            ws.cell(row=row, column=1, value="TOTAL EXPENDITURE").font = Font(bold=True)
            ws.cell(row=row, column=3, value=total_expenses).font = Font(bold=True, color="EF4444")
            row += 2
            
            # Surplus/Deficit
            surplus = total_income - total_expenses
            label = "SURPLUS" if surplus >= 0 else "DEFICIT"
            ws.cell(row=row, column=1, value=f"{label} FOR THE PERIOD").font = Font(bold=True, size=14)
            ws.cell(row=row, column=3, value=abs(surplus)).font = Font(bold=True, size=14, color="10B981" if surplus >= 0 else "EF4444")
            
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 18
            
        elif report_type == "balance":
            ws.title = "Balance Sheet"
            
            # Get balance sheet data
            query = {"user_id": user_id, "date": {"$lte": end_date}}
            txns = await db.transactions.find(query, {"_id": 0}).to_list(5000)
            assets = await db.fixed_assets.find({"user_id": user_id}, {"_id": 0}).to_list(100)
            loans = await db.loans.find({"user_id": user_id}, {"_id": 0}).to_list(100)
            
            total_income = sum(t["amount"] for t in txns if t["type"] == "income")
            total_expenses = sum(t["amount"] for t in txns if t["type"] == "expense")
            total_investments = sum(t["amount"] for t in txns if t["type"] == "investment")
            
            ws.merge_cells('A1:C1')
            ws['A1'] = f"STATEMENT OF FINANCIAL POSITION - {user_name}"
            ws['A1'].font = title_font
            ws['A2'] = f"As at: {end_date}"
            ws['A2'].font = subtitle_font
            
            row = 4
            
            # Assets
            ws.cell(row=row, column=1, value="I. ASSETS").font = Font(bold=True, size=14, color="3B82F6")
            row += 1
            
            # Fixed Assets
            total_fixed = sum(a.get("current_value", a["purchase_value"]) for a in assets)
            ws.cell(row=row, column=1, value="Fixed Assets")
            ws.cell(row=row, column=3, value=total_fixed)
            row += 1
            
            # Investments
            ws.cell(row=row, column=1, value="Investments")
            ws.cell(row=row, column=3, value=total_investments)
            row += 1
            
            # Cash & Bank
            cash_balance = total_income - total_expenses - total_investments
            ws.cell(row=row, column=1, value="Cash & Bank Balance")
            ws.cell(row=row, column=3, value=max(0, cash_balance))
            row += 1
            
            total_assets = total_fixed + total_investments + max(0, cash_balance)
            ws.cell(row=row, column=1, value="TOTAL ASSETS").font = Font(bold=True)
            ws.cell(row=row, column=3, value=total_assets).font = Font(bold=True, color="3B82F6")
            row += 2
            
            # Liabilities
            ws.cell(row=row, column=1, value="II. LIABILITIES").font = Font(bold=True, size=14, color="EF4444")
            row += 1
            
            total_liabilities = 0
            for loan in loans:
                ws.cell(row=row, column=1, value=loan["name"])
                ws.cell(row=row, column=3, value=loan.get("outstanding_principal", loan["principal_amount"]))
                total_liabilities += loan.get("outstanding_principal", loan["principal_amount"])
                row += 1
            
            ws.cell(row=row, column=1, value="TOTAL LIABILITIES").font = Font(bold=True)
            ws.cell(row=row, column=3, value=total_liabilities).font = Font(bold=True, color="EF4444")
            row += 2
            
            # Net Worth
            net_worth = total_assets - total_liabilities
            ws.cell(row=row, column=1, value="III. NET WORTH").font = Font(bold=True, size=14, color="10B981")
            row += 1
            ws.cell(row=row, column=1, value="Closing Net Worth").font = Font(bold=True)
            ws.cell(row=row, column=3, value=net_worth).font = Font(bold=True, color="10B981")
            
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 18
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"Visor_{report_type.title()}_{end_date}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    elif format == "pdf":
        # Create PDF using fpdf2
        pdf = FPDF()
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)
        
        # Header
        pdf.set_font("Helvetica", "B", 20)
        pdf.set_text_color(59, 130, 246)  # Blue
        
        if report_type == "ledger":
            pdf.cell(0, 10, "GENERAL LEDGER", ln=True, align="C")
        elif report_type == "pnl":
            pdf.cell(0, 10, "INCOME & EXPENDITURE STATEMENT", ln=True, align="C")
        else:
            pdf.cell(0, 10, "STATEMENT OF FINANCIAL POSITION", ln=True, align="C")
        
        pdf.set_font("Helvetica", "", 12)
        pdf.set_text_color(100, 100, 100)
        pdf.cell(0, 8, f"User: {user_name}", ln=True, align="C")
        pdf.cell(0, 8, f"Period: {start_date} to {end_date}", ln=True, align="C")
        pdf.cell(0, 8, f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}", ln=True, align="C")
        pdf.ln(10)
        
        # Content
        query = {"user_id": user_id, "date": {"$gte": start_date, "$lte": end_date}}
        txns = await db.transactions.find(query, {"_id": 0}).sort("date", 1).to_list(2000)
        
        if report_type == "ledger":
            # Table header
            pdf.set_fill_color(59, 130, 246)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Helvetica", "B", 10)
            
            pdf.cell(25, 8, "Date", border=1, fill=True)
            pdf.cell(60, 8, "Particulars", border=1, fill=True)
            pdf.cell(30, 8, "Category", border=1, fill=True)
            pdf.cell(25, 8, "Debit", border=1, fill=True, align="R")
            pdf.cell(25, 8, "Credit", border=1, fill=True, align="R")
            pdf.cell(25, 8, "Balance", border=1, fill=True, align="R")
            pdf.ln()
            
            pdf.set_text_color(0, 0, 0)
            pdf.set_font("Helvetica", "", 9)
            
            balance = 0
            for txn in txns[:50]:  # Limit for PDF
                debit = txn['amount'] if txn['type'] == 'income' else 0
                credit = txn['amount'] if txn['type'] in ['expense', 'investment'] else 0
                balance += debit - credit
                
                pdf.cell(25, 7, txn['date'][:10], border=1)
                pdf.cell(60, 7, txn['description'][:25], border=1)
                pdf.cell(30, 7, txn['category'][:15], border=1)
                pdf.cell(25, 7, f"{debit:,.0f}" if debit > 0 else "-", border=1, align="R")
                pdf.cell(25, 7, f"{credit:,.0f}" if credit > 0 else "-", border=1, align="R")
                pdf.cell(25, 7, f"{balance:,.0f}", border=1, align="R")
                pdf.ln()
        
        elif report_type == "pnl":
            total_income = sum(t["amount"] for t in txns if t["type"] == "income")
            total_expenses = sum(t["amount"] for t in txns if t["type"] == "expense")
            
            # Income by category
            income_cats = {}
            expense_cats = {}
            for t in txns:
                if t["type"] == "income":
                    income_cats[t["category"]] = income_cats.get(t["category"], 0) + t["amount"]
                elif t["type"] == "expense":
                    expense_cats[t["category"]] = expense_cats.get(t["category"], 0) + t["amount"]
            
            # Income Section
            pdf.set_font("Helvetica", "B", 14)
            pdf.set_text_color(16, 185, 129)  # Green
            pdf.cell(0, 10, "INCOME", ln=True)
            
            pdf.set_font("Helvetica", "", 11)
            pdf.set_text_color(0, 0, 0)
            
            for cat, amt in sorted(income_cats.items(), key=lambda x: -x[1]):
                pdf.cell(100, 8, cat)
                pdf.cell(50, 8, f"Rs. {amt:,.2f}", align="R")
                pdf.ln()
            
            pdf.set_font("Helvetica", "B", 12)
            pdf.set_text_color(16, 185, 129)
            pdf.cell(100, 10, "TOTAL INCOME")
            pdf.cell(50, 10, f"Rs. {total_income:,.2f}", align="R")
            pdf.ln(15)
            
            # Expenses Section
            pdf.set_font("Helvetica", "B", 14)
            pdf.set_text_color(239, 68, 68)  # Red
            pdf.cell(0, 10, "EXPENDITURE", ln=True)
            
            pdf.set_font("Helvetica", "", 11)
            pdf.set_text_color(0, 0, 0)
            
            for cat, amt in sorted(expense_cats.items(), key=lambda x: -x[1]):
                pdf.cell(100, 8, cat)
                pdf.cell(50, 8, f"Rs. {amt:,.2f}", align="R")
                pdf.ln()
            
            pdf.set_font("Helvetica", "B", 12)
            pdf.set_text_color(239, 68, 68)
            pdf.cell(100, 10, "TOTAL EXPENDITURE")
            pdf.cell(50, 10, f"Rs. {total_expenses:,.2f}", align="R")
            pdf.ln(15)
            
            # Surplus/Deficit
            surplus = total_income - total_expenses
            pdf.set_font("Helvetica", "B", 16)
            if surplus >= 0:
                pdf.set_text_color(16, 185, 129)
                pdf.cell(100, 12, "SURPLUS FOR THE PERIOD")
            else:
                pdf.set_text_color(239, 68, 68)
                pdf.cell(100, 12, "DEFICIT FOR THE PERIOD")
            pdf.cell(50, 12, f"Rs. {abs(surplus):,.2f}", align="R")
        
        elif report_type == "balance":
            query = {"user_id": user_id, "date": {"$lte": end_date}}
            txns = await db.transactions.find(query, {"_id": 0}).to_list(5000)
            assets = await db.fixed_assets.find({"user_id": user_id}, {"_id": 0}).to_list(100)
            loans = await db.loans.find({"user_id": user_id}, {"_id": 0}).to_list(100)
            
            total_income = sum(t["amount"] for t in txns if t["type"] == "income")
            total_expenses = sum(t["amount"] for t in txns if t["type"] == "expense")
            total_investments = sum(t["amount"] for t in txns if t["type"] == "investment")
            cash_balance = total_income - total_expenses - total_investments
            
            # Assets
            pdf.set_font("Helvetica", "B", 14)
            pdf.set_text_color(59, 130, 246)
            pdf.cell(0, 10, "I. ASSETS", ln=True)
            
            pdf.set_font("Helvetica", "", 11)
            pdf.set_text_color(0, 0, 0)
            
            total_fixed = sum(a.get("current_value", a["purchase_value"]) for a in assets)
            
            pdf.cell(100, 8, "Fixed Assets")
            pdf.cell(50, 8, f"Rs. {total_fixed:,.2f}", align="R")
            pdf.ln()
            
            pdf.cell(100, 8, "Investments")
            pdf.cell(50, 8, f"Rs. {total_investments:,.2f}", align="R")
            pdf.ln()
            
            pdf.cell(100, 8, "Cash & Bank Balance")
            pdf.cell(50, 8, f"Rs. {max(0, cash_balance):,.2f}", align="R")
            pdf.ln()
            
            total_assets = total_fixed + total_investments + max(0, cash_balance)
            pdf.set_font("Helvetica", "B", 12)
            pdf.set_text_color(59, 130, 246)
            pdf.cell(100, 10, "TOTAL ASSETS")
            pdf.cell(50, 10, f"Rs. {total_assets:,.2f}", align="R")
            pdf.ln(15)
            
            # Liabilities
            pdf.set_font("Helvetica", "B", 14)
            pdf.set_text_color(239, 68, 68)
            pdf.cell(0, 10, "II. LIABILITIES", ln=True)
            
            pdf.set_font("Helvetica", "", 11)
            pdf.set_text_color(0, 0, 0)
            
            total_liabilities = 0
            for loan in loans:
                outstanding = loan.get("outstanding_principal", loan["principal_amount"])
                pdf.cell(100, 8, loan["name"])
                pdf.cell(50, 8, f"Rs. {outstanding:,.2f}", align="R")
                pdf.ln()
                total_liabilities += outstanding
            
            if not loans:
                pdf.cell(100, 8, "No liabilities")
                pdf.ln()
            
            pdf.set_font("Helvetica", "B", 12)
            pdf.set_text_color(239, 68, 68)
            pdf.cell(100, 10, "TOTAL LIABILITIES")
            pdf.cell(50, 10, f"Rs. {total_liabilities:,.2f}", align="R")
            pdf.ln(15)
            
            # Net Worth
            net_worth = total_assets - total_liabilities
            pdf.set_font("Helvetica", "B", 14)
            pdf.set_text_color(16, 185, 129)
            pdf.cell(0, 10, "III. NET WORTH", ln=True)
            
            pdf.set_font("Helvetica", "B", 16)
            pdf.cell(100, 12, "Closing Net Worth")
            pdf.cell(50, 12, f"Rs. {net_worth:,.2f}", align="R")
        
        # Footer
        pdf.ln(20)
        pdf.set_font("Helvetica", "I", 9)
        pdf.set_text_color(150, 150, 150)
        pdf.cell(0, 8, "Generated by Visor Finance App", ln=True, align="C")
        
        # Save to buffer
        buffer = io.BytesIO()
        pdf.output(buffer)
        buffer.seek(0)
        
        filename = f"Visor_{report_type.title()}_{end_date}.pdf"
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    else:
        raise HTTPException(status_code=400, detail="Invalid format. Use 'excel' or 'pdf'")

# ══════════════════════════════════════
#  GMAIL INTEGRATION - Email Transaction Parsing
# ══════════════════════════════════════

import warnings
from google_auth_oauthlib.flow import Flow
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request as GoogleRequest
from googleapiclient.discovery import build
import base64
from bank_parser import parse_transaction_text

GOOGLE_CLIENT_ID = os.environ.get("GOOGLE_CLIENT_ID", "")
GOOGLE_CLIENT_SECRET = os.environ.get("GOOGLE_CLIENT_SECRET", "")
GMAIL_REDIRECT_URI = os.environ.get("GMAIL_REDIRECT_URI", "")
GMAIL_SCOPES = [
    "https://www.googleapis.com/auth/gmail.readonly",
    "openid",
    "https://www.googleapis.com/auth/userinfo.email",
    "https://www.googleapis.com/auth/userinfo.profile",
]

# Auto-detect redirect URI from frontend URL
if not GMAIL_REDIRECT_URI:
    _fe_url = os.environ.get("FRONTEND_URL", "https://invest-live.preview.emergentagent.com")
    GMAIL_REDIRECT_URI = f"{_fe_url}/api/gmail/callback"

def _gmail_client_config():
    return {
        "web": {
            "client_id": GOOGLE_CLIENT_ID,
            "client_secret": GOOGLE_CLIENT_SECRET,
            "auth_uri": "https://accounts.google.com/o/oauth2/auth",
            "token_uri": "https://oauth2.googleapis.com/token",
        }
    }

@api_router.get("/gmail/connect")
async def gmail_connect(user=Depends(get_current_user)):
    """Initiate Gmail OAuth flow."""
    if not GOOGLE_CLIENT_ID or not GOOGLE_CLIENT_SECRET:
        raise HTTPException(status_code=500, detail="Gmail OAuth not configured")

    flow = Flow.from_client_config(_gmail_client_config(), scopes=GMAIL_SCOPES, redirect_uri=GMAIL_REDIRECT_URI)
    url, state = flow.authorization_url(access_type="offline", prompt="consent")

    # Store state → user mapping
    await db.gmail_oauth_states.insert_one({
        "state": state,
        "user_id": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
    })

    return {"auth_url": url}

@api_router.get("/gmail/callback")
async def gmail_callback(code: str = None, state: str = None, error: str = None):
    """Handle Gmail OAuth callback."""
    if error:
        return RedirectResponse(url="/?gmail_error=" + error)

    if not code or not state:
        raise HTTPException(status_code=400, detail="Missing code or state")

    # Verify state
    state_doc = await db.gmail_oauth_states.find_one({"state": state}, {"_id": 0})
    if not state_doc:
        raise HTTPException(status_code=400, detail="Invalid OAuth state")

    user_id = state_doc["user_id"]
    await db.gmail_oauth_states.delete_one({"state": state})

    flow = Flow.from_client_config(_gmail_client_config(), scopes=GMAIL_SCOPES, redirect_uri=GMAIL_REDIRECT_URI)

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        flow.fetch_token(code=code)

    creds = flow.credentials

    # Store tokens
    token_doc = {
        "user_id": user_id,
        "access_token": creds.token,
        "refresh_token": creds.refresh_token,
        "token_uri": creds.token_uri,
        "client_id": creds.client_id,
        "client_secret": creds.client_secret,
        "expires_at": creds.expiry.isoformat() if creds.expiry else None,
        "connected_at": datetime.now(timezone.utc).isoformat(),
    }

    await db.gmail_tokens.update_one(
        {"user_id": user_id}, {"$set": token_doc}, upsert=True
    )

    # Redirect back to app settings
    return RedirectResponse(url="/?gmail_connected=true")

@api_router.get("/gmail/status")
async def gmail_status(user=Depends(get_current_user)):
    """Check if Gmail is connected for the current user."""
    token = await db.gmail_tokens.find_one({"user_id": user["id"]}, {"_id": 0, "access_token": 0, "refresh_token": 0, "client_secret": 0})
    if not token:
        return {"connected": False}

    last_sync = await db.gmail_sync_log.find_one(
        {"user_id": user["id"]}, {"_id": 0}, sort=[("synced_at", -1)]
    )

    return {
        "connected": True,
        "connected_at": token.get("connected_at"),
        "last_sync": last_sync.get("synced_at") if last_sync else None,
        "total_parsed": last_sync.get("total_parsed", 0) if last_sync else 0,
    }

async def _get_gmail_creds(user_id: str) -> Credentials:
    """Get valid Gmail credentials, refreshing if needed."""
    token = await db.gmail_tokens.find_one({"user_id": user_id}, {"_id": 0})
    if not token:
        raise HTTPException(status_code=400, detail="Gmail not connected")

    creds = Credentials(
        token=token["access_token"],
        refresh_token=token.get("refresh_token"),
        token_uri=token["token_uri"],
        client_id=token["client_id"],
        client_secret=token["client_secret"],
    )

    # Refresh if expired
    if token.get("expires_at"):
        from dateutil.parser import parse as parse_date
        expires = parse_date(token["expires_at"])
        if expires.tzinfo is None:
            expires = expires.replace(tzinfo=timezone.utc)
        if datetime.now(timezone.utc) >= expires:
            creds.refresh(GoogleRequest())
            await db.gmail_tokens.update_one(
                {"user_id": user_id},
                {"$set": {"access_token": creds.token, "expires_at": creds.expiry.isoformat() if creds.expiry else None}},
            )

    return creds

@api_router.post("/gmail/sync")
async def gmail_sync(user=Depends(get_current_user)):
    """Fetch and parse bank transaction emails from Gmail."""
    creds = await _get_gmail_creds(user["id"])
    service = build("gmail", "v1", credentials=creds)

    # Search for bank transaction emails
    bank_queries = [
        "from:(alerts@hdfcbank.net OR alert@icicibank.com OR sbi OR axisbank OR kotak) subject:(transaction OR debit OR credit OR spent)",
        "subject:(bank transaction alert) newer_than:30d",
        "(debit OR credit) (INR OR Rs) newer_than:30d",
    ]

    parsed_transactions = []
    seen_msg_ids = set()

    # Check existing synced message IDs to avoid duplicates
    existing = await db.gmail_synced_msgs.find(
        {"user_id": user["id"]}, {"msg_id": 1, "_id": 0}
    ).to_list(10000)
    seen_msg_ids = {d["msg_id"] for d in existing}

    for query in bank_queries:
        try:
            result = service.users().messages().list(
                userId="me", q=query, maxResults=50
            ).execute()
            messages = result.get("messages", [])

            for msg_meta in messages:
                msg_id = msg_meta["id"]
                if msg_id in seen_msg_ids:
                    continue
                seen_msg_ids.add(msg_id)

                try:
                    msg = service.users().messages().get(
                        userId="me", id=msg_id, format="full"
                    ).execute()

                    # Extract email body
                    body_text = _extract_email_body(msg)
                    sender = _get_header(msg, "From")

                    if body_text:
                        txn = parse_transaction_text(body_text, sender)
                        if txn:
                            txn["gmail_msg_id"] = msg_id
                            txn["email_subject"] = _get_header(msg, "Subject")
                            parsed_transactions.append(txn)

                except Exception as e:
                    logger.warning(f"Failed to parse email {msg_id}: {e}")
                    continue

        except Exception as e:
            logger.warning(f"Gmail search failed for query: {e}")
            continue

    # Save parsed transactions
    new_count = 0
    for txn in parsed_transactions:
        txn_id = str(uuid.uuid4())
        txn_doc = {
            "id": txn_id,
            "user_id": user["id"],
            "type": txn["type"],
            "amount": txn["amount"],
            "category": txn["category"],
            "description": txn["description"],
            "date": txn["date"],
            "is_recurring": False,
            "source": "gmail",
            "bank": txn.get("bank", ""),
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.transactions.insert_one(txn_doc)
        await db.gmail_synced_msgs.insert_one({
            "user_id": user["id"],
            "msg_id": txn["gmail_msg_id"],
            "txn_id": txn_id,
            "synced_at": datetime.now(timezone.utc).isoformat(),
        })
        new_count += 1

    # Log sync
    await db.gmail_sync_log.insert_one({
        "user_id": user["id"],
        "synced_at": datetime.now(timezone.utc).isoformat(),
        "emails_scanned": len(seen_msg_ids),
        "total_parsed": new_count,
    })

    return {
        "success": True,
        "new_transactions": new_count,
        "emails_scanned": len(seen_msg_ids),
    }

@api_router.delete("/gmail/disconnect")
async def gmail_disconnect(user=Depends(get_current_user)):
    """Remove Gmail connection."""
    await db.gmail_tokens.delete_many({"user_id": user["id"]})
    await db.gmail_oauth_states.delete_many({"user_id": user["id"]})
    return {"success": True}

def _extract_email_body(msg: dict) -> str:
    """Extract plain text body from Gmail message."""
    payload = msg.get("payload", {})

    # Direct body
    if payload.get("body", {}).get("data"):
        return base64.urlsafe_b64decode(payload["body"]["data"]).decode("utf-8", errors="ignore")

    # Multipart
    parts = payload.get("parts", [])
    for part in parts:
        if part.get("mimeType") == "text/plain" and part.get("body", {}).get("data"):
            return base64.urlsafe_b64decode(part["body"]["data"]).decode("utf-8", errors="ignore")

    # Fallback: HTML
    for part in parts:
        if part.get("mimeType") == "text/html" and part.get("body", {}).get("data"):
            html = base64.urlsafe_b64decode(part["body"]["data"]).decode("utf-8", errors="ignore")
            import re as _re
            return _re.sub(r"<[^>]+>", " ", html)

    return msg.get("snippet", "")

def _get_header(msg: dict, name: str) -> str:
    """Get email header value."""
    headers = msg.get("payload", {}).get("headers", [])
    for h in headers:
        if h["name"].lower() == name.lower():
            return h["value"]
    return ""


# ══════════════════════════════════════
#  SMS PARSING (Android Only)
# ══════════════════════════════════════

@api_router.post("/sms/parse")
async def parse_sms_batch(messages: list[dict], user=Depends(get_current_user)):
    """Parse a batch of SMS messages sent from Android client."""
    parsed = []
    for sms in messages:
        body = sms.get("body", "")
        sender = sms.get("sender", "")
        txn = parse_transaction_text(body, sender)
        if txn:
            txn_id = str(uuid.uuid4())
            txn_doc = {
                "id": txn_id,
                "user_id": user["id"],
                "type": txn["type"],
                "amount": txn["amount"],
                "category": txn["category"],
                "description": txn["description"],
                "date": txn["date"],
                "is_recurring": False,
                "source": "sms",
                "bank": txn.get("bank", ""),
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
            await db.transactions.insert_one(txn_doc)
            parsed.append({"id": txn_id, "description": txn["description"], "amount": txn["amount"], "type": txn["type"]})

    return {"success": True, "parsed_count": len(parsed), "transactions": parsed}


# ══════════════════════════════════════
#  APP SETUP
# ══════════════════════════════════════

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
async def startup():
    await seed_demo_data()
    await seed_market_data()
    asyncio.create_task(market_data_scheduler())

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
