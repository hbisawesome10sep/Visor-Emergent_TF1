"""
Statement Parser — Intelligent XLSX parser for Groww & Zerodha
Handles: Stock holdings, Mutual Fund holdings from various export formats.
"""
import re
import logging
from datetime import datetime, timezone
from typing import Optional
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# ── Column pattern matching ──────────────────────────────────────────────────
_COL_PATTERNS = {
    "name": [
        r"(?:company|stock|scrip|instrument|fund|scheme|security|isin)\s*name",
        r"^name$", r"^instrument$", r"^scrip$", r"^company$", r"^scheme$",
        r"^stock$", r"^security$", r"^fund$",
    ],
    "ticker": [
        r"(?:nse|bse)?\s*(?:symbol|ticker|code|scrip\s*code)",
        r"^symbol$", r"^trading\s*symbol$", r"^nse\s*symbol$",
    ],
    "isin": [r"^isin", r"isin\s*(?:no|number|code)?"],
    "quantity": [
        r"^(?:qty|quantity|units|unit|balance|holding\s*qty|no\.?\s*of\s*shares)",
        r"quantity\s*(?:available|held|total)",
        r"(?:current|free)\s*balance", r"^shares$", r"^units$",
    ],
    "buy_price": [
        r"(?:avg|average|weighted\s*avg)\.?\s*(?:cost|price|nav|buy\s*price|rate)",
        r"^buy\s*(?:avg|price|rate)$", r"^avg\.?\s*cost$", r"^purchase\s*price$",
        r"^cost\s*price$", r"^avg\.?\s*nav$", r"^invested\s*nav$",
        r"^average\s*price$",
    ],
    "current_price": [
        r"(?:ltp|cmp|current|present|last|mkt|market)[\s_]*(?:price|nav|rate)",
        r"(?:previous|prev)[\s_]*(?:closing|close)[\s_]*(?:price)?",
        r"^closing\s*price$",
        r"^ltp$", r"^cmp$", r"^nav$", r"^current\s*nav$", r"^closing\s*price$",
    ],
    "invested_value": [
        r"(?:invested?|cost|buy|purchase)\s*(?:value|amount|total)",
        r"^invested$", r"^cost\s*value$", r"^investment$", r"^invested\s*value$",
        r"^total\s*invest", r"^amount\s*invested$",
    ],
    "current_value": [
        r"(?:current|present|market|mkt|portfolio)\s*(?:value|val|amount)",
        r"^cur\.?\s*val\.?$", r"^present\s*value$",
    ],
    "pnl": [
        r"(?:p\s*&?\s*l|profit|gain|unrealised|unrealized)",
        r"^pnl$", r"^returns?$",
    ],
    "xirr": [r"^xirr$", r"^xirr\s*%?$"],
    "buy_date": [
        r"(?:buy|purchase|trade|transaction)\s*date",
        r"^date$",
    ],
    "folio": [r"folio\s*(?:no\.?|number)?", r"^folio$"],
    "amc": [r"^amc$", r"^fund\s*house$", r"^asset\s*management"],
    "sub_category": [r"^sub[\s\-_]?category$", r"^sub[\s\-_]?type$"],
    "category": [r"^category$", r"^type$", r"^asset\s*class$"],
    "source": [r"^source$", r"^platform$"],
    "exchange": [r"^exchange$", r"^exch$"],
    "sector": [r"^sector$", r"^industry$", r"^segment$"],
}


def _match_col(header: str) -> Optional[str]:
    """Match a column header to a known field name."""
    h = header.strip().lower()
    for field, patterns in _COL_PATTERNS.items():
        for pat in patterns:
            if re.search(pat, h, re.IGNORECASE):
                return field
    return None


def _detect_source(headers: list[str], sheet_name: str, all_rows: list) -> str:
    """Detect if the statement is from Groww, Zerodha, or unknown."""
    all_text = " ".join(headers).lower() + " " + sheet_name.lower()
    # Also scan first 30 rows for source clues
    for row in all_rows[:30]:
        for cell in (row or []):
            if cell:
                t = str(cell).lower()
                if "groww" in t:
                    return "Groww"
                if "zerodha" in t or "console" in t or "kite" in t:
                    return "Zerodha"
    if "groww" in all_text:
        return "Groww"
    if "zerodha" in all_text or "console" in all_text:
        return "Zerodha"
    return "Unknown"


def _detect_category(name: str, ticker: str, isin: str, cat_cell: str, statement_type: str) -> str:
    """Detect if a holding is Stock or Mutual Fund."""
    if statement_type == "mf_statement":
        return "Mutual Fund"
    if statement_type == "stock_statement":
        return "Stock"
    # Auto-detect from data
    name_lower = name.lower()
    cat_lower = cat_cell.lower() if cat_cell else ""
    if cat_lower in ("equity", "debt", "hybrid", "solution oriented", "other"):
        return "Mutual Fund"
    if any(kw in name_lower for kw in ["fund", "scheme", "growth", "direct", "regular", "flexi", "cap fund", "debt fund", "hybrid", "index fund"]):
        return "Mutual Fund"
    if isin and isin.startswith("INF"):
        return "Mutual Fund"
    return "Stock"


def _to_float(val) -> float:
    """Safely convert a cell value to float."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "").replace("₹", "").replace("Rs", "").replace("INR", "").replace("%", "").strip()
    s = re.sub(r"[^\d.\-]", "", s)
    try:
        return float(s) if s else 0.0
    except ValueError:
        return 0.0


def _to_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _extract_summary(rows: list) -> dict:
    """Extract portfolio summary from Groww-style header section."""
    summary = {}
    for i, row in enumerate(rows[:20]):
        if not row:
            continue
        first_cell = _to_str(row[0]).lower()
        if "holding summary" in first_cell or "portfolio summary" in first_cell:
            # Next non-empty row should have headers, then values
            for j in range(i + 1, min(i + 5, len(rows))):
                hdr_row = rows[j]
                if not hdr_row or not hdr_row[0]:
                    continue
                hdr_text = _to_str(hdr_row[0]).lower()
                if "total" in hdr_text or "invest" in hdr_text:
                    # This is the header row
                    headers = [_to_str(c).lower() for c in hdr_row]
                    # Next row should have values
                    if j + 1 < len(rows):
                        vals = rows[j + 1]
                        for k, h in enumerate(headers):
                            if k < len(vals):
                                if "total" in h and "invest" in h:
                                    summary["total_invested"] = _to_float(vals[k])
                                elif "current" in h or "portfolio" in h:
                                    summary["total_current_value"] = _to_float(vals[k])
                                elif "xirr" in h:
                                    summary["xirr"] = _to_float(vals[k])
                                elif "profit" in h and "%" not in h:
                                    summary["pnl"] = _to_float(vals[k])
                    break
    return summary


def _extract_personal_details(rows: list) -> dict:
    """Extract personal details from Groww-style header section."""
    details = {}
    for row in rows[:10]:
        if not row:
            continue
        key = _to_str(row[0]).lower()
        val = _to_str(row[1]) if len(row) > 1 and row[1] else ""
        if key == "name":
            details["name"] = val
        elif "pan" in key:
            details["pan"] = val
        elif "mobile" in key:
            details["mobile"] = val
    return details


def _build_ticker(name: str, ticker: str, isin: str, category: str) -> str:
    """Build a Yahoo Finance compatible ticker."""
    if ticker:
        t = ticker.strip().upper().replace(" ", "")
        if not t.endswith(".NS") and not t.endswith(".BO") and category == "Stock":
            t = f"{t}.NS"
        return t
    if category == "Mutual Fund" and isin:
        return isin
    return ""


def parse_holdings_xlsx(file_bytes: bytes, statement_type: str = "auto", filename: str = "") -> dict:
    """
    Parse an XLSX file and extract holdings.
    
    Args:
        file_bytes: Raw bytes of the XLSX file
        statement_type: 'stock_statement', 'mf_statement', or 'auto'
        filename: Original filename for source detection
    
    Returns:
        dict with 'holdings' list, 'metadata', and 'summary'
    """
    import io
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    
    all_holdings = []
    metadata = {
        "source": "Unknown",
        "sheets_parsed": [],
        "total_rows_scanned": 0,
        "parse_errors": [],
    }
    portfolio_summary = {}
    personal_details = {}
    
    parsed_sheet_types = set()
    
    for sheet_name in wb.sheetnames:
        # Skip "Combined" sheet if Equity/MF sheets already parsed (avoids duplicates)
        sheet_lower = sheet_name.lower()
        if "combined" in sheet_lower and parsed_sheet_types:
            continue
        
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        
        # Extract summary and personal details
        summary = _extract_summary(rows)
        if summary:
            portfolio_summary.update(summary)
        pdetails = _extract_personal_details(rows)
        if pdetails:
            personal_details.update(pdetails)
        
        # Find header row — scan up to 35 rows (Groww puts headers at row ~20)
        header_row_idx = None
        col_map = {}
        
        for i, row in enumerate(rows[:35]):
            if not row:
                continue
            headers = [_to_str(c) for c in row]
            temp_map = {}
            for j, h in enumerate(headers):
                if not h:
                    continue
                field = _match_col(h)
                if field and field not in temp_map:
                    temp_map[field] = j
            
            # Need at least one identifier (name/isin/ticker) + one value
            has_identifier = "name" in temp_map or "isin" in temp_map or "ticker" in temp_map
            has_value = "quantity" in temp_map or "invested_value" in temp_map or "current_value" in temp_map
            if has_identifier and has_value:
                header_row_idx = i
                col_map = temp_map
                source = _detect_source(headers, sheet_name, rows)
                if source != "Unknown":
                    metadata["source"] = source
                break
        
        if header_row_idx is None:
            continue
        
        metadata["sheets_parsed"].append(sheet_name)
        if "equity" in sheet_lower:
            parsed_sheet_types.add("equity")
        elif "mutual" in sheet_lower:
            parsed_sheet_types.add("mf")
        logger.info(f"Sheet '{sheet_name}': header at row {header_row_idx}, columns mapped: {list(col_map.keys())}")
        
        # Parse data rows
        for row in rows[header_row_idx + 1:]:
            if not row:
                continue
            metadata["total_rows_scanned"] += 1
            
            def get(field):
                idx = col_map.get(field)
                return row[idx] if idx is not None and idx < len(row) else None
            
            name = _to_str(get("name"))
            isin = _to_str(get("isin"))
            ticker = _to_str(get("ticker"))
            quantity = _to_float(get("quantity"))
            buy_price = _to_float(get("buy_price"))
            current_price = _to_float(get("current_price"))
            invested_value = _to_float(get("invested_value"))
            current_value = _to_float(get("current_value"))
            pnl = _to_float(get("pnl"))
            xirr_val = _to_float(get("xirr"))
            folio = _to_str(get("folio"))
            amc = _to_str(get("amc"))
            cat_cell = _to_str(get("category"))
            sub_cat = _to_str(get("sub_category"))
            source_cell = _to_str(get("source"))
            sector = _to_str(get("sector"))
            buy_date_raw = get("buy_date")
            
            # Use ticker/symbol as name when name column is missing (Zerodha format)
            if not name and ticker:
                name = ticker
            
            # Skip empty/invalid rows
            if not name and not isin:
                continue
            if quantity <= 0 and invested_value <= 0:
                continue
            
            # Determine category
            category = _detect_category(name, ticker, isin, cat_cell, statement_type)
            
            # Build ticker
            final_ticker = _build_ticker(name, ticker, isin, category)
            
            # Calculate missing values
            if invested_value > 0 and quantity > 0 and buy_price <= 0:
                buy_price = round(invested_value / quantity, 4)
            if invested_value <= 0 and quantity > 0 and buy_price > 0:
                invested_value = round(quantity * buy_price, 2)
            if current_value > 0 and quantity > 0 and current_price <= 0:
                current_price = round(current_value / quantity, 4)
            if current_value <= 0 and quantity > 0 and current_price > 0:
                current_value = round(quantity * current_price, 2)
            
            # Calculate gain/loss
            gain_loss = round(current_value - invested_value, 2) if current_value > 0 and invested_value > 0 else pnl
            gain_loss_pct = round((gain_loss / invested_value) * 100, 2) if invested_value > 0 and gain_loss else 0
            
            # Handle buy date
            buy_date_str = ""
            if buy_date_raw:
                if isinstance(buy_date_raw, datetime):
                    buy_date_str = buy_date_raw.strftime("%Y-%m-%d")
                else:
                    for fmt in ["%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d-%b-%Y", "%d %b %Y"]:
                        try:
                            buy_date_str = datetime.strptime(str(buy_date_raw).strip(), fmt).strftime("%Y-%m-%d")
                            break
                        except ValueError:
                            continue
            
            holding = {
                "name": name or isin,
                "ticker": final_ticker,
                "isin": isin,
                "category": category,
                "sub_category": sub_cat,
                "sector": sector,
                "amc": amc,
                "folio": folio,
                "source_platform": source_cell,
                "quantity": round(quantity, 4),
                "buy_price": round(buy_price, 4),
                "buy_date": buy_date_str,
                "invested_value": round(invested_value, 2),
                "current_price": round(current_price, 4) if current_price > 0 else None,
                "current_value": round(current_value, 2) if current_value > 0 else None,
                "gain_loss": gain_loss,
                "gain_loss_pct": gain_loss_pct,
                "xirr": xirr_val if xirr_val != 0 else None,
            }
            all_holdings.append(holding)
    
    wb.close()
    
    # Detect source from filename if not found in sheet
    if metadata["source"] == "Unknown":
        fn = filename.lower()
        if "groww" in fn:
            metadata["source"] = "Groww"
        elif "zerodha" in fn or "console" in fn or "kite" in fn:
            metadata["source"] = "Zerodha"
    
    # Detect source from holdings source_platform field
    if metadata["source"] == "Unknown":
        sources = set(h.get("source_platform", "").lower() for h in all_holdings)
        if "groww" in sources:
            metadata["source"] = "Groww"
    
    metadata["holdings_found"] = len(all_holdings)
    
    logger.info(f"Parsed {filename}: source={metadata['source']}, "
                f"sheets={metadata['sheets_parsed']}, "
                f"holdings={len(all_holdings)}, "
                f"rows_scanned={metadata['total_rows_scanned']}")
    
    return {
        "holdings": all_holdings,
        "metadata": metadata,
        "summary": portfolio_summary,
        "personal_details": personal_details,
    }
